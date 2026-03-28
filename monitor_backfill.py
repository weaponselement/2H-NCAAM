import time
import json
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

print("Starting 15-min progress monitor for backfill...")
print("="*70)

poll_num = 0
while True:
    poll_num += 1
    now = datetime.now().strftime("%H:%M:%S")
    
    # Check workbook
    try:
        wb = load_workbook('logs/NCAAM Results.xlsx', data_only=True)
        ws = wb['Game_Log']
        rows = list(ws.iter_rows(values_only=True))
        h = [str(c) if c is not None else '' for c in rows[0]]
        i = {k: v for v, k in enumerate(h)}
        data = [r for r in rows[1:] if any(v is not None for v in r)]
        d = [str(r[i['Date']]).split(' ')[0] for r in data if i.get('Date') is not None and r[i['Date']] not in (None, '')]
        nov_jan = [k for k in d if '2025-11-01' <= k <= '2026-01-31']
        wb_total = len(data)
        wb_nov_jan = len(nov_jan)
    except Exception as e:
        wb_total = 0
        wb_nov_jan = 0
    
    # Check baselines
    baseline_files = list(Path('data/processed/baselines').glob('last4_2025-11*.json'))
    baseline_files.extend(Path('data/processed/baselines').glob('last4_2025-12*.json'))
    baseline_files.extend(Path('data/processed/baselines').glob('last4_2026-01*.json'))
    baseline_count = len(baseline_files)
    
    # Check PBP downloads
    pbp_files = list(Path('data/raw/pbp_live').glob('*/pbp_first_half_backfill.json'))
    pbp_count = len(pbp_files)
    
    print(f"[POLL {poll_num:02d} @ {now}]")
    print(f"  Workbook total rows: {wb_total}")
    print(f"  Workbook Nov-Jan rows: {wb_nov_jan}")
    print(f"  Baselines (Nov-Jan): {baseline_count}")
    print(f"  PBP files cached: {pbp_count}")
    print("-" * 70)
    
    if wb_nov_jan > 0:
        print(f"✓ WORKBOOK UPDATED! Nov-Jan rows detected: {wb_nov_jan}")
        print("  Backfill phase 4 (workbook write) has completed!")
        break
    
    time.sleep(900)  # 15 minutes
