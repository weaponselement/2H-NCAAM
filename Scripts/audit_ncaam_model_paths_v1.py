import os
import glob
from collections import Counter
from openpyxl import load_workbook

ROOT = r"C:\NCAA Model"
SCRIPTS_DIR = os.path.join(ROOT, "Scripts")
DATA_DIR = os.path.join(ROOT, "data")
LOGS_DIR = os.path.join(ROOT, "logs")
WORKBOOK = os.path.join(LOGS_DIR, "NCAAM Results.xlsx")

EXPECTED_DIRS = [
    ROOT,
    SCRIPTS_DIR,
    DATA_DIR,
    os.path.join(DATA_DIR, "raw"),
    os.path.join(DATA_DIR, "raw", "pbp"),
    os.path.join(DATA_DIR, "raw", "pbp_live"),
    os.path.join(DATA_DIR, "processed"),
    os.path.join(DATA_DIR, "processed", "baselines"),
    os.path.join(DATA_DIR, "processed", "reports"),
    os.path.join(DATA_DIR, "processed", "selected_games"),
    os.path.join(DATA_DIR, "processed", "slates"),
    LOGS_DIR,
]

EXPECTED_SCRIPTS = [
    "step4b_feature_report_from_file_v5_test.py",
    "log_prediction_to_results_v1.py",
    "update_results_postgame_v1.py",
    "update_new_results_only_v1.py",
    "update_all_results_v1.py",
]

KEY_GLOBS = {
    "baseline_manifests": os.path.join(DATA_DIR, "processed", "baselines", "last*.json"),
    "selected_games": os.path.join(DATA_DIR, "processed", "selected_games", "selected_games_*.json"),
    "reports_v5": os.path.join(DATA_DIR, "processed", "reports", "feature_report_v5_test_*.json"),
    "reports_v4": os.path.join(DATA_DIR, "processed", "reports", "feature_report_v4_*.json"),
    "slates": os.path.join(DATA_DIR, "processed", "slates", "slate_d1_*.json"),
    "pbp_live_first_half": os.path.join(DATA_DIR, "raw", "pbp_live", "*", "pbp_first_half_*.json"),
    "pbp_live_full": os.path.join(DATA_DIR, "raw", "pbp_live", "*", "pbp_*.json"),
    "pbp_baselines": os.path.join(DATA_DIR, "raw", "pbp", "*", "*.json"),
}


def count_files(pattern):
    return len(glob.glob(pattern))


def workbook_summary(path):
    if not os.path.exists(path):
        return {"exists": False}

    wb = load_workbook(path, data_only=True)
    if "Game_Log" not in wb.sheetnames:
        return {"exists": True, "has_game_log": False}

    ws = wb["Game_Log"]

    row_count = 0
    gids = []
    missing_results = 0

    for row in range(2, ws.max_row + 1):
        gid = ws[f"B{row}"].value
        if gid in (None, ""):
            continue
        gid = str(gid).strip()
        if not gid:
            continue

        row_count += 1
        gids.append(gid)

        actual = ws[f"L{row}"].value
        if actual in (None, ""):
            missing_results += 1

    counts = Counter(gids)
    dupes = {gid: n for gid, n in counts.items() if n > 1}

    return {
        "exists": True,
        "has_game_log": True,
        "rows_with_gameid": row_count,
        "unique_gameids": len(set(gids)),
        "duplicate_gameids": len(dupes),
        "missing_results_rows": missing_results,
    }


def main():
    print("=" * 70)
    print("NCAA MODEL PATH / DATA AUDIT")
    print("=" * 70)

    print("\n[Directories]")
    for d in EXPECTED_DIRS:
        print(f"{'[OK]' if os.path.exists(d) else '[MISSING]'} {d}")

    print("\n[Scripts]")
    for name in EXPECTED_SCRIPTS:
        path = os.path.join(SCRIPTS_DIR, name)
        print(f"{'[OK]' if os.path.exists(path) else '[MISSING]'} {path}")

    print("\n[Key File Counts]")
    for label, pattern in KEY_GLOBS.items():
        print(f"{label:22s}: {count_files(pattern)}")

    print("\n[Workbook]")
    summary = workbook_summary(WORKBOOK)
    if not summary.get("exists"):
        print(f"[MISSING] {WORKBOOK}")
    elif not summary.get("has_game_log"):
        print(f"[BAD] Workbook exists but Game_Log sheet not found: {WORKBOOK}")
    else:
        print(f"[OK] {WORKBOOK}")
        print(f"rows_with_gameid   : {summary['rows_with_gameid']}")
        print(f"unique_gameids     : {summary['unique_gameids']}")
        print(f"duplicate_gameids  : {summary['duplicate_gameids']}")
        print(f"missing_results_rows: {summary['missing_results_rows']}")

    print("\nAudit complete.")


if __name__ == "__main__":
    main()