#!/usr/bin/env python3
"""
Enhanced evaluator that tests different historical game lookback windows.
Accepts --window parameter to test last N games (4, 5, 6, 7, 8).
Outputs structured JSON results for comparison.
"""

import argparse
import csv
import math
import json
import sys
import statistics
from pathlib import Path
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error

# Import utilities
import importlib.util
spec = importlib.util.spec_from_file_location("model_feature_utils", 
    Path(__file__).resolve().parent / "model_feature_utils.py")
mfu = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mfu)

WORKBOOK_PATH = Path(__file__).resolve().parent.parent / 'logs' / 'NCAAM Results.xlsx'
SHEET_NAME = 'Game_Log'
CANONICAL_LINES_PATH = Path(__file__).resolve().parent.parent / 'data' / 'processed' / 'market_lines' / 'canonical_lines.csv'
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_ROOT = PROJECT_ROOT / 'data'
LOGS_DIR = PROJECT_ROOT / 'logs'

PBP_PRIOR_KEYS = [
    'last4_three_rate',
    'last4_paint_share',
    'last4_ft_rate',
    'last4_turnover_rate',
    'last4_orb_rate',
    'last4_possessions_per_team_1h',
    'last4_pbp_coverage_count',
]

def safe_float(value):
    try:
        return float(value)
    except Exception:
        return None

def build_rows():
    wb = load_workbook(str(WORKBOOK_PATH), data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else '' for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        data.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return data

def load_lines():
    out = {}
    with open(str(CANONICAL_LINES_PATH), 'r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            gid = str(row.get('game_id') or '').strip()
            if not gid:
                continue
            out[gid] = safe_float(row.get('total_game'))
    return out


def load_team_stats_for_window(date_str, window):
    """Load per-team avg scored/allowed from window-specific baseline when available."""
    baselines_dir = PROJECT_ROOT / 'data' / 'processed' / 'baselines'
    window_baseline_path = baselines_dir / f'lastN_{window}_{date_str}.json'
    if not window_baseline_path.exists():
        return mfu.load_team_stats(date_str)

    try:
        with open(window_baseline_path, 'r', encoding='utf-8') as f:
            baseline = json.load(f)
    except Exception:
        return mfu.load_team_stats(date_str)

    teams = baseline.get('teams', {}) if isinstance(baseline, dict) else {}
    stats = {}
    for team, games in teams.items():
        if not isinstance(games, list) or not games:
            continue
        scores_for = [safe_float(g.get('score_for')) for g in games]
        scores_against = [safe_float(g.get('score_against')) for g in games]
        scores_for = [x for x in scores_for if x is not None]
        scores_against = [x for x in scores_against if x is not None]
        if scores_for and scores_against:
            stats[team] = {
                'avg_scored': statistics.mean(scores_for),
                'avg_allowed': statistics.mean(scores_against),
            }
    return stats

def get_team_priors(last4_priors, team_name):
    team_data = (last4_priors or {}).get(str(team_name or '').strip(), {})
    return {key: float(team_data.get(key, 0.0) or 0.0) for key in PBP_PRIOR_KEYS}

def expected_total_from_team_stats(home_avg_scored, home_avg_allowed, away_avg_scored, away_avg_allowed):
    expected_home_total = (home_avg_scored + away_avg_allowed) / 2.0
    expected_away_total = (away_avg_scored + home_avg_allowed) / 2.0
    return expected_home_total + expected_away_total

def build_feature_row(record, team_stats_by_date, last4_priors_by_date):
    date_str = str(record.get('Date') or '').split(' ')[0].strip()
    home_team = str(record.get('Home') or '').strip()
    away_team = str(record.get('Away') or '').strip()
    actual_total = safe_float(record.get('ActualTotal'))
    game_id = str(record.get('GameID') or '').strip()

    if not date_str or not home_team or not away_team or actual_total is None or not game_id:
        return None

    stats = team_stats_by_date.get(date_str, {})
    home_avg_scored, home_avg_allowed = mfu.resolve_team_stats(stats, home_team)
    away_avg_scored, away_avg_allowed = mfu.resolve_team_stats(stats, away_team)
    expected_total = expected_total_from_team_stats(
        home_avg_scored,
        home_avg_allowed,
        away_avg_scored,
        away_avg_allowed,
    )

    priors = last4_priors_by_date.get(date_str, {})
    home_prior = get_team_priors(priors, home_team)
    away_prior = get_team_priors(priors, away_team)

    blended_possessions_1h = (home_prior['last4_possessions_per_team_1h'] + away_prior['last4_possessions_per_team_1h']) / 2.0
    blended_three_rate = (home_prior['last4_three_rate'] + away_prior['last4_three_rate']) / 2.0
    blended_paint_share = (home_prior['last4_paint_share'] + away_prior['last4_paint_share']) / 2.0
    blended_ft_rate = (home_prior['last4_ft_rate'] + away_prior['last4_ft_rate']) / 2.0
    blended_turnover_rate = (home_prior['last4_turnover_rate'] + away_prior['last4_turnover_rate']) / 2.0
    blended_orb_rate = (home_prior['last4_orb_rate'] + away_prior['last4_orb_rate']) / 2.0
    blended_coverage = (home_prior['last4_pbp_coverage_count'] + away_prior['last4_pbp_coverage_count']) / 2.0

    features = {
        'home_avg_scored': home_avg_scored,
        'home_avg_allowed': home_avg_allowed,
        'away_avg_scored': away_avg_scored,
        'away_avg_allowed': away_avg_allowed,
        'expected_total': expected_total,
        'home_offense_diff': home_avg_scored - away_avg_allowed,
        'away_offense_diff': away_avg_scored - home_avg_allowed,
        'blended_possessions_1h': blended_possessions_1h,
        'blended_possessions_full': blended_possessions_1h * 2.0,
        'blended_three_rate': blended_three_rate,
        'blended_paint_share': blended_paint_share,
        'blended_ft_rate': blended_ft_rate,
        'blended_turnover_rate': blended_turnover_rate,
        'blended_orb_rate': blended_orb_rate,
        'blended_pbp_coverage': blended_coverage,
        'home_last4_three_rate': home_prior['last4_three_rate'],
        'away_last4_three_rate': away_prior['last4_three_rate'],
        'home_last4_ft_rate': home_prior['last4_ft_rate'],
        'away_last4_ft_rate': away_prior['last4_ft_rate'],
        'home_last4_turnover_rate': home_prior['last4_turnover_rate'],
        'away_last4_turnover_rate': away_prior['last4_turnover_rate'],
        'home_last4_orb_rate': home_prior['last4_orb_rate'],
        'away_last4_orb_rate': away_prior['last4_orb_rate'],
        'home_last4_possessions_1h': home_prior['last4_possessions_per_team_1h'],
        'away_last4_possessions_1h': away_prior['last4_possessions_per_team_1h'],
        'three_rate_gap': home_prior['last4_three_rate'] - away_prior['last4_three_rate'],
        'ft_rate_gap': home_prior['last4_ft_rate'] - away_prior['last4_ft_rate'],
        'turnover_rate_gap': home_prior['last4_turnover_rate'] - away_prior['last4_turnover_rate'],
        'orb_rate_gap': home_prior['last4_orb_rate'] - away_prior['last4_orb_rate'],
    }

    return {
        'game_id': game_id,
        'date': date_str,
        'home': home_team,
        'away': away_team,
        'actual_total': actual_total,
        'features': features,
        'expected_total': expected_total,
    }

def side_hit(pred_total, closing_total, actual_total):
    if pred_total is None or closing_total is None or actual_total is None:
        return None
    if pred_total == closing_total:
        return None
    if actual_total == closing_total:
        return None
    pred_side = 'OVER' if pred_total > closing_total else 'UNDER'
    actual_side = 'OVER' if actual_total > closing_total else 'UNDER'
    return int(pred_side == actual_side)

def payout_per_unit_risk(odds):
    if odds == 0:
        raise ValueError('Odds cannot be 0')
    if odds > 0:
        return odds / 100.0
    return 100.0 / abs(odds)

def compute_gap_policy_results(rows, pred_key, odds, min_bets):
    """Compute gap-based policy performance and return best threshold metrics."""
    per_win = payout_per_unit_risk(odds)
    graded = []
    for row in rows:
        closing_total = row.get('closing_total')
        pred_total = row.get(pred_key)
        actual_total = row.get('actual_total')
        hit = side_hit(pred_total, closing_total, actual_total)
        if hit is None:
            continue
        graded.append({'hit': hit, 'gap': abs(pred_total - closing_total)})

    if not graded:
        return None

    best = None
    best_roi = -float('inf')
    for threshold in [0, 1, 2, 3, 4, 5, 6, 8, 10]:
        wagers = [x for x in graded if x['gap'] >= threshold]
        n = len(wagers)
        if n < min_bets:
            continue
        wins = sum(x['hit'] for x in wagers)
        losses = n - wins
        units = wins * per_win - losses
        roi = (units / n) * 100.0
        hit_rate = (wins / n) * 100.0
        
        if roi > best_roi:
            best_roi = roi
            best = {
                'threshold': threshold,
                'bets': n,
                'wins': wins,
                'hit_rate': hit_rate,
                'units': units,
                'roi': roi,
            }

    return best

def main():
    parser = argparse.ArgumentParser(description='Test pregame model at different lookback windows')
    parser.add_argument('--window', type=int, default=4, help='Lookback window (4, 5, 6, 7, 8)')
    parser.add_argument('--source-date', type=str, default='2026-03-26', help='Date for baseline files')
    parser.add_argument('--odds', type=int, default=-110, help='Odds for ROI simulation')
    parser.add_argument('--min-bets', type=int, default=20, help='Min bets to report')
    parser.add_argument('--full-gap', type=float, default=10.0, help='Full stake threshold')
    parser.add_argument('--half-gap', type=float, default=6.0, help='Half stake threshold')
    parser.add_argument('--mode', type=str, default='walkforward', choices=['single', 'walkforward'],
                        help='Evaluation mode: single (one 75/25 split) or walkforward (rolling folds)')
    parser.add_argument('--min-train-dates', type=int, default=60,
                        help='Minimum training dates before first test fold (walkforward mode)')
    parser.add_argument('--fold-size', type=int, default=30,
                        help='Number of dates per test fold (walkforward mode)')
    parser.add_argument('--start-date', type=str, default='',
                        help='Optional inclusive start date (YYYY-MM-DD) for evaluation rows')
    parser.add_argument('--end-date', type=str, default='',
                        help='Optional inclusive end date (YYYY-MM-DD) for evaluation rows')
    args = parser.parse_args()

    window = args.window
    print(f"\n{'='*70}", flush=True)
    print(f"Testing Window: last {window} games per team | Mode: {args.mode}", flush=True)
    print(f"{'='*70}\n", flush=True)

    # Load data
    workbook_rows = build_rows()
    lines = load_lines()

    if args.start_date or args.end_date:
        filtered_rows = []
        for row in workbook_rows:
            date_str = str(row.get('Date') or '').split(' ')[0].strip()
            if not date_str:
                continue
            if args.start_date and date_str < args.start_date:
                continue
            if args.end_date and date_str > args.end_date:
                continue
            filtered_rows.append(row)
        print(
            f"Date filter active: {args.start_date or 'BEGIN'} -> {args.end_date or 'END'} | "
            f"rows: {len(workbook_rows)} -> {len(filtered_rows)}",
            flush=True,
        )
        workbook_rows = filtered_rows

    unique_dates = sorted({str(r.get('Date') or '').split(' ')[0].strip() for r in workbook_rows if r.get('Date') not in (None, '')})
    all_dates = sorted(unique_dates)

    print(f'Loading team stats for {len(all_dates)} dates...', flush=True)
    team_stats_by_date = {date_str: load_team_stats_for_window(date_str, window) for date_str in all_dates}

    # Load window-specific baselines if they exist, otherwise use default last4
    last4_priors_by_date = {}
    for date_str in all_dates:
        baselines_dir = PROJECT_ROOT / 'data' / 'processed' / 'baselines'
        window_baseline_path = baselines_dir / f'lastN_{window}_{date_str}.json'

        if window_baseline_path.exists():
            try:
                with open(window_baseline_path, 'r') as f:
                    baseline = json.load(f)
                from step4b_feature_report_from_file_v5_test import load_game_pbp_features
                priors = {}
                teams = baseline.get('teams', {})
                for team_seo, games in teams.items():
                    if not isinstance(games, list) or not games:
                        continue
                    prior_rows = []
                    for g in games:
                        gid = str((g or {}).get('gameID') or '').strip()
                        if not gid:
                            continue
                        pbp = load_game_pbp_features(str(DATA_ROOT), gid)
                        if pbp:
                            prior_rows.append(pbp)

                    if not prior_rows:
                        priors[team_seo] = {k: v for k, v in mfu.DEFAULT_PBP_FEATURES.items() if 'home_last4' in k}
                        continue

                    def _mean(key, default):
                        vals = [float(r.get(key)) for r in prior_rows if r.get(key) is not None]
                        return statistics.mean(vals) if vals else float(default)

                    priors[team_seo] = {
                        'last4_three_rate': _mean('home_three_rate', mfu.DEFAULT_PBP_FEATURES['home_last4_three_rate']),
                        'last4_paint_share': _mean('home_paint_share', mfu.DEFAULT_PBP_FEATURES['home_last4_paint_share']),
                        'last4_ft_rate': _mean('home_ft_rate', mfu.DEFAULT_PBP_FEATURES['home_last4_ft_rate']),
                        'last4_turnover_rate': _mean('home_turnover_rate', mfu.DEFAULT_PBP_FEATURES['home_last4_turnover_rate']),
                        'last4_orb_rate': _mean('home_orb_rate', mfu.DEFAULT_PBP_FEATURES['home_last4_orb_rate']),
                        'last4_possessions_per_team_1h': _mean('possessions_per_team_1h', mfu.DEFAULT_PBP_FEATURES['home_last4_possessions_per_team_1h']),
                        'last4_pbp_coverage_count': float(len(prior_rows)),
                    }
                last4_priors_by_date[date_str] = priors
            except Exception as e:
                print(f"Error loading window baseline for {date_str}: {e}", flush=True)
                last4_priors_by_date[date_str] = mfu.load_last4_pbp_priors(date_str, data_root=str(DATA_ROOT))
        else:
            last4_priors_by_date[date_str] = mfu.load_last4_pbp_priors(date_str, data_root=str(DATA_ROOT))

    # Build feature rows
    feature_rows = []
    for record in workbook_rows:
        built = build_feature_row(record, team_stats_by_date, last4_priors_by_date)
        if built is None:
            continue
        built['closing_total'] = lines.get(built['game_id'])
        feature_rows.append(built)

    feature_rows = [r for r in feature_rows if r.get('closing_total') is not None]
    print(f'Rows with closings: {len(feature_rows)}', flush=True)

    if not feature_rows:
        print('ERROR: No rows with closing lines')
        return

    feature_names = list(feature_rows[0]['features'].keys())
    dates_with_data = sorted({r['date'] for r in feature_rows})

    # ------------------------------------------------------------------ #
    #  SINGLE split (original 75/25 behaviour)                            #
    # ------------------------------------------------------------------ #
    if args.mode == 'single':
        test_dates_count = max(1, int(math.ceil(len(dates_with_data) * 0.25)))
        if len(dates_with_data) - test_dates_count < 20:
            test_dates_count = max(1, len(dates_with_data) - 20)
        split_index = len(dates_with_data) - test_dates_count
        train_dates = set(dates_with_data[:split_index])
        test_dates = set(dates_with_data[split_index:])

        print(f'Total dates: {len(dates_with_data)} | Train: {len(train_dates)} | Test: {len(test_dates)}', flush=True)
        print(f'Test window: {min(test_dates)} -> {max(test_dates)}', flush=True)

        train_rows = [r for r in feature_rows if r['date'] in train_dates]
        test_rows  = [r for r in feature_rows if r['date'] in test_dates]
        print(f'Train: {len(train_rows)} | Test: {len(test_rows)}\n', flush=True)

        if not train_rows or not test_rows:
            print('ERROR: Insufficient rows')
            return

        x_train = [[r['features'][name] for name in feature_names] for r in train_rows]
        y_train  = [r['actual_total'] for r in train_rows]
        x_test   = [[r['features'][name] for name in feature_names] for r in test_rows]

        model = RandomForestRegressor(n_estimators=400, min_samples_leaf=3, random_state=42, n_jobs=-1)
        model.fit(x_train, y_train)
        preds = model.predict(x_test)

        for row, pred in zip(test_rows, preds):
            row['model_pred_total']    = float(pred)
            row['baseline_pred_total'] = float(row['expected_total'])

        all_test_rows = test_rows
        importances = sorted(zip(feature_names, model.feature_importances_), key=lambda x: x[1], reverse=True)

    # ------------------------------------------------------------------ #
    #  WALK-FORWARD rolling folds                                         #
    # ------------------------------------------------------------------ #
    else:
        min_train = args.min_train_dates
        fold_size = args.fold_size

        if len(dates_with_data) < min_train + fold_size:
            print(f'ERROR: Only {len(dates_with_data)} dates — need at least {min_train + fold_size} for walk-forward')
            return

        # Build fold definitions: each fold trains on everything before split_idx,
        # tests on the next fold_size dates.
        folds = []
        split_idx = min_train
        while split_idx < len(dates_with_data):
            fold_train_dates = set(dates_with_data[:split_idx])
            fold_test_dates  = set(dates_with_data[split_idx:split_idx + fold_size])
            if fold_test_dates:
                folds.append((fold_train_dates, fold_test_dates))
            split_idx += fold_size

        print(f'Walk-forward: {len(folds)} folds | min_train={min_train} dates | fold_size={fold_size} dates', flush=True)
        print(f'Date range: {dates_with_data[0]} -> {dates_with_data[-1]}\n', flush=True)

        all_test_rows   = []
        last_importances = None

        for fold_idx, (fold_train_dates, fold_test_dates) in enumerate(folds):
            train_rows = [r for r in feature_rows if r['date'] in fold_train_dates]
            test_rows  = [r for r in feature_rows if r['date'] in fold_test_dates]

            if not train_rows or not test_rows:
                continue

            x_train = [[r['features'][name] for name in feature_names] for r in train_rows]
            y_train  = [r['actual_total'] for r in train_rows]
            x_test   = [[r['features'][name] for name in feature_names] for r in test_rows]

            model = RandomForestRegressor(n_estimators=400, min_samples_leaf=3, random_state=42, n_jobs=-1)
            model.fit(x_train, y_train)
            preds = model.predict(x_test)

            for row, pred in zip(test_rows, preds):
                row['model_pred_total']    = float(pred)
                row['baseline_pred_total'] = float(row['expected_total'])

            all_test_rows.extend(test_rows)
            last_importances = list(zip(feature_names, model.feature_importances_))

            fold_sorted = sorted(fold_test_dates)
            print(f'  Fold {fold_idx+1:2d}: train={len(train_rows):4d} | test={len(test_rows):3d} | {fold_sorted[0]} -> {fold_sorted[-1]}', flush=True)

        if last_importances is None:
            print('ERROR: No folds completed')
            return

        importances = sorted(last_importances, key=lambda x: x[1], reverse=True)
        print(f'\nTotal graded test rows across all folds: {len(all_test_rows)}', flush=True)

    # ------------------------------------------------------------------ #
    #  Aggregate metrics (shared by both modes)                           #
    # ------------------------------------------------------------------ #
    y_test       = [r['actual_total'] for r in all_test_rows]
    mae_baseline = mean_absolute_error(y_test, [r['expected_total']      for r in all_test_rows])
    mae_rf       = mean_absolute_error(y_test, [r['model_pred_total']    for r in all_test_rows])

    baseline_policy = compute_gap_policy_results(all_test_rows, 'baseline_pred_total', args.odds, args.min_bets)
    rf_policy       = compute_gap_policy_results(all_test_rows, 'model_pred_total',    args.odds, args.min_bets)

    top_features = {name: float(val) for name, val in importances[:10]}

    results = {
        'window': window,
        'mode': args.mode,
        'date_filter': {
            'start_date': args.start_date or None,
            'end_date': args.end_date or None,
        },
        'metrics': {
            'mae_baseline':    round(mae_baseline, 3),
            'mae_rf':          round(mae_rf, 3),
            'total_test_rows': len(all_test_rows),
        },
        'baseline_policy': baseline_policy,
        'rf_policy':        rf_policy,
        'top_10_features':  top_features,
    }

    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = LOGS_DIR / f'window_test_results_{window}.json'
    with open(output_path, 'w') as f:
        json.dump(results, f, indent=2)

    print(f'\nBaseline MAE: {mae_baseline:.3f}')
    print(f'RF Model MAE: {mae_rf:.3f}')
    if baseline_policy:
        print(f'\nBaseline best policy:')
        print(f'  gap>={baseline_policy["threshold"]}: {baseline_policy["bets"]} bets, {baseline_policy["hit_rate"]:.1f}% hit, +{baseline_policy["units"]:.2f}u, {baseline_policy["roi"]:.2f}% ROI')
    if rf_policy:
        print(f'\nRF model best policy:')
        print(f'  gap>={rf_policy["threshold"]}: {rf_policy["bets"]} bets, {rf_policy["hit_rate"]:.1f}% hit, +{rf_policy["units"]:.2f}u, {rf_policy["roi"]:.2f}% ROI')

    print(f'\nTop 5 features:')
    for i, (name, val) in enumerate(list(importances)[:5], 1):
        print(f'  {i}. {name}: {val:.4f}')

    print(f'\nResults saved to {output_path}\n')

if __name__ == '__main__':
    main()
