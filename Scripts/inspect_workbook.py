from openpyxl import load_workbook
wb = load_workbook('logs/NCAAM Results.xlsx', data_only=True)
ws = wb['Game_Log']
print([c.value for c in ws[1]])
print(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

if ws['S1'].value or ws['T1'].value:
	print('narrow range headers', ws['S1'].value, ws['T1'].value)
