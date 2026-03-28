import argparse
import csv
from openpyxl import load_workbook


WORKBOOK_PATH = 'logs/NCAAM Results.xlsx'
SHEET_NAME = 'Game_Log'
CANONICAL_LINES_PATH = 'data/processed/market_lines/canonical_lines.csv'


def safe_float(value):
    try:
        return float(value)
    except Exception:
        return None


def parse_range(value):
    if value in (None, ''):
        return None
    s = str(value)
    if '-' not in s:
        return None
    left, right = s.split('-', 1)
    try:
        lo = float(left.strip())
        hi = float(right.strip())
    except Exception:
        return None
    if lo > hi:
        lo, hi = hi, lo
    return lo, hi


def build_row_dicts():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else '' for c in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return out


def load_lines():
    out = {}
    with open(CANONICAL_LINES_PATH, 'r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            gid = str(row.get('game_id') or '').strip()
            if not gid:
                continue
            out[gid] = {
                'total_game': safe_float(row.get('total_game')),
                'total_2h': safe_float(row.get('total_2h')),
            }
    return out


def build_stats(records, line_key, pred_key, actual_key):
    usable = []
    for r in records:
        line = r.get(line_key)
        pred_range = parse_range(r.get(pred_key))
        actual = safe_float(r.get(actual_key))
        if line is None or pred_range is None or actual is None:
            continue
        mid = (pred_range[0] + pred_range[1]) / 2.0
        if mid == line:
            continue
        pred_side = 'OVER' if mid > line else 'UNDER'
        actual_side = 'OVER' if actual > line else 'UNDER'
        usable.append({
            'hit': 1 if pred_side == actual_side else 0,
            'gap': abs(mid - line),
            'pred_side': pred_side,
            'actual_side': actual_side,
            'line': line,
            'mid': mid,
            'actual': actual,
        })
    return usable


def print_summary(name, usable):
    n = len(usable)
    if n == 0:
        print(f'{name}: no usable rows')
        return
    hit = sum(x['hit'] for x in usable)
    over_actual = sum(1 for x in usable if x['actual_side'] == 'OVER')
    over_pred = sum(1 for x in usable if x['pred_side'] == 'OVER')
    print(f'{name}: n={n} hit={100.0 * hit / n:.1f}%')
    print(f'  actual_over_rate={100.0 * over_actual / n:.1f}% pred_over_rate={100.0 * over_pred / n:.1f}%')
    for threshold in [0, 1, 2, 3, 4, 5, 6]:
        subset = [x for x in usable if x['gap'] >= threshold]
        if len(subset) < 20:
            continue
        h = sum(x['hit'] for x in subset)
        print(f'  gap>={threshold}: n={len(subset)} hit={100.0 * h / len(subset):.1f}%')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--wagered-only', action='store_true', help='Use only WageredFlag=Y rows')
    parser.add_argument('--trigger-version', default='', help='Filter by TriggerVersion value, e.g. v1')
    parser.add_argument('--start-date', default='', help='Filter Date >= YYYY-MM-DD')
    args = parser.parse_args()

    rows = build_row_dicts()
    lines = load_lines()

    joined = []
    for r in rows:
        gid = str(r.get('GameID') or '').strip()
        if not gid:
            continue
        line = lines.get(gid)
        if line is None:
            continue

        if args.wagered_only:
            if str(r.get('WageredFlag') or '').strip().upper() != 'Y':
                continue

        if args.trigger_version:
            if str(r.get('TriggerVersion') or '').strip() != args.trigger_version:
                continue

        if args.start_date:
            date_value = str(r.get('Date') or '').strip()
            if not date_value or date_value < args.start_date:
                continue

        rec = dict(r)
        rec['line_total_game'] = line.get('total_game')
        rec['line_total_2h'] = line.get('total_2h')
        joined.append(rec)

    print(f'Joined rows with line coverage: {len(joined)}')

    full_stats = build_stats(joined, 'line_total_game', 'PredTotalRange', 'ActualTotal')
    second_half_stats = build_stats(joined, 'line_total_2h', 'Pred2HRange', 'Actual2H')

    print_summary('FULL_GAME_TOTAL_OU', full_stats)
    print_summary('SECOND_HALF_TOTAL_OU', second_half_stats)

    if len(second_half_stats) == 0:
        print('Note: second-half O/U lines are missing or unusable in current canonical_lines.csv for matched rows.')


if __name__ == '__main__':
    main()
