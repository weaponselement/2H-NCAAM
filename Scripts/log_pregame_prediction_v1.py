#!/usr/bin/env python3
"""
Log pregame predictions to workbook.

Usage:
  python log_pregame_prediction_v1.py \
    --date 2026-03-28 \
    --home michigan-st \
    --away uconn \
    --game-id 6530122 \
    --market-line 136.5 \
    --pred-total 145.6 \
    --pred-gap 9.1 \
    --lean OVER \
    --trigger "LEAN (gap 8-9)"

This writes (or updates) a row in the workbook with the pregame prediction.
After the game, use --actual-winner and --actual-total to fill results and calculate PredictionHit.
"""

import argparse
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "logs" / "NCAAM Results.xlsx"


def safe_float(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def find_or_create_row(ws, date_str, game_id, home, away):
    """Find existing row or return next available row index."""
    game_id_str = str(game_id).strip()
    
    # Search for existing row
    for row_idx in range(2, ws.max_row + 1):
        row_gid = ws.cell(row=row_idx, column=2).value
        if row_gid and str(row_gid).strip() == game_id_str:
            return row_idx
    
    # Find next empty row
    for row_idx in range(2, 10000):
        if not ws.cell(row=row_idx, column=2).value:
            return row_idx
    
    return ws.max_row + 1


def log_prediction(date_str, home, away, game_id, market_line, pred_total, 
                   pred_gap, lean, trigger, actual_winner=None, actual_total=None):
    """Write prediction to workbook."""
    
    wb = load_workbook(str(WORKBOOK_PATH))
    ws = wb['Game_Log']
    
    row_idx = find_or_create_row(ws, date_str, game_id, home, away)
    
    # Write/update pregame data
    ws.cell(row=row_idx, column=1).value = date_str  # Date
    ws.cell(row=row_idx, column=2).value = game_id   # GameID
    ws.cell(row=row_idx, column=3).value = away      # Away
    ws.cell(row=row_idx, column=4).value = home      # Home
    ws.cell(row=row_idx, column=5).value = safe_float(market_line)  # MarketTotalLine
    ws.cell(row=row_idx, column=6).value = safe_float(pred_total)   # PregamePredTotal
    ws.cell(row=row_idx, column=7).value = safe_float(pred_gap)     # PregamePredGap
    ws.cell(row=row_idx, column=8).value = lean                     # PregameLean
    ws.cell(row=row_idx, column=9).value = trigger                  # PregameTrigger
    
    # Write results if provided
    if actual_winner is not None:
        ws.cell(row=row_idx, column=10).value = actual_winner       # ActualWinner
    
    if actual_total is not None:
        actual_total_f = safe_float(actual_total)
        ws.cell(row=row_idx, column=11).value = actual_total_f      # ActualTotal
        
        # Calculate PredictionHit (did directional lean come true?)
        if lean and market_line is not None and actual_total_f is not None:
            market_line_f = safe_float(market_line)
            if market_line_f is not None:
                actual_over = actual_total_f > market_line_f
                pred_over = lean.upper() == "OVER"
                hit = 1 if (actual_over == pred_over) else 0
                ws.cell(row=row_idx, column=12).value = hit         # PredictionHit
    
    wb.save(str(WORKBOOK_PATH))
    return row_idx


def main():
    parser = argparse.ArgumentParser(description="Log pregame prediction to workbook")
    parser.add_argument("--date", required=True, help="Date (YYYY-MM-DD)")
    parser.add_argument("--home", required=True, help="Home team SEO")
    parser.add_argument("--away", required=True, help="Away team SEO")
    parser.add_argument("--game-id", required=True, help="GameID")
    parser.add_argument("--market-line", type=float, help="Market total line")
    parser.add_argument("--pred-total", type=float, help="Model prediction")
    parser.add_argument("--pred-gap", type=float, help="Signed gap (pred - line)")
    parser.add_argument("--lean", help="OVER or UNDER")
    parser.add_argument("--trigger", help="Trigger tier")
    parser.add_argument("--actual-winner", help="Actual winner (for postgame)")
    parser.add_argument("--actual-total", type=float, help="Actual total (for postgame)")
    
    args = parser.parse_args()
    
    row_idx = log_prediction(
        date_str=args.date,
        home=args.home,
        away=args.away,
        game_id=args.game_id,
        market_line=args.market_line,
        pred_total=args.pred_total,
        pred_gap=args.pred_gap,
        lean=args.lean,
        trigger=args.trigger,
        actual_winner=args.actual_winner,
        actual_total=args.actual_total,
    )
    
    print(f"✓ Logged to workbook row {row_idx}")
    if args.actual_winner and args.actual_total:
        print(f"  Result: {args.actual_winner}")
    else:
        print(f"  Waiting for game result...")


if __name__ == "__main__":
    main()
