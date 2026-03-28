"""
backfill_prior_season_v1.py

Backfills one full prior NCAAB season into NCAAM Results.xlsx.

Phases:
  1. Scan scoreboard API for all completed D1 games in the target season
     -> collects game_id, team SEOs, final scores, caches scoreboard JSONs
  2. Build rolling last4 baselines for each game date
     -> writes data/processed/baselines/last4_YYYY-MM-DD.json for each date
  3. Download PBP for each game (threaded, resumable)
     -> extracts first-half plays, saves to data/raw/pbp_live/{id}/pbp_first_half_backfill.json
     -> derives halftime score from first-half plays
  4. Write rows to Game_Log workbook
     -> skips game_ids already present
     -> leaves Pred* columns blank (fill with update_all_predictions_ml.py after retrain)

After running:
  python Scripts/tune_totals_spreads_v1.py        (retrain with full dataset)
  python Scripts/update_all_predictions_ml.py     (fill Pred* columns for all rows)

Usage:
  python Scripts/backfill_prior_season_v1.py
  python Scripts/backfill_prior_season_v1.py --season 2024-25 --workers 8 --max-rps 4
  python Scripts/backfill_prior_season_v1.py --dry-run    (no workbook writes)
  python Scripts/backfill_prior_season_v1.py --skip-pbp  (skip PBP download, use cached only)
"""

import argparse
import json
import os
import random
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

# Add Scripts/ dir to path for local imports
sys.path.insert(0, str(Path(__file__).parent))
from paths import DATA_DIR, NCAAM_RESULTS_XLSX
from model_feature_utils import (
    FEATURE_NAMES, build_feature_vector, load_market_lines,
    load_team_stats, resolve_team_stats,
)
from step4b_feature_report_from_file_v5_test import load_game_pbp_features

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
API_BASE = "https://ncaa-api.henrygd.me"
HEADERS = {"User-Agent": "Mozilla/5.0"}
DATA_ROOT = str(DATA_DIR)

# Available seasons to backfill.  Add more as needed.
SEASON_DATE_RANGES = {
    "2024-25": (date(2024, 11, 4), date(2025, 3, 22)),
    "2025-26": (date(2025, 11, 4), date(2026, 1, 31)),
}
DEFAULT_SEASON = "2024-25"
DEFAULT_WORKERS = 8
DEFAULT_MAX_RPS = 4.0

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

class RateLimiter:
    """Thread-safe token-bucket rate limiter."""
    def __init__(self, max_rps: float):
        self.interval = 1.0 / max_rps if max_rps > 0 else 0.0
        self._lock = threading.Lock()
        self._next_time = time.monotonic()

    def wait(self):
        if self.interval <= 0:
            return
        with self._lock:
            now = time.monotonic()
            if now < self._next_time:
                time.sleep(self._next_time - now)
            self._next_time = max(self._next_time + self.interval, time.monotonic())


def fetch_json(url: str, timeout: int = 30, retries: int = 3):
    """HTTP GET with retry/backoff. Returns parsed JSON or None."""
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout)
            if r.status_code == 200:
                return r.json()
            if r.status_code in (429, 500, 502, 503, 504):
                wait = 2 ** attempt + random.uniform(0, 1.0)
                time.sleep(wait)
                continue
            r.raise_for_status()
        except requests.RequestException as exc:
            if attempt < retries:
                time.sleep(2 ** attempt)
            else:
                raise
    return None


def iter_dates(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def pace_label_from_halftime_total(ht_total: float) -> str:
    if ht_total <= 60:
        return "grinder"
    if ht_total <= 70:
        return "moderate"
    return "run_and_gun"


# ---------------------------------------------------------------------------
# Phase 1: Scan scoreboard
# ---------------------------------------------------------------------------

def scan_season_scoreboard(season: str, cache_root: Path) -> list:
    """
    Iterate every date in the season, fetch (or load cached) scoreboard,
    return list of completed-game dicts with final scores.
    """
    start, end = SEASON_DATE_RANGES[season]
    cache_root.mkdir(parents=True, exist_ok=True)

    games = []
    dates_with_games = 0

    for d in iter_dates(start, end):
        date_str = d.strftime("%Y-%m-%d")
        yyyy, mm, dd = d.strftime("%Y"), d.strftime("%m"), d.strftime("%d")

        cache_file = cache_root / f"scoreboard_{date_str}.json"

        if cache_file.exists():
            try:
                sb = json.loads(cache_file.read_text(encoding="utf-8"))
            except Exception:
                sb = None
        else:
            url = f"{API_BASE}/scoreboard/basketball-men/d1/{yyyy}/{mm}/{dd}"
            try:
                sb = fetch_json(url, timeout=30, retries=3)
                if sb:
                    cache_file.write_text(
                        json.dumps(sb, ensure_ascii=False), encoding="utf-8"
                    )
            except Exception as exc:
                print(f"  [WARN] Scoreboard fetch failed for {date_str}: {exc}")
                sb = None

        if not sb:
            continue

        day_games = []
        for item in sb.get("games", []):
            g = item.get("game", {})
            gid = str(g.get("gameID", "")).strip()
            if not gid:
                continue

            away_names = (g.get("away") or {}).get("names") or {}
            home_names = (g.get("home") or {}).get("names") or {}
            away_seo = away_names.get("seo")
            home_seo = home_names.get("seo")
            away_score_raw = (g.get("away") or {}).get("score")
            home_score_raw = (g.get("home") or {}).get("score")

            if not away_seo or not home_seo:
                continue
            try:
                away_score = int(away_score_raw)
                home_score = int(home_score_raw)
            except (TypeError, ValueError):
                continue  # skip non-final / not-started games

            # Basic sanity: NCAAB scores are typically 40-130 per team
            if not (30 <= away_score <= 150 and 30 <= home_score <= 150):
                continue

            day_games.append({
                "date": date_str,
                "game_id": gid,
                "away_seo": away_seo,
                "home_seo": home_seo,
                "away_short": away_names.get("short") or away_seo,
                "home_short": home_names.get("short") or home_seo,
                "away_score": away_score,
                "home_score": home_score,
            })

        if day_games:
            dates_with_games += 1
            if dates_with_games % 10 == 1:
                print(f"  {date_str}: {len(day_games)} games  (cumulative: {len(games) + len(day_games)})")
        games.extend(day_games)

    print(f"Total completed games: {len(games)} across {dates_with_games} dates")
    return games


# ---------------------------------------------------------------------------
# Phase 2: Build last4 baselines
# ---------------------------------------------------------------------------

def build_baselines_for_season(games: list, baselines_dir: Path, n: int = 4):
    """
    Walk games in chronological order, snapshot rolling last-n-game registry
    before each date's games, write last{n}_{date}.json if not already present.
    """
    baselines_dir.mkdir(parents=True, exist_ok=True)

    # Group by date, sorted
    games_by_date: dict[str, list] = {}
    for g in games:
        games_by_date.setdefault(g["date"], []).append(g)

    registry: dict[str, list] = {}  # team_seo -> list of game result dicts
    written = 0
    skipped = 0

    for date_str in sorted(games_by_date.keys()):
        out_path = baselines_dir / f"last{n}_{date_str}.json"

        if not out_path.exists():
            # Snapshot state BEFORE today's games
            snapshot = {team: list(hist[-n:]) for team, hist in registry.items() if hist}
            out_data = {
                "run_date": date_str,
                "source": "backfill_prior_season",
                "games_per_team": n,
                "teams": snapshot,
            }
            out_path.write_text(
                json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            written += 1
        else:
            skipped += 1

        # Update registry with today's games
        for g in games_by_date[date_str]:
            for role, seo, score_for, score_against, opp in (
                ("away", g["away_seo"], g["away_score"], g["home_score"], g["home_seo"]),
                ("home", g["home_seo"], g["home_score"], g["away_score"], g["away_seo"]),
            ):
                registry.setdefault(seo, []).append({
                    "date": date_str,
                    "gameID": g["game_id"],
                    "opponent_seo": opp,
                    "home_away": role,
                    "score_for": score_for,
                    "score_against": score_against,
                })

    print(f"Baselines: wrote {written} new, skipped {skipped} existing")


# ---------------------------------------------------------------------------
# Phase 3: Download PBP
# ---------------------------------------------------------------------------

def pbp_output_path(game_id: str) -> Path:
    return (
        Path(DATA_ROOT) / "raw" / "pbp_live" / str(game_id) / "pbp_first_half_backfill.json"
    )


def extract_first_half_plays(payload: dict) -> list:
    """Return first-half play list from PBP API response."""
    periods = payload.get("periods", [])
    if not isinstance(periods, list):
        return []
    for p in periods:
        if not isinstance(p, dict):
            continue
        pn = p.get("periodNumber")
        pd_label = (p.get("periodDisplay") or "").lower()
        if pn == 1 or "1st half" in pd_label or "first half" in pd_label:
            plays = p.get("playbyplayStats")
            return plays if isinstance(plays, list) else []
    return []


def extract_halftime_score(plays: list):
    """Return (away_ht, home_ht) from last scored play in first half, or None."""
    for p in reversed(plays):
        if not isinstance(p, dict):
            continue
        if "homeScore" in p and "visitorScore" in p:
            try:
                return int(p["visitorScore"]), int(p["homeScore"])  # (away, home)
            except (TypeError, ValueError):
                pass
    return None


def download_pbp_for_game(game_id: str, rate_limiter: RateLimiter) -> tuple:
    """
    Download or load cached PBP.  Returns (game_id, (away_ht, home_ht) or None).
    """
    out_path = pbp_output_path(game_id)

    # Load from cache
    if out_path.exists():
        try:
            data = json.loads(out_path.read_text(encoding="utf-8"))
            ht = extract_halftime_score(data.get("first_half_plays", []))
            return game_id, ht
        except Exception:
            pass  # re-download if corrupt

    url = f"{API_BASE}/game/{game_id}/play-by-play"
    rate_limiter.wait()
    try:
        r = requests.get(url, headers=HEADERS, timeout=45)
        if r.status_code != 200:
            return game_id, None
        payload = r.json()
    except Exception:
        return game_id, None

    plays = extract_first_half_plays(payload)
    if not plays:
        return game_id, None

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps({"first_half_plays": plays}, ensure_ascii=False),
        encoding="utf-8",
    )
    return game_id, extract_halftime_score(plays)


def download_all_pbp(games: list, workers: int, max_rps: float) -> dict:
    """Download PBP for all games, return {game_id: (away_ht, home_ht) or None}."""
    rate_limiter = RateLimiter(max_rps)
    halftime_scores: dict = {}

    already_done = [g for g in games if pbp_output_path(g["game_id"]).exists()]
    to_download = [g for g in games if not pbp_output_path(g["game_id"]).exists()]

    print(f"  Already downloaded: {len(already_done)}")
    print(f"  To download:        {len(to_download)}")

    # Load cached ones (fast)
    for g in already_done:
        out_path = pbp_output_path(g["game_id"])
        try:
            data = json.loads(out_path.read_text(encoding="utf-8"))
            halftime_scores[g["game_id"]] = extract_halftime_score(
                data.get("first_half_plays", [])
            )
        except Exception:
            halftime_scores[g["game_id"]] = None

    if not to_download:
        return halftime_scores

    done_count = [0]
    done_lock = threading.Lock()
    total = len(to_download)

    def download_one(g):
        result = download_pbp_for_game(g["game_id"], rate_limiter)
        with done_lock:
            done_count[0] += 1
            if done_count[0] % 200 == 0 or done_count[0] == total:
                print(f"  PBP: {done_count[0]}/{total} downloaded")
        return result

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(download_one, g): g for g in to_download}
        for fut in as_completed(futures):
            gid, ht = fut.result()
            halftime_scores[gid] = ht

    return halftime_scores


# ---------------------------------------------------------------------------
# Phase 4: Write to workbook
# ---------------------------------------------------------------------------

def get_existing_game_ids(ws) -> set:
    """Scan column B of Game_Log, return set of all present game_ids."""
    ids = set()
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val not in (None, ""):
            ids.add(str(val).strip())
    return ids


def first_empty_row(ws) -> int:
    """Return the first row in column B that is empty."""
    row = 2
    while ws.cell(row=row, column=2).value not in (None, ""):
        row += 1
    return row


def write_rows_to_workbook(games: list, halftime_scores: dict, dry_run: bool):
    """Load workbook, append new rows, save."""
    wb = load_workbook(str(NCAAM_RESULTS_XLSX))
    ws = wb["Game_Log"]

    existing_ids = get_existing_game_ids(ws)
    next_row = first_empty_row(ws)
    print(f"  Existing rows in workbook: {len(existing_ids)}, next empty row: {next_row}")

    market_lines_cache = load_market_lines()

    added = 0
    skipped_existing = 0
    skipped_no_ht = 0

    for g in games:
        game_id = g["game_id"]

        if game_id in existing_ids:
            skipped_existing += 1
            continue

        ht = halftime_scores.get(game_id)
        if ht is None:
            skipped_no_ht += 1
            continue

        away_ht, home_ht = ht
        halftime_total = away_ht + home_ht

        # Basic sanity: halftime scores should be 20-70 per team
        if not (10 <= away_ht <= 80 and 10 <= home_ht <= 80):
            skipped_no_ht += 1
            continue

        halftime_score_str = f"{away_ht}-{home_ht}"

        # Build team stats and PBP features
        stats = load_team_stats(g["date"])
        home_avg_scored, home_avg_allowed = resolve_team_stats(stats, g["home_seo"])
        away_avg_scored, away_avg_allowed = resolve_team_stats(stats, g["away_seo"])
        pbp_features = load_game_pbp_features(DATA_ROOT, game_id)

        # Derive pace profile label from halftime total
        pace_profile = pace_label_from_halftime_total(float(halftime_total))

        # Compute actual outcomes
        actual_total = g["away_score"] + g["home_score"]
        actual_margin = g["home_score"] - g["away_score"]  # positive = home won
        actual_2h = actual_total - halftime_total
        actual_winner = g["home_seo"] if actual_margin > 0 else g["away_seo"]

        if not dry_run:
            row = next_row
            ws.cell(row=row, column=1).value = g["date"]            # A: Date
            ws.cell(row=row, column=2).value = game_id              # B: GameID
            ws.cell(row=row, column=3).value = g["away_seo"]        # C: Away
            ws.cell(row=row, column=4).value = g["home_seo"]        # D: Home
            ws.cell(row=row, column=5).value = halftime_score_str   # E: HalftimeScore
            ws.cell(row=row, column=6).value = pace_profile          # F: PaceProfile
            # Cols 7-11 (G-K): Pred* — leave blank for update_all_predictions_ml.py
            ws.cell(row=row, column=12).value = actual_winner        # L: ActualWinner
            ws.cell(row=row, column=13).value = actual_margin        # M: ActualMargin
            ws.cell(row=row, column=14).value = actual_2h            # N: Actual2H
            ws.cell(row=row, column=15).value = actual_total         # O: ActualTotal
            # P-T: WinnerCorrect, TwoH_Error, Total_Error, narrow ranges — leave blank
            next_row += 1
            added += 1

            # Checkpoint save every 250 rows
            if added % 250 == 0:
                wb.save(str(NCAAM_RESULTS_XLSX))
                print(f"  Saved checkpoint: {added} rows written so far")
        else:
            added += 1  # count for dry-run reporting

    if not dry_run and added > 0:
        wb.save(str(NCAAM_RESULTS_XLSX))

    print(f"\n  Rows added:                     {added}")
    print(f"  Skipped (already in workbook):  {skipped_existing}")
    print(f"  Skipped (no halftime PBP):      {skipped_no_ht}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Backfill prior NCAAB season data into NCAAM Results.xlsx"
    )
    parser.add_argument(
        "--season", default=DEFAULT_SEASON,
        choices=list(SEASON_DATE_RANGES.keys()),
        help=f"Season to backfill (default: {DEFAULT_SEASON})",
    )
    parser.add_argument(
        "--workers", type=int, default=DEFAULT_WORKERS,
        help="Number of parallel PBP download workers",
    )
    parser.add_argument(
        "--max-rps", type=float, default=DEFAULT_MAX_RPS,
        help="Max HTTP requests per second across all workers",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Download data and compute everything, but don't write to workbook",
    )
    parser.add_argument(
        "--skip-pbp", action="store_true",
        help="Skip PBP download phase, use only already-cached files",
    )
    parser.add_argument(
        "--phases", default="1234",
        help="Which phases to run, e.g. --phases 12 to only run scoreboard+baselines",
    )
    args = parser.parse_args()

    cache_root = Path(DATA_ROOT) / "cache" / "scoreboard_daily"
    baselines_dir = Path(DATA_ROOT) / "processed" / "baselines"

    print(f"\n{'='*60}")
    print(f"  Backfilling {args.season} season")
    print(f"  Date range: {SEASON_DATE_RANGES[args.season][0]} to {SEASON_DATE_RANGES[args.season][1]}")
    print(f"  Workers: {args.workers}   Max RPS: {args.max_rps}")
    print(f"  Dry run: {args.dry_run}   Skip PBP: {args.skip_pbp}")
    print(f"{'='*60}\n")

    # ── Phase 1: Scan scoreboard ────────────────────────────────────────────
    if "1" in args.phases:
        print("=== Phase 1: Scanning scoreboard ===")
        games = scan_season_scoreboard(args.season, cache_root)
    else:
        print("=== Phase 1: SKIPPED — loading from inventory ===")
        # Reconstruct from scoreboard cache
        games = scan_season_scoreboard(args.season, cache_root)

    if not games:
        print("No games found. Exiting.")
        return

    # ── Phase 2: Baselines ──────────────────────────────────────────────────
    if "2" in args.phases:
        print("\n=== Phase 2: Building last4 baselines ===")
        build_baselines_for_season(games, baselines_dir)
    else:
        print("\n=== Phase 2: SKIPPED ===")

    # ── Phase 3: Download PBP ───────────────────────────────────────────────
    if "3" in args.phases and not args.skip_pbp:
        print(f"\n=== Phase 3: Downloading PBP for {len(games)} games ===")
        halftime_scores = download_all_pbp(games, args.workers, args.max_rps)
        ht_found = sum(1 for v in halftime_scores.values() if v is not None)
        print(f"  Halftime scores extracted: {ht_found}/{len(games)}")
    else:
        if args.skip_pbp:
            print("\n=== Phase 3: SKIPPED (--skip-pbp) — loading cached only ===")
        else:
            print("\n=== Phase 3: SKIPPED ===")
        halftime_scores = {}
        for g in games:
            out_path = pbp_output_path(g["game_id"])
            if out_path.exists():
                try:
                    data = json.loads(out_path.read_text(encoding="utf-8"))
                    halftime_scores[g["game_id"]] = extract_halftime_score(
                        data.get("first_half_plays", [])
                    )
                except Exception:
                    halftime_scores[g["game_id"]] = None

    # ── Phase 4: Write workbook ─────────────────────────────────────────────
    if "4" in args.phases:
        action = "DRY RUN (no writes)" if args.dry_run else "Writing to workbook"
        print(f"\n=== Phase 4: {action} ===")
        write_rows_to_workbook(games, halftime_scores, dry_run=args.dry_run)
    else:
        print("\n=== Phase 4: SKIPPED ===")

    print(f"\n{'='*60}")
    if args.dry_run:
        print("  Dry run complete.  Re-run without --dry-run to write workbook.")
    else:
        print("  Backfill complete.  Next steps:")
        print("    python Scripts/tune_totals_spreads_v1.py")
        print("    python Scripts/update_all_predictions_ml.py")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
