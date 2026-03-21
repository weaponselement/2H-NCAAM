import os
import re
import glob
import json
from collections import defaultdict
from openpyxl import load_workbook

DATA_ROOT = r"C:\NCAA Model\data"
WORKBOOK = r"C:\NCAA Model\logs\NCAAM Results.xlsx"
SHEET_NAME = "Game_Log"

SLATES_DIR = os.path.join(DATA_ROOT, "processed", "slates")
SELECTED_DIR = os.path.join(DATA_ROOT, "processed", "selected_games")
BASELINES_DIR = os.path.join(DATA_ROOT, "processed", "baselines")
REPORTS_DIR = os.path.join(DATA_ROOT, "processed", "reports")
PBP_LIVE_DIR = os.path.join(DATA_ROOT, "raw", "pbp_live")


def extract_date(name):
    m = re.search(r"(\d{4}-\d{2}-\d{2})", name)
    return m.group(1) if m else None


def latest_report_path(game_id):
    matches = glob.glob(os.path.join(REPORTS_DIR, f"feature_report_v5_test_{game_id}_*.json"))
    matches.sort(key=os.path.getmtime, reverse=True)
    return matches[0] if matches else None


def latest_first_half_path(game_id):
    matches = glob.glob(os.path.join(PBP_LIVE_DIR, str(game_id), "pbp_first_half_*.json"))
    matches.sort(key=os.path.getmtime, reverse=True)
    return matches[0] if matches else None


def main():
    slate_dates = {
        extract_date(os.path.basename(p))
        for p in glob.glob(os.path.join(SLATES_DIR, "slate_d1_*.json"))
    }
    selected_dates = {
        extract_date(os.path.basename(p))
        for p in glob.glob(os.path.join(SELECTED_DIR, "selected_games_*.json"))
    }
    baseline_dates = {
        extract_date(os.path.basename(p))
        for p in glob.glob(os.path.join(BASELINES_DIR, "last*.json"))
    }

    slate_dates.discard(None)
    selected_dates.discard(None)
    baseline_dates.discard(None)

    print("=" * 70)
    print("DATE-LEVEL COVERAGE")
    print("=" * 70)
    print(f"Slate dates       : {len(slate_dates)}")
    print(f"Selected dates    : {len(selected_dates)}")
    print(f"Baseline dates    : {len(baseline_dates)}")
    print("")

    missing_selected = sorted(slate_dates - selected_dates)
    missing_baselines = sorted(slate_dates - baseline_dates)

    print("Missing selected_games dates:")
    if missing_selected:
        for d in missing_selected:
            print(f"  - {d}")
    else:
        print("  none")

    print("")
    print("Missing baseline dates:")
    if missing_baselines:
        for d in missing_baselines:
            print(f"  - {d}")
    else:
        print("  none")

    print("")
    print("=" * 70)
    print("WORKBOOK GAME COVERAGE")
    print("=" * 70)

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET_NAME]

    total = 0
    missing_pbp = []
    missing_report = []
    missing_both = []
    by_date = defaultdict(lambda: {"rows": 0, "missing_pbp": 0, "missing_report": 0})

    for row in range(2, ws.max_row + 1):
        game_id = ws[f"B{row}"].value
        date_val = ws[f"A{row}"].value

        if game_id in (None, ""):
            continue

        gid = str(game_id).strip()
        if not gid:
            continue

        date_str = str(date_val).strip() if date_val not in (None, "") else "UNKNOWN"

        total += 1
        by_date[date_str]["rows"] += 1

        pbp_path = latest_first_half_path(gid)
        report_path = latest_report_path(gid)

        has_pbp = pbp_path is not None
        has_report = report_path is not None

        if not has_pbp and not has_report:
            missing_both.append((date_str, gid))
            by_date[date_str]["missing_pbp"] += 1
            by_date[date_str]["missing_report"] += 1
        elif not has_pbp:
            missing_pbp.append((date_str, gid))
            by_date[date_str]["missing_pbp"] += 1
        elif not has_report:
            missing_report.append((date_str, gid))
            by_date[date_str]["missing_report"] += 1

    print(f"Workbook rows checked : {total}")
    print(f"Missing halftime PBP  : {len(missing_pbp)}")
    print(f"Missing v5 report     : {len(missing_report)}")
    print(f"Missing both          : {len(missing_both)}")

    print("")
    print("Dates with any missing coverage:")
    bad_dates = []
    for d in sorted(by_date.keys()):
        info = by_date[d]
        if info["missing_pbp"] or info["missing_report"]:
            bad_dates.append(d)
            print(
                f"  {d} -> rows={info['rows']}, "
                f"missing_pbp={info['missing_pbp']}, "
                f"missing_report={info['missing_report']}"
            )

    if not bad_dates:
        print("  none")

    print("")
    print("First 25 missing halftime PBP entries:")
    for d, gid in missing_pbp[:25]:
        print(f"  - {d} | {gid}")

    print("")
    print("First 25 missing v5 report entries:")
    for d, gid in missing_report[:25]:
        print(f"  - {d} | {gid}")

    print("")
    print("First 25 missing both entries:")
    for d, gid in missing_both[:25]:
        print(f"  - {d} | {gid}")


if __name__ == "__main__":
    main()