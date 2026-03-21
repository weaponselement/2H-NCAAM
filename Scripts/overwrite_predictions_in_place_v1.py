import glob
import json
import os
from openpyxl import load_workbook
from paths import DATA_DIR, NCAAM_RESULTS_XLSX

RESULTS_XLSX = str(NCAAM_RESULTS_XLSX)
DATA_ROOT = str(DATA_DIR)
SHEET_NAME = "Game_Log"


def latest_report_path(game_id: str):
    pattern = os.path.join(
        DATA_ROOT,
        "processed",
        "reports",
        f"feature_report_v5_test_{game_id}_*.json"
    )
    matches = glob.glob(pattern)

    if not matches:
        return None

    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return matches[0]


def load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def halftime_score_from_report(report: dict):
    halftime = report.get("halftime_score", {}) or {}
    away_ht = halftime.get("away")
    home_ht = halftime.get("home")

    if away_ht is None or home_ht is None:
        return ""

    return f"{away_ht}-{home_ht}"


def extract_projection_range(projection: dict, key_path):
    obj = projection
    for key in key_path:
        if not isinstance(obj, dict):
            return ""
        obj = obj.get(key, {})
    if isinstance(obj, str):
        return obj
    return ""


def main():
    if not os.path.exists(RESULTS_XLSX):
        print(f"Workbook not found: {RESULTS_XLSX}")
        return

    wb = load_workbook(RESULTS_XLSX)
    ws = wb[SHEET_NAME]

    updated = 0
    missing_reports = []

    for row in range(2, ws.max_row + 1):
        game_id = ws.cell(row=row, column=2).value

        if game_id in (None, ""):
            continue

        game_id = str(game_id).strip()
        if not game_id:
            continue

        report_path = latest_report_path(game_id)
        if not report_path:
            missing_reports.append(game_id)
            continue

        report = load_json(report_path)

        teams = report.get("teams", {}) or {}
        away = normalize_text((teams.get("away") or {}).get("seo", ""))
        home = normalize_text((teams.get("home") or {}).get("seo", ""))

        game_state = report.get("game_state", {}) or {}
        projection = report.get("projection", {}) or {}

        halftime_score = halftime_score_from_report(report)
        pred_margin = normalize_text(projection.get("winner_margin_range", ""))
        pred_2h = extract_projection_range(
            projection,
            ["second_half_points_projection", "range"]
        )
        pred_total = extract_projection_range(
            projection,
            ["full_game_total_projection", "range"]
        )

        ws[f"A{row}"] = normalize_text(report.get("run_date", ws[f"A{row}"].value))
        ws[f"B{row}"] = game_id
        if away:
            ws[f"C{row}"] = away
        if home:
            ws[f"D{row}"] = home
        if halftime_score:
            ws[f"E{row}"] = halftime_score

        ws[f"F{row}"] = normalize_text(game_state.get("pace_profile", ""))
        ws[f"G{row}"] = normalize_text(projection.get("winner_projection", ""))
        ws[f"H{row}"] = pred_margin
        ws[f"I{row}"] = pred_2h
        ws[f"J{row}"] = pred_total
        ws[f"K{row}"] = normalize_text(projection.get("confidence", ""))

        updated += 1
        print(f"Updated row {row}: game {game_id}")

    wb.save(RESULTS_XLSX)

    print("")
    print("OVERWRITE COMPLETE")
    print(f"Workbook: {RESULTS_XLSX}")
    print(f"Rows updated: {updated}")
    print(f"Missing reports: {len(missing_reports)}")

    if missing_reports:
        print("GameIDs missing reports:")
        for gid in missing_reports:
            print(f"  - {gid}")


if __name__ == "__main__":
    main()