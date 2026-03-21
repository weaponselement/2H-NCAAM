import argparse
import glob
import json
import os
import requests
from openpyxl import load_workbook
from paths import DATA_DIR

API_BASE = "https://ncaa-api.henrygd.me"
HEADERS = {"User-Agent": "Mozilla/5.0"}
DATA_ROOT = str(DATA_DIR)


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def fetch_scoreboard(date_str):
    yyyy, mm, dd = date_str.split("-")
    url = f"{API_BASE}/scoreboard/basketball-men/d1/{yyyy}/{mm}/{dd}"
    r = requests.get(url, headers=HEADERS, timeout=45)
    r.raise_for_status()
    return r.json()


def latest_report_path(game_id):
    pattern = os.path.join(DATA_ROOT, "processed", "reports", f"feature_report_v4_{game_id}_*.json")
    matches = glob.glob(pattern)

    if not matches:
        return None

    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return matches[0]


def parse_range(text):
    if not text or "-" not in text:
        return None, None

    a, b = text.split("-", 1)

    try:
        return int(a.strip()), int(b.strip())
    except:
        return None, None


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    parser.add_argument("--workbook", required=True)

    args = parser.parse_args()

    selected_games_path = os.path.join(
        DATA_ROOT,
        "processed",
        "selected_games",
        f"selected_games_{args.date}.json"
    )

    selected_games = load_json(selected_games_path)

    scoreboard = fetch_scoreboard(args.date)

    wb = load_workbook(args.workbook)
    ws = wb["Game Log v2"]

    row = ws.max_row + 1

    for game in selected_games:

        game_id = str(game.get("gameID") or game.get("game_id"))

        report_path = latest_report_path(game_id)

        if not report_path:
            continue

        report = load_json(report_path)

        projection = report.get("projection", {})

        away_team = game.get("away_short")
        home_team = game.get("home_short")

        away_ht = report.get("halftime_score", {}).get("away")
        home_ht = report.get("halftime_score", {}).get("home")

        pace_profile = report.get("game_state", {}).get("pace_profile")

        confidence = projection.get("confidence")
        winner_projection = projection.get("winner_projection")

        margin_low, margin_high = parse_range(projection.get("winner_margin_range"))

        sh_low, sh_high = parse_range(
            projection.get("second_half_points_projection", {}).get("range")
        )

        fg_low, fg_high = parse_range(
            projection.get("full_game_total_projection", {}).get("range")
        )

        final_game = None

        for item in scoreboard.get("games", []):
            if str(item["game"]["gameID"]) == game_id:
                final_game = item["game"]
                break

        if not final_game:
            continue

        away_final = int(final_game["away"]["score"])
        home_final = int(final_game["home"]["score"])

        ws[f"A{row}"] = args.date
        ws[f"B{row}"] = int(game_id)
        ws[f"C{row}"] = away_team
        ws[f"D{row}"] = home_team
        ws[f"E{row}"] = away_ht
        ws[f"F{row}"] = home_ht
        ws[f"G{row}"] = pace_profile
        ws[f"H{row}"] = confidence
        ws[f"I{row}"] = winner_projection
        ws[f"J{row}"] = margin_low
        ws[f"K{row}"] = margin_high
        ws[f"L{row}"] = sh_low
        ws[f"M{row}"] = sh_high
        ws[f"N{row}"] = fg_low
        ws[f"O{row}"] = fg_high
        ws[f"P{row}"] = away_final
        ws[f"Q{row}"] = home_final

        row += 1

    wb.save(args.workbook)

    print("Postgame export complete")


if __name__ == "__main__":
    main()