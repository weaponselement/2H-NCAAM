"""
ncaab_historical_lines_covers_v1.py
------------------------------------
Scrape NCAAB closing lines from Covers.com for dates in NCAAM Results.xlsx,
match each game to a workbook GameID, populate canonical_lines.csv.

Usage:
  python Scripts/ncaab_historical_lines_covers_v1.py [--date 2026-02-18] [--all] [--dry-run]

Options:
  --date YYYY-MM-DD    Scrape one date only
  --all                Scrape all dates in the workbook
  --dry-run            Scrape but don't write canonical_lines.csv
  --since YYYY-MM-DD   Only scrape dates >= this (useful for incremental updates)
"""

import argparse
import csv
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

# Force UTF-8 output on Windows to avoid charmap encoding errors
if sys.platform == "win32":
    import codecs
    sys.stdout = codecs.getwriter("utf-8")(sys.stdout.buffer, errors="replace")
    sys.stderr = codecs.getwriter("utf-8")(sys.stderr.buffer, errors="replace")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = REPO_ROOT / "logs" / "NCAAM Results.xlsx"
CANONICAL_PATH = REPO_ROOT / "data" / "processed" / "market_lines" / "canonical_lines.csv"

COVERS_URL = "https://www.covers.com/sports/ncaab/matchups?selectedDate={date}"
SLEEP_BETWEEN_REQUESTS = 2.0  # be polite

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

CANONICAL_FIELDS = [
    "game_id", "date", "away_seo", "home_seo",
    "spread_home", "spread_away", "ml_home", "ml_away",
    "total_game", "total_2h", "sgo_event_id", "source_file", "staged_timestamp",
]


# ---------------------------------------------------------------------------
# HTML Parsing
# ---------------------------------------------------------------------------

def unescape_html(text: str) -> str:
    """Minimal HTML entity decode."""
    return (
        text.replace("&#x2B;", "+")
            .replace("&#x2D;", "-")
            .replace("&#43;", "+")
            .replace("&#45;", "-")
            .replace("&amp;", "&")
            .replace("&nbsp;", " ")
            .replace("&#x27;", "'")
    )


def split_game_blocks(html: str) -> list:
    """
    Split the full page HTML into per-game article elements.
    Each game is in <article id="ncaab-{coverId}" ...>.
    """
    pattern = re.compile(r'<article\s+id="ncaab-\d+"')
    positions = [m.start() for m in pattern.finditer(html)]
    if not positions:
        return []

    blocks = []
    for i, start in enumerate(positions):
        end = positions[i + 1] if i + 1 < len(positions) else len(html)
        blocks.append(html[start:end])
    return blocks


def parse_game_block(block: str) -> dict | None:
    """
    Parse one game block and return dict:
      away_slug, home_slug, away_score, home_score,
      spread_home, total_game, covers_game_id
    Returns None if essential data is missing.
    """
    result = {}

    # -- Covers game ID from article element ----------------------------------
    art_m = re.search(r'<article\s+id="ncaab-(\d+)"', block)
    if art_m:
        result["covers_game_id"] = art_m.group(1)

    # -- Team names and shortnames from data attributes -----------------------
    away_name_m = re.search(r'data-away-team-name=(.+?)\s+data-away-team-displayname=', block)
    home_name_m = re.search(r'data-home-team-name=(.+?)\s+data-home-team-displayname=', block)
    away_short_m = re.search(r'data-away-team-shortname=(\S+)', block)
    home_short_m = re.search(r'data-home-team-shortname=(\S+)', block)
    header_away_name = away_name_m.group(1).strip().lower() if away_name_m else ""
    header_home_name = home_name_m.group(1).strip().lower() if home_name_m else ""
    away_shortname = away_short_m.group(1).lower() if away_short_m else ""
    home_shortname = home_short_m.group(1).lower() if home_short_m else ""

    # Fallback: parse "AWAY @ HOME" from gamebox-header strong tag
    if not header_away_name or not header_home_name:
        header_m = re.search(
            r'id="gamebox-header"[^>]*>.*?<strong class="text-uppercase">([^<]+)</strong>',
            block, re.DOTALL
        )
        if header_m:
            matchup_raw = unescape_html(header_m.group(1)).strip()
            if " @ " in matchup_raw:
                header_away_name, header_home_name = matchup_raw.split(" @ ", 1)
                header_away_name = header_away_name.strip().lower()
                header_home_name = header_home_name.strip().lower()

    # -- Team slugs (from team anchor hrefs) ----------------------------------
    slug_pattern = re.compile(
        r'gamebox-team-anchor[^"]*"[^>]+href="/sport/basketball/ncaab/teams/main/([^"]+)"'
    )
    slugs = slug_pattern.findall(block)
    if len(slugs) >= 2:
        result["away_slug"] = slugs[0]
        result["home_slug"] = slugs[1]
    elif header_away_name and header_home_name:
        # Fall back to data-attribute names as slugs (they're lowercase, space-separated)
        result["away_slug"] = header_away_name.replace(" ", "-")
        result["home_slug"] = header_home_name.replace(" ", "-")
    else:
        return None  # Can't proceed without team slugs

    # -- Scores (mobile view in <strong class="team-score away ..."> etc.) ----
    away_score_m = re.search(r'"team-score away[^"]*">(\d+)</strong>', block)
    home_score_m = re.search(r'"team-score home[^"]*">(\d+)</strong>', block)
    if away_score_m and home_score_m:
        result["away_score"] = int(away_score_m.group(1))
        result["home_score"] = int(home_score_m.group(1))

    # -- Spread and total from summary-box text --------------------------------
    # Normal case: "Team covered the spread of +X.X. The total score of Y was over/under Z.Z."
    summary_m = re.search(
        r'<p class="m-0 summary-box[^"]*">\s*([^<]+)\s+covered the spread of\s+<strong>([\+\-]?[\d.]+)</strong>'
        r'\.\s+The total score of (\d+) was\s+<strong>(over|under) ([\d.]+)</strong>',
        block, re.DOTALL
    )
    if summary_m:
        covering_team_text = unescape_html(summary_m.group(1)).strip()
        covered_spread_val = float(summary_m.group(2))
        actual_total = int(summary_m.group(3))
        total_line = float(summary_m.group(5))
        result["actual_total"] = actual_total
        result["total_game"] = total_line

        # Determine if covering team is away or home
        # Compare covering_team_text (e.g. "Lafayette") to header names
        # Extract covering team abbreviation from "Cover By" block span
        cover_by_m = re.search(
            r'Cover By.*?<span>\s*([A-Z0-9\-]+)\s+(?:&#x2B;|&#x2D;|[\+\-])',
            block
        )
        cover_abv = cover_by_m.group(1).lower() if cover_by_m else ""

        # Priority: exact shortname match (most reliable, no ambiguity)
        if cover_abv and cover_abv == away_shortname and cover_abv != home_shortname:
            result["spread_home"] = -covered_spread_val
            result["spread_away"] = covered_spread_val
            return result
        if cover_abv and cover_abv == home_shortname and cover_abv != away_shortname:
            result["spread_home"] = covered_spread_val
            result["spread_away"] = -covered_spread_val
            return result

        cov_lower = covering_team_text.lower()
        away_lower = header_away_name.lower()
        home_lower = header_home_name.lower()

        is_away_cover = (
            _name_matches(cov_lower, away_lower)
            or (cover_abv and away_shortname.startswith(cover_abv))
        )
        is_home_cover = (
            _name_matches(cov_lower, home_lower)
            or (cover_abv and home_shortname.startswith(cover_abv))
        )

        if is_away_cover and not is_home_cover:
            result["spread_home"] = -covered_spread_val
            result["spread_away"] = covered_spread_val
        elif is_home_cover and not is_away_cover:
            result["spread_home"] = covered_spread_val
            result["spread_away"] = -covered_spread_val
        else:
            result["spread_home"] = None
            result["_spread_ambiguous"] = (
                f"covering='{covering_team_text}'({cover_abv}) away='{header_away_name}'({away_shortname}) "
                f"home='{header_home_name}'({home_shortname}) spread={covered_spread_val}"
            )
        return result

    # -- Push on total: "spread covered, total pushed" -------------------------
    push_total_m = re.search(
        r'<p class="m-0 summary-box[^"]*">\s*([^<]+)\s+covered the spread of\s+<strong>([\+\-]?[\d.]+)</strong>'
        r'\.\s+The total score of (\d+) pushed the pre-game total of ([\d.]+)',
        block, re.DOTALL
    )
    if push_total_m:
        covering_team_text = unescape_html(push_total_m.group(1)).strip()
        covered_spread_val = float(push_total_m.group(2))
        actual_total = int(push_total_m.group(3))
        total_line = float(push_total_m.group(4))
        result["actual_total"] = actual_total
        result["total_game"] = total_line
        cov_lower = covering_team_text.lower()
        away_lower = header_away_name.lower()
        home_lower = header_home_name.lower()
        cover_by_m = re.search(
            r'Cover By.*?<span>\s*([A-Z0-9\-]+)\s+(?:&#x2B;|&#x2D;|[\+\-])',
            block
        )
        cover_abv = cover_by_m.group(1).lower() if cover_by_m else ""
        is_away = (_name_matches(cov_lower, away_lower) or (cover_abv and cover_abv == away_shortname))
        is_home = (_name_matches(cov_lower, home_lower) or (cover_abv and cover_abv == home_shortname))
        if is_away and not is_home:
            result["spread_home"] = -covered_spread_val
            result["spread_away"] = covered_spread_val
        elif is_home and not is_away:
            result["spread_home"] = covered_spread_val
            result["spread_away"] = -covered_spread_val
        else:
            result["spread_home"] = None
        return result

    # -- Push on spread:  team "pushed the spread" (no clear cover) -----------
    spread_push_m = re.search(
        r'<p class="m-0 summary-box[^"]*">\s*(?:The spread pushed[^.]*|[^<]+pushed the spread[^.]*)\.',
        block, re.DOTALL
    )
    ou_m = re.search(r'<p class="m-0 summary-box[^"]*">.*?<strong>(over|under) ([\d.]+)</strong>', block, re.DOTALL)
    if spread_push_m and ou_m:
        # Total still present
        result["spread_home"] = None
        result["total_game"] = float(ou_m.group(2))
        return result

    # -- No odds data (game not yet played or data missing) -------------------
    # Just return what we have (team slugs, maybe scores)
    if "away_slug" in result:
        return result
    return None


def _name_matches(cov_lower: str, team_header_lower: str) -> bool:
    """
    Check if the covering team name matches a header team name.
    'cov_lower' is from the summary text (e.g. "lafayette"),
    'team_header_lower' is from the header  (e.g. "lafayette").
    """
    if cov_lower == team_header_lower:
        return True
    # Covering text might be prefixed (e.g. "illinois" vs "illinois-chicago")
    if team_header_lower.startswith(cov_lower):
        return True
    if cov_lower.startswith(team_header_lower):
        return True
    # Partial word overlap (covers 2+ tokens)
    cov_parts = set(cov_lower.split())
    hdl_parts = set(team_header_lower.split())
    if cov_parts & hdl_parts:
        return True
    return False


def scrape_date(date_str: str) -> list:
    """Fetch and parse Covers.com matchups for a given date. Returns list of game dicts."""
    url = COVERS_URL.format(date=date_str)
    try:
        r = requests.get(url, headers=HTTP_HEADERS, timeout=15)
        r.raise_for_status()
    except Exception as exc:
        print(f"  [ERROR] {date_str}: {exc}")
        return []

    blocks = split_game_blocks(r.text)
    games = []
    for block in blocks:
        parsed = parse_game_block(block)
        if parsed and "away_slug" in parsed:
            parsed["date"] = date_str
            games.append(parsed)
    return games


# ---------------------------------------------------------------------------
# Workbook loading
# ---------------------------------------------------------------------------

def safe_float(value):
    try:
        return float(value)
    except Exception:
        return None

def load_workbook_games() -> list:
    """Load workbook Game_Log rows (those with ActualTotal)."""
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Game_Log"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        row = dict(zip(headers, r))
        if row.get("ActualTotal") is None:
            continue
        date = str(row.get("Date", "")).split(" ")[0]
        actual_total = safe_float(row.get("ActualTotal"))
        if actual_total is None:
            continue
        actual_margin = safe_float(row.get("ActualMargin", 0) or 0)
        try:
            if actual_margin is None:
                raise ValueError("missing margin")
            # margin = home - away, total = home + away
            # home = (total + margin) / 2
            home_score_f = (actual_total + actual_margin) / 2.0
            away_score_f = actual_total - home_score_f
        except Exception:
            home_score_f = None
            away_score_f = None
        data.append({
            "game_id": str(row.get("GameID", "")).strip(),
            "date": date,
            "away_seo": str(row.get("Away", "")).strip(),
            "home_seo": str(row.get("Home", "")).strip(),
            "actual_total": actual_total,
            "actual_margin": actual_margin if actual_margin is not None else 0.0,
            "home_score_f": home_score_f,
            "away_score_f": away_score_f,
        })
    return data


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def slug_from_covers(covers_slug: str) -> str:
    """
    Convert Covers team slug to an SDIO-like form for matching.
    Covers: 'lafayette-leopards', 'penn-state-nittany-lions', 'holy-cross-crusaders'
    SDIO:   'lafayette',           'penn-st',                  'holy-cross'

    Strategy: strip common mascot suffixes, apply state abbrevation rule.
    Also consult a known override table.
    """
    # Strip NCAAB suffix (if any)
    slug = covers_slug.lower().rstrip("-").strip()

    # Known exact overrides (covers_slug → sdio_slug)
    OVERRIDES = {
        "penn-state-nittany-lions": "penn-st",
        "michigan-state-spartans": "michigan-st",
        "ohio-state-buckeyes": "ohio-st",
        "iowa-state-cyclones": "iowa-st",
        "kansas-state-wildcats": "kansas-st",
        "florida-state-seminoles": "florida-st",
        "mississippi-state-bulldogs": "mississippi-st",
        "north-carolina-state-wolfpack": "nc-state",
        "nc-state-wolfpack": "nc-state",
        "murray-state-racers": "murray-st",
        "south-dakota-state-jackrabbits": "south-dakota-st",
        "north-dakota-state-bison": "north-dakota-st",
        "middle-tennessee-blue-raiders": "middle-tenn",
        "east-tennessee-state-buccaneers": "east-tenn-st",
        "southeast-missouri-state-redhawks": "se-missouri-st",
        "southern-illinois-salukis": "southern-ill",
        "sam-houston-bearkats": "sam-houston",
        "san-jose-state-spartans": "san-jose-st",
        "utah-state-aggies": "utah-st",
        "boise-state-broncos": "boise-st",
        "fresno-state-bulldogs": "fresno-st",
        "colorado-state-rams": "colorado-st",
        "portland-state-vikings": "portland-st",
        "south-carolina-state-bulldogs": "south-carolina-st",
        "appalachian-state-mountaineers": "app-st",
        "georgia-state-panthers": "georgia-st",
        "kennesaw-state-owls": "kennesaw-st",
        "jacksonville-state-gamecocks": "jacksonville-st",
        "north-carolina-central-eagles": "nc-central",
        "mcneese-state-cowboys": "mcneese-st",
        "north-carolina-a-t-aggies": "nc-at",
        "north-carolina-at-aggies": "nc-at",
        "north-carolina-a&t-aggies": "nc-at",
        "uic-flames": "ill-chicago",
        "illinois-chicago-flames": "ill-chicago",
        "byu-cougars": "byu",
        "brigham-young-cougars": "byu",
        "uconn-huskies": "connecticut",
        "connecticut-huskies": "connecticut",
        "miami-fl-hurricanes": "miami-fl",
        "miami-oh-redhawks": "miami-oh",
        "miami-redhawks": "miami-oh",
        "miami-hurricanes": "miami-fl",
        "usc-trojans": "southern-california",
        "southern-california-trojans": "southern-california",
        "tcu-horned-frogs": "tcu",
        "texas-christian-horned-frogs": "tcu",
        "smu-mustangs": "smu",
        "southern-methodist-mustangs": "smu",
        "ucf-knights": "ucf",
        "central-florida-knights": "ucf",
        "fiu-panthers": "fiu",
        "florida-international-panthers": "fiu",
        "fgcu-eagles": "fgcu",
        "florida-gulf-coast-eagles": "fgcu",
        "lsu-tigers": "lsu",
        "louisiana-state-tigers": "lsu",
        "ole-miss-rebels": "ole-miss",
        "mississippi-rebels": "ole-miss",
        "la-monroe-warhawks": "la-monroe",
        "louisiana-monroe-warhawks": "la-monroe",
        "neb-omaha-mavericks": "neb-omaha",
        "omaha-mavericks": "neb-omaha",
        "ut-martin-skyhawks": "ut-martin",
        "tennessee-martin-skyhawks": "ut-martin",
        "fla-atlantic-owls": "fla-atlantic",
        "florida-atlantic-owls": "fla-atlantic",
        "queens-university-royals": "queens-nc",
        "queens-nc-royals": "queens-nc",
        "vmi-keydets": "vmi",
        "virginia-military-keydets": "vmi",
        "uc-irvine-anteaters": "uc-irvine",
        "uc-santa-barbara-gauchos": "uc-santa-barbara",
        "uc-san-diego-tritons": "uc-san-diego",
        "uc-davis-aggies": "uc-davis",
        "uc-riverside-highlanders": "uc-riverside",
        "cal-poly-mustangs": "cal-poly",
        "cal-state-fullerton-titans": "cal-st-fullerton",
        "cal-state-bakersfield-roadrunners": "cal-st-bakersfield",
        "csun-matadors": "cal-st-northridge",
        "cal-state-northridge-matadors": "cal-st-northridge",
        "cal-state-long-beach-beach": "long-beach-st",
        "long-beach-state-beach": "long-beach-st",
        "east-carolina-pirates": "east-carolina",
        "wichita-state-shockers": "wichita-st",
        "cleveland-state-vikings": "cleveland-st",
        "youngstown-state-penguins": "youngstown-st",
        "seton-hall-pirates": "seton-hall",
        "st-johns-red-storm": "st-johns-ny",
        "st.-john-s-red-storm": "st-johns-ny",
        "georgia-southern-eagles": "ga-southern",
        "loyola-chicago-ramblers": "loyola-chicago",
        "loyola-maryland-greyhounds": "loyola-maryland",
        "saint-joseph-s-hawks": "saint-josephs",
        "saint-josephs-hawks": "saint-josephs",
        "saint-marys-gaels": "saint-marys-ca",
        "saint-mary-s-gaels": "saint-marys-ca",
        "saint-peter-s-peacocks": "st-peters",
        "st.-bonaventure-bonnies": "st-bonaventure",
        "eastern-kentucky-colonels": "eastern-ky",
        "florida-a-m-rattlers": "florida-am",
        "alabama-a-m-bulldogs": "alabama-am",
        "arkansas-pine-bluff-golden-lions": "ark-pine-bluff",
        "purdue-fort-wayne-mastodons": "purdue-fort-wayne",
        "indiana-state-sycamores": "indiana-st",
        "utah-valley-wolverines": "utah-valley",
        "northern-kentucky-norse": "northern-ky",
        "west-georgia-wolves": "west-ga",
        "south-florida-bulls": "south-fla",
        "southeastern-louisiana-lions": "southeastern-la",
        "bellarmine-knights": "bellarmine",
        "seattle-redhawks": "seattle",
        "seattle-u-redhawks": "seattle",
        "pacific-tigers": "pacific",
        "texas-am-aggies": "texas-am",
        "texas-a-m-aggies": "texas-am",
        "north-alabama-lions": "north-alabama",
        "western-kentucky-hilltoppers": "western-ky",
        "western-carolina-catamounts": "western-caro",
    }
    if slug in OVERRIDES:
        return OVERRIDES[slug]

    # Generic rule: strip last hyphenated word (mascot)
    # e.g. "lafayette-leopards" → "lafayette"
    # But only if the result looks reasonable (not stripping real name parts)
    parts = slug.split("-")
    if len(parts) >= 2:
        # Remove last token (mascot) as a trial
        candidate = "-".join(parts[:-1])
        # Apply -state → -st rule
        if candidate.endswith("-state"):
            candidate = candidate[:-6] + "-st"
        return candidate

    return slug


def build_score_lookup(wb_games: list) -> dict:
    """
    Build dict: (date, away_score_int, home_score_int) → game_id for fast matching.
    Uses integer scores to avoid float precision issues.
    """
    lookup = {}
    for g in wb_games:
        if g["home_score_f"] is None:
            continue
        key = (
            g["date"],
            round(g["away_score_f"]),
            round(g["home_score_f"]),
        )
        lookup[key] = g["game_id"]
    return lookup


def build_team_lookup(wb_games: list) -> dict:
    """Dict: (date, away_seo, home_seo) → game_id for team-name matching."""
    lookup = {}
    for g in wb_games:
        lookup[(g["date"], g["away_seo"], g["home_seo"])] = g["game_id"]
    return lookup


def match_game(covers_game: dict, date: str,
               score_lookup: dict, team_lookup: dict,
               wb_games_by_date: dict) -> tuple[str | None, str]:
    """
    Match a Covers game to a workbook GameID.
    Returns (game_id, match_method) or (None, 'NO_MATCH').
    """
    away_score = covers_game.get("away_score")
    home_score = covers_game.get("home_score")
    away_slug = covers_game.get("away_slug", "")
    home_slug = covers_game.get("home_slug", "")

    # 1. Score-based exact match (most reliable)
    if away_score is not None and home_score is not None:
        # Try exact date match
        key = (date, away_score, home_score)
        gid = score_lookup.get(key)
        if gid:
            return gid, "SCORE_EXACT"

        # Try date ±1 day (UTC offset for late games)
        for delta in [-1, 1]:
            from datetime import timedelta
            alt_date = (datetime.strptime(date, "%Y-%m-%d") + timedelta(days=delta)).strftime("%Y-%m-%d")
            key2 = (alt_date, away_score, home_score)
            gid2 = score_lookup.get(key2)
            if gid2:
                return gid2, f"SCORE_DATE±{delta}"

    # 2. Team-name match (convert Covers slugs to SDIO-like)
    away_sdio = slug_from_covers(away_slug)
    home_sdio = slug_from_covers(home_slug)

    # Try exact seo match
    key_team = (date, away_sdio, home_sdio)
    gid = team_lookup.get(key_team)
    if gid:
        return gid, "TEAM_EXACT"

    # Try common Team slug variants
    games_on_date = wb_games_by_date.get(date, [])
    best_score = 0.0
    best_gid = None
    for wg in games_on_date:
        away_sim = _slug_similarity(away_sdio, wg["away_seo"])
        home_sim = _slug_similarity(home_sdio, wg["home_seo"])
        combined = away_sim * home_sim
        if combined > best_score:
            best_score = combined
            best_gid = wg["game_id"]

    if best_gid and best_score >= 0.6:
        return best_gid, f"TEAM_FUZZY({best_score:.2f})"

    return None, "NO_MATCH"


def _slug_similarity(a: str, b: str) -> float:
    """Quick token overlap similarity between two slugs."""
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.85
    a_parts = set(a.split("-"))
    b_parts = set(b.split("-"))
    overlap = len(a_parts & b_parts)
    union = len(a_parts | b_parts)
    return overlap / union if union else 0.0


# ---------------------------------------------------------------------------
# Canonical CSV I/O
# ---------------------------------------------------------------------------

def load_existing_canonical() -> dict:
    """Load existing canonical_lines.csv keyed by game_id."""
    existing = {}
    if not CANONICAL_PATH.exists():
        return existing
    with CANONICAL_PATH.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            gid = row.get("game_id", "").strip()
            # Skip synthetic data rows that we'll be replacing
            src = row.get("source_file", "")
            if "synthetic" in src.lower():
                continue
            if gid:
                existing[gid] = row
    return existing


def write_canonical(rows: list):
    """Write the canonical_lines.csv."""
    CANONICAL_PATH.parent.mkdir(parents=True, exist_ok=True)
    with CANONICAL_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CANONICAL_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in CANONICAL_FIELDS})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Scrape Covers.com NCAAB historical lines")
    parser.add_argument("--date", help="Single date YYYY-MM-DD")
    parser.add_argument("--all", action="store_true", help="All workbook dates")
    parser.add_argument("--since", help="Only dates >= this YYYY-MM-DD")
    parser.add_argument("--dry-run", action="store_true", help="Don't write output")
    args = parser.parse_args()

    print("Loading workbook...")
    wb_games = load_workbook_games()
    print(f"  {len(wb_games)} games with results in workbook")

    # Build lookups
    score_lookup = build_score_lookup(wb_games)
    team_lookup = build_team_lookup(wb_games)
    wb_games_by_date: dict = {}
    for g in wb_games:
        wb_games_by_date.setdefault(g["date"], []).append(g)

    all_dates = sorted(wb_games_by_date.keys())

    # Determine which dates to scrape
    if args.date:
        dates_to_scrape = [args.date]
    elif getattr(args, "all"):
        dates_to_scrape = all_dates
    else:
        # Default: all dates that don't already have lines
        existing = load_existing_canonical()
        covered_dates = set()
        for gid, row in existing.items():
            covered_dates.add(row.get("date", ""))
        dates_to_scrape = [d for d in all_dates if d not in covered_dates]
        if not dates_to_scrape:
            print("All dates already have lines! Use --all to re-scrape.")
            sys.exit(0)
        print(f"  Scraping {len(dates_to_scrape)} dates not yet in canonical_lines.csv")

    if args.since:
        dates_to_scrape = [d for d in dates_to_scrape if d >= args.since]

    # Load existing canonical (skip synthetic), will merge below
    existing_canonical = load_existing_canonical() if not getattr(args, "all") else {}

    total_scraped = 0
    total_matched = 0
    total_lines = 0
    new_rows: dict = {}  # game_id → canonical row

    for date in dates_to_scrape:
        print(f"\n[{date}]  fetching Covers.com...")
        games = scrape_date(date)
        print(f"  {len(games)} game blocks parsed")

        matched = 0
        with_lines = 0
        for cg in games:
            gid, method = match_game(cg, date, score_lookup, team_lookup, wb_games_by_date)
            if not gid:
                away_s = cg.get("away_slug", "?")
                home_s = cg.get("home_slug", "?")
                print(f"    UNMATCHED: {away_s} @ {home_s}")
                continue
            matched += 1

            spread_h = cg.get("spread_home")
            total_g = cg.get("total_game")
            if spread_h is not None or total_g is not None:
                with_lines += 1

            row = {
                "game_id": gid,
                "date": date,
                "away_seo": cg.get("away_slug", ""),
                "home_seo": cg.get("home_slug", ""),
                "spread_home": f"{spread_h:.1f}" if spread_h is not None else "",
                "spread_away": f"{-spread_h:.1f}" if spread_h is not None else "",
                "ml_home": "",  # Covers doesn't provide moneylines
                "ml_away": "",
                "total_game": f"{total_g:.1f}" if total_g is not None else "",
                "total_2h": "",  # Covers doesn't provide 2H total
                "sgo_event_id": f"covers-{cg.get('covers_game_id', '')}",
                "source_file": "covers_scraper",
                "staged_timestamp": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            }
            new_rows[gid] = row

            ambig = cg.get("_spread_ambiguous")
            if ambig:
                print(f"    AMBIGUOUS SPREAD for GameID={gid}: {ambig}")

        print(f"  matched {matched}/{len(games)}, {with_lines} have spread+total")
        total_scraped += len(games)
        total_matched += matched
        total_lines += with_lines

        if len(dates_to_scrape) > 1:
            time.sleep(SLEEP_BETWEEN_REQUESTS)

    # Merge with existing
    merged = {**existing_canonical, **new_rows}
    all_rows = sorted(merged.values(), key=lambda r: (r.get("date", ""), r.get("game_id", "")))

    print(f"\n=== Summary ===")
    print(f"  Scraped: {total_scraped} Covers game blocks")
    print(f"  Matched to workbook: {total_matched}")
    print(f"  With spread+total lines: {total_lines}")
    print(f"  Total rows in canonical: {len(all_rows)}")

    if args.dry_run:
        print("  DRY RUN — not writing canonical_lines.csv")
        return

    write_canonical(all_rows)
    print(f"  Wrote {len(all_rows)} rows → {CANONICAL_PATH}")

    # Print match rate vs. full workbook
    wb_with_results = len(wb_games)
    pct = 100.0 * total_matched / wb_with_results if wb_with_results else 0
    print(f"\n  Coverage: {total_matched}/{wb_with_results} ({pct:.1f}%) workbook games have lines")


if __name__ == "__main__":
    main()
