import requests
from openpyxl import load_workbook

RESULTS_XLSX = r"C:\NCAA Model\logs\NCAAM Results.xlsx"
API_BASE = "https://ncaa-api.henrygd.me"
HEADERS = {"User-Agent": "Mozilla/5.0"}


def fetch_scoreboard(date_str: str, cache: dict):
    if date_str in cache:
        return cache[date_str]

    yyyy, mm, dd = date_str.split("-")
    url = f"{API_BASE}/scoreboard/basketball-men/d1/{yyyy}/{mm}/{dd}"
    r = requests.get(url, headers=HEADERS, timeout=45)
    r.raise_for_status()
    data = r.json()
    cache[date_str] = data
    return data


def find_game(scoreboard: dict, game_id: str):
    for item in scoreboard.get("games", []):
        game = item.get("game", {})
        if str(game.get("gameID")) == str(game_id):
            return game
    return None


def parse_range(text):
    if not text or "-" not in str(text):
        return None, None
    left, right = str(text).split("-", 1)
    try:
        return int(left.strip()), int(right.strip())
    except Exception:
        return None, None


def main():
    wb = load_workbook(RESULTS_XLSX)
    ws = wb["Game_Log"]

    cache = {}
    updated = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        game_id = ws[f"B{row}"].value
        if game_id in (None, ""):
            continue

        date_str = str(ws[f"A{row}"].value).strip()
        away_team = str(ws[f"C{row}"].value).strip()
        home_team = str(ws[f"D{row}"].value).strip()
        halftime_score = str(ws[f"E{row}"].value).strip()
        pred_winner = str(ws[f"G{row}"].value).strip()

        scoreboard = fetch_scoreboard(date_str, cache)
        game = find_game(scoreboard, str(game_id))

        if game is None:
            print(f"Skipped row {row}: game {game_id} not found on scoreboard for {date_str}")
            skipped += 1
            continue

        try:
            away_final = int(game["away"]["score"])
            home_final = int(game["home"]["score"])
        except Exception:
            print(f"Skipped row {row}: game {game_id} is not final yet")
            skipped += 1
            continue

        actual_winner = away_team if away_final > home_final else home_team
        actual_margin = abs(away_final - home_final)
        actual_total = away_final + home_final

        actual_2h = ""
        if "-" in halftime_score:
            try:
                away_ht, home_ht = halftime_score.split("-", 1)
                away_ht = int(away_ht.strip())
                home_ht = int(home_ht.strip())
                actual_2h = (away_final - away_ht) + (home_final - home_ht)
            except Exception:
                actual_2h = ""

        pred_2h_low, pred_2h_high = parse_range(ws[f"I{row}"].value)
        pred_total_low, pred_total_high = parse_range(ws[f"J{row}"].value)

        winner_correct = ""
        if pred_winner:
            winner_correct = "YES" if pred_winner.lower() == actual_winner.lower() else "NO"

        twoh_error = ""
        if actual_2h != "" and pred_2h_low is not None and pred_2h_high is not None:
            pred_2h_mid = (pred_2h_low + pred_2h_high) / 2
            twoh_error = round(actual_2h - pred_2h_mid, 1)

        total_error = ""
        if pred_total_low is not None and pred_total_high is not None:
            pred_total_mid = (pred_total_low + pred_total_high) / 2
            total_error = round(actual_total - pred_total_mid, 1)

        ws[f"L{row}"] = actual_winner
        ws[f"M{row}"] = actual_margin
        ws[f"N{row}"] = actual_2h
        ws[f"O{row}"] = actual_total
        ws[f"P{row}"] = winner_correct
        ws[f"Q{row}"] = twoh_error
        ws[f"R{row}"] = total_error

        print(f"Updated row {row}: game {game_id}")
        updated += 1

    wb.save(RESULTS_XLSX)

    print("")
    print("Update complete.")
    print(f"Rows updated: {updated}")
    print(f"Rows skipped: {skipped}")
    print(f"Workbook: {RESULTS_XLSX}")


if __name__ == "__main__":
    main()