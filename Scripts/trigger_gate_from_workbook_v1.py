import argparse
from openpyxl import load_workbook
from pathlib import Path


WORKBOOK_PATH = Path('logs/NCAAM Results.xlsx')
SHEET_NAME = 'Game_Log'
TRIGGER_VERSION = 'v1'
DEFAULT_ACTIONABLE_MIN_N = 60
FULL_STAKE_MIN_HIT = 75.0
FULL_STAKE_MIN_N = 75
HALF_STAKE_MIN_HIT = 70.0
HALF_STAKE_MIN_N = 60

# High-confidence trigger set mined from historical workbook data.
# Each trigger has historical hit rate and sample count used for transparency.
TRIGGERS = [
    {
        'name': 'Lead 0-3 & HT total 71-80',
        'min_n': 90,
        'hist_hit': 78.9,
        'check': lambda conf, lead_b, total_b, pace: lead_b == '0-3' and total_b == '71-80',
    },
    {
        'name': 'Lead 4-7 & HT total 71-80',
        'min_n': 111,
        'hist_hit': 78.4,
        'check': lambda conf, lead_b, total_b, pace: lead_b == '4-7' and total_b == '71-80',
    },
    {
        'name': 'MEDIUM-HIGH & Lead 4-7 & HT total 71-80',
        'min_n': 61,
        'hist_hit': 77.0,
        'check': lambda conf, lead_b, total_b, pace: conf == 'MEDIUM-HIGH' and lead_b == '4-7' and total_b == '71-80',
    },
    {
        'name': 'MEDIUM & HT total 71-80',
        'min_n': 101,
        'hist_hit': 74.3,
        'check': lambda conf, lead_b, total_b, pace: conf == 'MEDIUM' and total_b == '71-80',
    },
    {
        'name': 'MEDIUM-HIGH & HT total 71-80',
        'min_n': 272,
        'hist_hit': 72.1,
        'check': lambda conf, lead_b, total_b, pace: conf == 'MEDIUM-HIGH' and total_b == '71-80',
    },
]


def parse_halftime_score(value):
    if value in (None, '') or '-' not in str(value):
        return None, None
    left, right = str(value).split('-', 1)
    try:
        away = float(left.strip())
        home = float(right.strip())
        return away, home
    except Exception:
        return None, None


def lead_bucket(home_lead):
    if home_lead is None:
        return 'unknown'
    x = abs(home_lead)
    if x <= 3:
        return '0-3'
    if x <= 7:
        return '4-7'
    if x <= 12:
        return '8-12'
    return '13+'


def total_bucket(halftime_total):
    if halftime_total is None:
        return 'unknown'
    if halftime_total <= 60:
        return '<=60'
    if halftime_total <= 70:
        return '61-70'
    if halftime_total <= 80:
        return '71-80'
    return '81+'


def choose_trigger(confidence, lead_b, total_b, pace):
    matches = [t for t in TRIGGERS if t['check'](confidence, lead_b, total_b, pace)]
    if not matches:
        return None
    # Prefer larger-sample trigger first, then higher hit-rate.
    matches.sort(key=lambda t: (t['min_n'], t['hist_hit']), reverse=True)
    return matches[0]


def upsert_header(ws, headers, idx, name):
    if name in idx:
        return
    ws.cell(row=1, column=len(headers) + 1, value=name)
    headers.append(name)
    idx[name] = len(headers) - 1


def decide_stake_tier(trigger_decision, hist_hit, hist_n):
    if trigger_decision != 'ACTIONABLE':
        return 'PASS'
    if hist_hit is None or hist_n is None:
        return 'PASS'
    if hist_hit >= FULL_STAKE_MIN_HIT and hist_n >= FULL_STAKE_MIN_N:
        return 'FULL'
    if HALF_STAKE_MIN_HIT <= hist_hit < FULL_STAKE_MIN_HIT and hist_n >= HALF_STAKE_MIN_N:
        return 'HALF'
    return 'PASS'


def hype_line(wagered_flag):
    if wagered_flag == 'Y':
        return 'Full Send, Bro!'
    return "Pump your brakes, Jay. This aint the one, Fam"


def pick_text(row, keys):
    for key in keys:
        value = row.get(key)
        if value not in (None, ''):
            return str(value).strip()
    return ''


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--game-id', help='Only update a single GameID row')
    parser.add_argument('--min-n', type=int, default=DEFAULT_ACTIONABLE_MIN_N, help='Minimum sample required for a trigger to be actionable')
    parser.add_argument('--print-actionable', action='store_true', help='Print actionable-only rows after updating the workbook')
    parser.add_argument('--print-game-card', action='store_true', help='Print a single clean decision card for --game-id')
    args = parser.parse_args()

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)

    wb = load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'{SHEET_NAME} not found in workbook')
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    for header_name in [
        'TriggerDecision',
        'TriggerName',
        'TriggerHistHit',
        'TriggerHistN',
        'WageredFlag',
        'TriggerVersion',
        'StakeTier',
    ]:
        upsert_header(ws, headers, idx, header_name)

    actionable = 0
    total = 0
    actionable_rows = []
    selected_game_output = None

    for excel_row, row_vals in enumerate(rows[1:], start=2):
        row = {h: row_vals[i] if i < len(row_vals) else None for h, i in idx.items() if i < len(row_vals)}
        game_id = str(row.get('GameID') or '').strip()
        if args.game_id and game_id != str(args.game_id).strip():
            continue
        halftime = row.get('HalftimeScore')
        confidence = str(row.get('Confidence') or '').strip().upper()
        pace = str(row.get('PaceProfile') or '').strip().lower()
        away_team = str(row.get('Away') or '').strip()
        home_team = str(row.get('Home') or '').strip()
        pred_winner = pick_text(row, ['PredWinner'])
        pred_margin_range = pick_text(row, ['PredMarginRange', 'PredMargin'])
        pred_2h_range = pick_text(row, ['Pred2HRange', 'Pred2HRange_Narrow'])
        pred_total_range = pick_text(row, ['PredTotalRange', 'PredTotalRange_Narrow'])

        away_ht, home_ht = parse_halftime_score(halftime)
        home_lead = (home_ht - away_ht) if away_ht is not None and home_ht is not None else None
        ht_total = (home_ht + away_ht) if away_ht is not None and home_ht is not None else None
        lead_b = lead_bucket(home_lead)
        total_b = total_bucket(ht_total)

        trig = choose_trigger(confidence, lead_b, total_b, pace)
        total += 1

        if trig is None or trig['min_n'] < args.min_n:
            trigger_decision = 'PASS'
            trigger_name = ''
            trigger_hist_hit = ''
            trigger_hist_n = ''
        else:
            actionable += 1
            trigger_decision = 'ACTIONABLE'
            trigger_name = trig['name']
            trigger_hist_hit = trig['hist_hit']
            trigger_hist_n = trig['min_n']

        stake_tier = decide_stake_tier(
            trigger_decision,
            trigger_hist_hit if trigger_hist_hit != '' else None,
            trigger_hist_n if trigger_hist_n != '' else None,
        )
        wagered_flag = 'Y' if stake_tier in {'FULL', 'HALF'} else 'N'

        ws.cell(row=excel_row, column=idx['TriggerDecision'] + 1, value=trigger_decision)
        ws.cell(row=excel_row, column=idx['TriggerName'] + 1, value=trigger_name)
        ws.cell(row=excel_row, column=idx['TriggerHistHit'] + 1, value=trigger_hist_hit)
        ws.cell(row=excel_row, column=idx['TriggerHistN'] + 1, value=trigger_hist_n)
        ws.cell(row=excel_row, column=idx['WageredFlag'] + 1, value=wagered_flag)
        ws.cell(row=excel_row, column=idx['TriggerVersion'] + 1, value=TRIGGER_VERSION)
        ws.cell(row=excel_row, column=idx['StakeTier'] + 1, value=stake_tier)

        if args.game_id and game_id == str(args.game_id).strip():
            selected_game_output = {
                'game_id': game_id,
                'away': away_team,
                'home': home_team,
                'halftime_score': str(halftime or '').strip(),
                'pred_winner': pred_winner,
                'pred_margin_range': pred_margin_range,
                'pred_2h_range': pred_2h_range,
                'pred_total_range': pred_total_range,
                'trigger_decision': trigger_decision,
                'trigger_name': trigger_name,
                'trigger_hist_hit': trigger_hist_hit,
                'trigger_hist_n': trigger_hist_n,
                'stake_tier': stake_tier,
                'wagered_flag': wagered_flag,
            }

        if trigger_decision == 'ACTIONABLE':
            matchup = f'{away_team} @ {home_team}'
            actionable_rows.append({
                'game_id': game_id,
                'matchup': matchup,
                'trigger_name': trigger_name,
                'hist_hit': trigger_hist_hit,
                'hist_n': trigger_hist_n,
                'stake_tier': stake_tier,
            })

    wb.save(WORKBOOK_PATH)
    print(f'Updated trigger gate columns in {WORKBOOK_PATH}')
    print(f'Actionable rows: {actionable}/{total} ({(actionable / total * 100.0) if total else 0.0:.1f}%)')
    print(f'Actionable minimum sample: n >= {args.min_n}')

    if args.print_game_card:
        print('')
        if selected_game_output is None:
            print(f'GameID {args.game_id} not found in workbook')
            return
        print('HALFTIME DECISION CARD')
        print(f"Game: {selected_game_output['game_id']} | {selected_game_output['away']} @ {selected_game_output['home']}")
        print(f"Halftime: {selected_game_output['halftime_score']}")
        print(
            f"Prediction: Winner={selected_game_output['pred_winner']} | "
            f"Margin={selected_game_output['pred_margin_range']} | "
            f"2H={selected_game_output['pred_2h_range']} | "
            f"Total={selected_game_output['pred_total_range']}"
        )
        print(
            f"Gate: {selected_game_output['trigger_decision']} | "
            f"Trigger={selected_game_output['trigger_name'] or 'none'} | "
            f"Hit={selected_game_output['trigger_hist_hit'] or '-'} | "
            f"N={selected_game_output['trigger_hist_n'] or '-'} | "
            f"Stake={selected_game_output['stake_tier']}"
        )
        print(f"Call: {hype_line(selected_game_output['wagered_flag'])}")
        return

    if args.print_actionable:
        print('')
        print('ACTIONABLE slate')
        if not actionable_rows:
            print('None')
        else:
            for item in actionable_rows:
                print(
                    f"{item['game_id']} | {item['matchup']} | {item['trigger_name']} | "
                    f"Hit={item['hist_hit']} | N={item['hist_n']} | Stake={item['stake_tier']}"
                )


if __name__ == '__main__':
    main()
