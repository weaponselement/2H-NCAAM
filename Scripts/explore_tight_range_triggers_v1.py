import argparse
import glob
import json
from collections import defaultdict
from openpyxl import load_workbook


WORKBOOK = 'logs/NCAAM Results.xlsx'
REPORT_GLOB = 'data/processed/reports/feature_report_v5_test_*.json'


def safe_float(v):
    try:
        return float(v)
    except Exception:
        return None


def parse_range(value):
    if value in (None, ''):
        return None
    s = str(value)
    if '-' not in s:
        return None
    left, right = s.split('-', 1)
    try:
        lo = float(left.strip())
        hi = float(right.strip())
    except Exception:
        return None
    if lo > hi:
        lo, hi = hi, lo
    return (lo, hi)


def parse_halftime_score(value):
    if value in (None, ''):
        return None, None
    s = str(value)
    if '-' not in s:
        return None, None
    left, right = s.split('-', 1)
    try:
        away = float(left.strip())
        home = float(right.strip())
    except Exception:
        return None, None
    return away, home


def outcome_side(pred_range, line):
    if pred_range is None or line is None:
        return None
    lo, hi = pred_range
    if lo > line:
        return 'OVER'
    if hi < line:
        return 'UNDER'
    return 'AMBIG'


def hit_tight(actual, pred_range, half_width):
    if actual is None or pred_range is None:
        return None
    lo, hi = pred_range
    center = (lo + hi) / 2.0
    return abs(actual - center) <= half_width


def flatten_report(report):
    gs = report.get('game_state') or {}
    fp = report.get('foul_pressure') or {}
    projection = report.get('projection') or {}
    return {
        'pace_profile': str(gs.get('pace_profile') or '').strip().lower(),
        'estimated_possessions_per_team_1h': safe_float(gs.get('estimated_possessions_per_team_1H')),
        'whistle_events_count': safe_float(gs.get('whistle_events_count')),
        'long_dead_ball_gaps': safe_float(gs.get('long_dead_ball_gaps')),
        'dead_ball_events': safe_float(gs.get('dead_ball_events')),
        'possession_change_markers': safe_float(gs.get('possession_change_markers')),
        'neutral_segment_deadballs': safe_float(gs.get('neutral_segment_deadballs')),
        'foul_pressure_text': str(fp.get('summary') or fp.get('label') or '').strip().lower(),
        'confidence': str(projection.get('confidence') or '').strip().upper(),
    }


def bucket_num(x, cuts, labels):
    if x is None:
        return 'unknown'
    for c, label in zip(cuts, labels):
        if x <= c:
            return label
    return labels[-1]


def build_trigger_keys(rec):
    pace = rec['pace_profile'] or 'unknown'
    conf = rec['confidence'] or 'unknown'

    pos_b = bucket_num(rec['estimated_possessions_per_team_1h'], [36, 39, 42], ['<=36', '37-39', '40-42', '43+'])
    whistle_b = bucket_num(rec['whistle_events_count'], [20, 26, 32], ['<=20', '21-26', '27-32', '33+'])
    posschg_b = bucket_num(rec['possession_change_markers'], [70, 78, 86], ['<=70', '71-78', '79-86', '87+'])
    dead_b = bucket_num(rec['dead_ball_events'], [60, 70, 80], ['<=60', '61-70', '71-80', '81+'])
    longgap_b = bucket_num(rec['long_dead_ball_gaps'], [4, 7, 10], ['<=4', '5-7', '8-10', '11+'])

    keys = [
        f'pace={pace}',
        f'conf={conf}',
        f'poss={pos_b}',
        f'whistle={whistle_b}',
        f'poschg={posschg_b}',
        f'dead={dead_b}',
        f'longgap={longgap_b}',
        f'pace={pace}|conf={conf}',
        f'pace={pace}|poss={pos_b}',
        f'pace={pace}|whistle={whistle_b}',
        f'pace={pace}|longgap={longgap_b}',
        f'conf={conf}|poss={pos_b}',
        f'conf={conf}|whistle={whistle_b}',
        f'conf={conf}|longgap={longgap_b}',
        f'pace={pace}|conf={conf}|whistle={whistle_b}',
        f'pace={pace}|conf={conf}|poss={pos_b}',
        f'pace={pace}|conf={conf}|longgap={longgap_b}',
    ]
    return keys


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--line-total', type=float, default=146.5)
    parser.add_argument('--min-n', type=int, default=40)
    parser.add_argument('--target-hit', type=float, default=70.0)
    args = parser.parse_args()

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb['Game_Log']
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    workbook_by_gid = {}
    for r in rows[1:]:
        gid = str(r[idx['GameID']]).strip() if idx.get('GameID') is not None and r[idx['GameID']] is not None else ''
        if gid:
            workbook_by_gid[gid] = r

    reports_by_gid = {}
    for path in glob.glob(REPORT_GLOB):
        with open(path, 'r', encoding='utf-8') as f:
            report = json.load(f)
        gid = str(report.get('gameID') or report.get('game_id') or '').strip()
        if gid:
            reports_by_gid[gid] = report

    base_rows = []
    for gid, row in workbook_by_gid.items():
        report = reports_by_gid.get(gid)
        if report is None:
            continue
        actual_2h = safe_float(row[idx['Actual2H']]) if idx.get('Actual2H') is not None else None
        actual_total = safe_float(row[idx['ActualTotal']]) if idx.get('ActualTotal') is not None else None
        pred_2h = parse_range(row[idx['Pred2HRange']]) if idx.get('Pred2HRange') is not None else None
        pred_total = parse_range(row[idx['PredTotalRange']]) if idx.get('PredTotalRange') is not None else None
        if actual_2h is None or actual_total is None or pred_2h is None or pred_total is None:
            continue
        away_ht, home_ht = parse_halftime_score(row[idx['HalftimeScore']] if idx.get('HalftimeScore') is not None else None)
        if away_ht is None or home_ht is None:
            continue
        rec = flatten_report(report)
        rec['gid'] = gid
        rec['home_lead'] = home_ht - away_ht
        rec['ht_total'] = home_ht + away_ht
        rec['actual_2h'] = actual_2h
        rec['actual_total'] = actual_total
        rec['pred_2h'] = pred_2h
        rec['pred_total'] = pred_total
        rec['total_side_vs_line'] = outcome_side(pred_total, args.line_total)
        base_rows.append(rec)

    print(f'Eligible merged rows: {len(base_rows)}')

    for half_width in (2.0, 3.0):
        all_n = len(base_rows)
        h2_hit = sum(1 for r in base_rows if hit_tight(r['actual_2h'], r['pred_2h'], half_width))
        t_hit = sum(1 for r in base_rows if hit_tight(r['actual_total'], r['pred_total'], half_width))
        side_n = sum(1 for r in base_rows if r['total_side_vs_line'] in {'OVER', 'UNDER'})
        side_hit = 0
        for r in base_rows:
            if r['total_side_vs_line'] not in {'OVER', 'UNDER'}:
                continue
            actual_side = 'OVER' if r['actual_total'] > args.line_total else 'UNDER'
            if actual_side == r['total_side_vs_line']:
                side_hit += 1
        print('')
        print(f'GLOBAL tight ±{int(half_width)}: 2H={h2_hit}/{all_n} ({(100.0*h2_hit/all_n) if all_n else 0:.1f}%), Total={t_hit}/{all_n} ({(100.0*t_hit/all_n) if all_n else 0:.1f}%)')
        print(f'GLOBAL directional vs line {args.line_total}: {side_hit}/{side_n} ({(100.0*side_hit/side_n) if side_n else 0:.1f}%), non-ambig share={(100.0*side_n/all_n) if all_n else 0:.1f}%')

    agg = defaultdict(lambda: {'n':0, 'h2_t2':0, 'tot_t2':0, 'tot_t3':0, 'side_n':0, 'side_hit':0})
    for r in base_rows:
        keys = build_trigger_keys(r)
        for k in keys:
            a = agg[k]
            a['n'] += 1
            if hit_tight(r['actual_2h'], r['pred_2h'], 2.0):
                a['h2_t2'] += 1
            if hit_tight(r['actual_total'], r['pred_total'], 2.0):
                a['tot_t2'] += 1
            if hit_tight(r['actual_total'], r['pred_total'], 3.0):
                a['tot_t3'] += 1
            if r['total_side_vs_line'] in {'OVER', 'UNDER'}:
                a['side_n'] += 1
                actual_side = 'OVER' if r['actual_total'] > args.line_total else 'UNDER'
                if actual_side == r['total_side_vs_line']:
                    a['side_hit'] += 1

    scored = []
    for k, a in agg.items():
        n = a['n']
        if n < args.min_n:
            continue
        h2_t2 = 100.0 * a['h2_t2'] / n
        tot_t2 = 100.0 * a['tot_t2'] / n
        tot_t3 = 100.0 * a['tot_t3'] / n
        side_hit = (100.0 * a['side_hit'] / a['side_n']) if a['side_n'] else 0.0
        scored.append((k, n, h2_t2, tot_t2, tot_t3, a['side_n'], side_hit))

    print('')
    print(f'Top trigger candidates (min n={args.min_n}) by Total tight ±2 then side-hit')
    scored.sort(key=lambda x: (x[3], x[6], x[2], x[1]), reverse=True)
    for row in scored[:30]:
        k, n, h2_t2, tot_t2, tot_t3, side_n, side_hit = row
        print(f'{k:55s} n={n:4d} | 2H±2={h2_t2:5.1f}% | Total±2={tot_t2:5.1f}% | Total±3={tot_t3:5.1f}% | side={side_hit:5.1f}% on {side_n}')

    print('')
    print(f'Actionable trigger candidates (Total±2 >= {args.target_hit} and n >= {args.min_n})')
    found = 0
    for row in scored:
        k, n, h2_t2, tot_t2, tot_t3, side_n, side_hit = row
        if tot_t2 >= args.target_hit:
            found += 1
            print(f'{k:55s} n={n:4d} | 2H±2={h2_t2:5.1f}% | Total±2={tot_t2:5.1f}% | side={side_hit:5.1f}% on {side_n}')
    if found == 0:
        print('None found. Try lowering min-n, using ±3, or switching to directional side-hit triggers.')


if __name__ == '__main__':
    main()
