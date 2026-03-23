from openpyxl import load_workbook
from pathlib import Path

path = Path('logs/NCAAM Results.xlsx')
if not path.exists():
    raise FileNotFoundError(path)

wb = load_workbook(path, data_only=True)
if 'Game_Log' not in wb.sheetnames:
    raise ValueError('Game_Log not in workbook')

ws = wb['Game_Log']
rows = list(ws.iter_rows(values_only=True))
if len(rows) < 2:
    raise ValueError('No data')

headers = [str(c) if c is not None else '' for c in rows[0]]
print('rows', len(rows)-1, 'cols', len(headers))
print('headers', headers)

# data rows as dicts
data = [dict(zip(headers, r)) for r in rows[1:]]

# Filter by date if specified
import sys
date_filter = None
if len(sys.argv) > 1:
    date_filter = sys.argv[1]
if date_filter:
    data = [r for r in data if str(r.get('Date', '')).strip() == date_filter]
    print(f'Filtered to date {date_filter}, rows: {len(data)}')

# detect GameID column
game_id_col = None
for c in headers:
    if str(c).strip().lower() in ('gameid', 'game_id', 'game id', 'b', 'game id'):
        game_id_col = c
        break
if not game_id_col:
    for c in headers:
        if 'game' in str(c).lower() and 'id' in str(c).lower():
            game_id_col = c
            break

if game_id_col:
    ids = [r.get(game_id_col) for r in data]
    dup_ids = [x for x in ids if x is not None and ids.count(x) > 1]
    dup_ids = sorted(set(dup_ids))
    print('GameID col:', game_id_col, 'duplicates:', len(dup_ids))
    if dup_ids:
        print('sample duplicates:', dup_ids[:10])
else:
    print('GameID column not found')

pred_cols = [c for c in headers if any(k in str(c).lower() for k in ('pred', 'projection', 'projected'))]
actual_cols = [c for c in headers if any(k in str(c).lower() for k in ('final', 'actual', 'result', 'score'))]
print('prediction columns:', pred_cols)
print('actual columns:', actual_cols)

print('first 5 rows:')
for i, r in enumerate(data[:5], 1):
    o = {k: r.get(k) for k in headers[:12]}
    print(i, o)


def parse_score(text):
    if text is None:
        return None
    s = str(text).strip()
    if '-' not in s:
        return None
    try:
        left, right = s.split('-', 1)
        return float(left.strip()), float(right.strip())
    except Exception:
        return None

# If we can compare margins or totals
# find columns for predictions and actual if numeric
def first_numeric(col):
    values = []
    for row in data:
        v = row.get(col)
        if isinstance(v, (int, float)):
            values.append(v)
        else:
            try:
                values.append(float(v))
            except Exception:
                pass
    return values

# candidate actual total / margin columns
candidates = [c for c in headers if any(k in str(c).lower() for k in ('2h', 'total', 'margin', 'final'))]
print('candidate numeric columns for analysis:', candidates)

# Compute accuracy metrics (existing metrics in sheet)
if 'WinnerCorrect' in headers:
    wins = [r.get('WinnerCorrect') for r in data if r.get('WinnerCorrect') is not None]
    correct = sum(1 for v in wins if str(v).strip().lower() in ('true', '1', 'yes', 'y', 'correct'))
    total = len(wins)
    print(f'WinnerCorrect count: {correct}/{total} ({(correct/total*100) if total else 0:.1f}%)')

recomputed_winner_total = 0
recomputed_winner_correct = 0
if 'PredWinner' in headers and 'ActualWinner' in headers:
    for r in data:
        pred = r.get('PredWinner')
        actual_name = r.get('ActualWinner')
        if pred in (None, '') or actual_name in (None, ''):
            continue
        # Skip rows where ActualWinner is actually a score string from a corrupted workbook state.
        if parse_score(actual_name) is not None:
            continue
        recomputed_winner_total += 1
        if str(pred).strip().lower() == str(actual_name).strip().lower():
            recomputed_winner_correct += 1
    if recomputed_winner_total:
        print(f'Recomputed winner accuracy: {recomputed_winner_correct}/{recomputed_winner_total} ({recomputed_winner_correct/recomputed_winner_total*100:.1f}%)')

# compute numeric errors if possible
# use Total_Error, TwoH_Error if present
for col in ['Total_Error', 'TwoH_Error']:
    if col in headers:
        vals = [r.get(col) for r in data if isinstance(r.get(col), (int, float))]
        if vals:
            import statistics
            print(f'{col}: mean {statistics.mean(vals):.2f}, median {statistics.median(vals):.2f}, std {statistics.stdev(vals):.2f}, max {max(vals):.2f}')

# parse PredMargin ranges to compare with ActualMargin when ActualMargin is numeric
if 'PredMargin' in headers and 'ActualMargin' in headers:
    margin_diff = []
    for r in data:
        p = r.get('PredMargin')
        a = r.get('ActualMargin')
        if p is None or a is None:
            continue
        try:
            # assume p like '6-11' or '1-5'
            lo, hi = [float(x) for x in str(p).split('-')]
            m = (lo+hi)/2
            margin_diff.append(abs(m - float(a)))
        except Exception:
            continue
    if margin_diff:
        import statistics
        print('PredMargin vs ActualMargin error: mean', statistics.mean(margin_diff), 'median', statistics.median(margin_diff), 'max', max(margin_diff))

# 2H and Total range hit / error

def parse_range(v):
    if v is None:
        return None
    s = str(v).strip()
    if '-' not in s:
        return None
    parts = [p.strip() for p in s.split('-')]
    try:
        low = float(parts[0])
        high = float(parts[1])
        return low, high
    except Exception:
        return None

if 'Pred2HRange' in headers and 'Actual2H' in headers:
    hits = 0
    total = 0
    diffs = []
    for r in data:
        pred = parse_range(r.get('Pred2HRange'))
        actual = r.get('Actual2H')
        if pred is None or actual is None:
            continue
        total += 1
        low, high = pred
        try:
            av = float(actual)
        except Exception:
            continue
        if low <= av <= high:
            hits += 1
        mid = (low + high) / 2
        diffs.append(abs(mid - av))
    if total:
        import statistics
        print(f'2H range hit: {hits}/{total} ({hits/total*100:.1f}%)')
        print('2H midpoint error: mean', statistics.mean(diffs), 'median', statistics.median(diffs), 'max', max(diffs))

if 'PredTotalRange' in headers and 'ActualTotal' in headers:
    hits = 0
    total = 0
    diffs = []
    for r in data:
        pred = parse_range(r.get('PredTotalRange'))
        actual = r.get('ActualTotal')
        if pred is None or actual is None:
            continue
        total += 1
        low, high = pred
        try:
            av = float(actual)
        except Exception:
            continue
        if low <= av <= high:
            hits += 1
        mid = (low + high) / 2
        diffs.append(abs(mid - av))
    if total:
        import statistics
        print(f'Total range hit: {hits}/{total} ({hits/total*100:.1f}%)')
        print('Total midpoint error: mean', statistics.mean(diffs), 'median', statistics.median(diffs), 'max', max(diffs))

print('analysis complete')
