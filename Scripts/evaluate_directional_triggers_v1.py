import argparse
import csv
from collections import defaultdict
from openpyxl import load_workbook

WORKBOOK_PATH = 'logs/NCAAM Results.xlsx'
SHEET_NAME = 'Game_Log'
CANONICAL_LINES_PATH = 'data/processed/market_lines/canonical_lines.csv'


def safe_float(value):
    try:
        return float(value)
    except Exception:
        return None


def parse_range(value):
    if value in (None, ''):
        return None
    s = str(value)
    if '-' not in s:
        return None
    left, right = s.split('-', 1)
    try:
        lo = float(left.strip())
        hi = float(right.strip())
    except Exception:
        return None
    if lo > hi:
        lo, hi = hi, lo
    return lo, hi


def midpoint_from_range(value):
    rng = parse_range(value)
    if rng is None:
        return None
    return (rng[0] + rng[1]) / 2.0


def read_workbook_rows():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else '' for c in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return out


def load_lines():
    out = {}
    with open(CANONICAL_LINES_PATH, 'r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            gid = str(row.get('game_id') or '').strip()
            if not gid:
                continue
            out[gid] = {
                'line_total_game': safe_float(row.get('total_game')),
                'line_total_2h': safe_float(row.get('total_2h')),
            }
    return out


def classify_direction(midpoint, line):
    if midpoint is None or line is None:
        return None
    if midpoint > line:
        return 'OVER'
    if midpoint < line:
        return 'UNDER'
    return None


def actual_direction(actual_points, line):
    if actual_points is None or line is None:
        return None
    if actual_points > line:
        return 'OVER'
    if actual_points < line:
        return 'UNDER'
    return None


def compute_stats(records, line_key, pred_key, actual_key, thresholds):
    usable = []
    for r in records:
        line = r.get(line_key)
        midpoint = midpoint_from_range(r.get(pred_key))
        actual = safe_float(r.get(actual_key))
        pred_side = classify_direction(midpoint, line)
        actual_side = actual_direction(actual, line)
        if pred_side is None or actual_side is None:
            continue
        edge = abs(midpoint - line)
        usable.append({
            'edge': edge,
            'hit': 1 if pred_side == actual_side else 0,
            'pred_side': pred_side,
            'actual_side': actual_side,
            'line': line,
            'midpoint': midpoint,
            'actual': actual,
            'confidence': str(r.get('Confidence') or '').strip(),
            'pace': str(r.get('PaceProfile') or '').strip(),
            'trigger_decision': str(r.get('TriggerDecision') or '').strip(),
            'stake_tier': str(r.get('StakeTier') or '').strip(),
        })

    summary = {
        'n': len(usable),
        'hit_rate': (sum(x['hit'] for x in usable) / len(usable)) if usable else None,
        'over_pred_rate': (sum(1 for x in usable if x['pred_side'] == 'OVER') / len(usable)) if usable else None,
        'over_actual_rate': (sum(1 for x in usable if x['actual_side'] == 'OVER') / len(usable)) if usable else None,
    }

    by_threshold = {}
    for t in thresholds:
        bucket = [x for x in usable if x['edge'] >= t]
        by_threshold[t] = {
            'n': len(bucket),
            'hit_rate': (sum(x['hit'] for x in bucket) / len(bucket)) if bucket else None,
            'avg_edge': (sum(x['edge'] for x in bucket) / len(bucket)) if bucket else None,
        }

    by_conf = defaultdict(list)
    for x in usable:
        by_conf[x['confidence']].append(x)

    conf_summary = {}
    for conf, rows in by_conf.items():
        if not conf:
            continue
        conf_summary[conf] = {
            'n': len(rows),
            'hit_rate': sum(x['hit'] for x in rows) / len(rows),
            'avg_edge': sum(x['edge'] for x in rows) / len(rows),
        }

    return summary, by_threshold, conf_summary


def fmt_pct(v):
    if v is None:
        return 'n/a'
    return f'{100.0 * v:.1f}%'


def fmt_num(v):
    if v is None:
        return 'n/a'
    return f'{v:.2f}'


def print_block(title, summary, by_threshold, conf_summary):
    print(title)
    print(f"  n={summary['n']} hit={fmt_pct(summary['hit_rate'])} pred_over={fmt_pct(summary['over_pred_rate'])} actual_over={fmt_pct(summary['over_actual_rate'])}")
    print('  Thresholds (abs(midpoint - line))')
    for t in sorted(by_threshold.keys()):
        s = by_threshold[t]
        print(f"    edge>={t}: n={s['n']} hit={fmt_pct(s['hit_rate'])} avg_edge={fmt_num(s['avg_edge'])}")

    if conf_summary:
        print('  By confidence')
        for conf in sorted(conf_summary.keys()):
            s = conf_summary[conf]
            print(f"    {conf}: n={s['n']} hit={fmt_pct(s['hit_rate'])} avg_edge={fmt_num(s['avg_edge'])}")


def main():
    parser = argparse.ArgumentParser(description='Evaluate directional O/U trigger accuracy from model midpoint vs closing lines.')
    parser.add_argument('--start-date', default='', help='Filter rows with Date >= YYYY-MM-DD')
    parser.add_argument('--end-date', default='', help='Filter rows with Date <= YYYY-MM-DD')
    parser.add_argument('--wagered-only', action='store_true', help='Use only rows with WageredFlag=Y')
    parser.add_argument('--trigger-version', default='', help='Filter by TriggerVersion (e.g. v1)')
    parser.add_argument('--thresholds', default='0,1,2,3,4,5,6,8,10', help='Comma-separated edge thresholds')
    args = parser.parse_args()

    thresholds = []
    for piece in args.thresholds.split(','):
        piece = piece.strip()
        if not piece:
            continue
        thresholds.append(float(piece))

    rows = read_workbook_rows()
    lines = load_lines()

    joined = []
    for r in rows:
        gid = str(r.get('GameID') or '').strip()
        if not gid:
            continue
        line_row = lines.get(gid)
        if line_row is None:
            continue

        date_val = str(r.get('Date') or '').strip()
        date_only = date_val.split(' ')[0] if date_val else ''
        if args.start_date and (not date_only or date_only < args.start_date):
            continue
        if args.end_date and (not date_only or date_only > args.end_date):
            continue

        if args.wagered_only and str(r.get('WageredFlag') or '').strip().upper() != 'Y':
            continue
        if args.trigger_version and str(r.get('TriggerVersion') or '').strip() != args.trigger_version:
            continue

        rec = dict(r)
        rec.update(line_row)
        joined.append(rec)

    print(f'Joined rows (workbook + canonical lines): {len(joined)}')

    full_summary, full_thresholds, full_conf = compute_stats(
        joined,
        line_key='line_total_game',
        pred_key='PredTotalRange',
        actual_key='ActualTotal',
        thresholds=thresholds,
    )

    second_half_summary, second_half_thresholds, second_half_conf = compute_stats(
        joined,
        line_key='line_total_2h',
        pred_key='Pred2HRange',
        actual_key='Actual2H',
        thresholds=thresholds,
    )

    print_block('FULL_GAME_OU_DIRECTIONAL', full_summary, full_thresholds, full_conf)
    print_block('SECOND_HALF_OU_DIRECTIONAL', second_half_summary, second_half_thresholds, second_half_conf)


if __name__ == '__main__':
    main()
