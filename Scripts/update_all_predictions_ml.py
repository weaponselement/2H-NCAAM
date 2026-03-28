from openpyxl import load_workbook
from pathlib import Path
import pickle
import json

from model_feature_utils import build_feature_vector, load_last4_pbp_priors, load_team_stats, load_market_lines, load_neutral_court_games, load_rest_context, parse_halftime_score, range_half_widths_for_halftime_total, resolve_team_stats
from step4b_feature_report_from_file_v5_test import load_game_pbp_features

# Load models
models_dir = Path('models')
models = {}
for target in ['ActualMargin', 'Actual2H', 'ActualTotal']:
    model_path = models_dir / f'{target.lower()}_model.pkl'
    if model_path.exists():
        with open(model_path, 'rb') as f:
            models[target] = pickle.load(f)
    else:
        print(f"Model {model_path} not found")
        models[target] = None

# Load model error stats for dynamic range widths
error_stats_path = models_dir / 'model_error_stats.json'
model_error_stats = {}
if error_stats_path.exists():
    with open(error_stats_path, 'r') as f:
        model_error_stats = json.load(f)
else:
    print(f"Warning: {error_stats_path} not found; using defaults")

strategy_path = models_dir / 'model_strategy.json'
model_strategy = {'total_prediction_strategy': 'direct_model'}
if strategy_path.exists():
    with open(strategy_path, 'r', encoding='utf-8') as f:
        model_strategy = json.load(f)

# Load workbook
path = Path('logs/NCAAM Results.xlsx')
if not path.exists():
    raise FileNotFoundError(path)

wb = load_workbook(path)
ws = wb['Game_Log']

if ws['S1'].value in (None, ''):
    ws['S1'] = 'Pred2HRange_Narrow'
if ws['T1'].value in (None, ''):
    ws['T1'] = 'PredTotalRange_Narrow'

# Load team stats for all dates
rows = list(ws.iter_rows(values_only=True))
headers = [str(c) if c is not None else '' for c in rows[0]]
data_indices = {h: i for i, h in enumerate(headers)}

team_stats_cache = {}
for row in rows[1:]:
    date = row[data_indices.get('Date')]
    if date:
        date_str = str(date).split(' ')[0]  # assume YYYY-MM-DD
        if date_str not in team_stats_cache:
            team_stats_cache[date_str] = load_team_stats(date_str)

pbp_feature_cache = {}
data_root = str(Path(__file__).resolve().parent.parent / 'data')

# Load market lines
market_lines_cache = load_market_lines()

# Load rest context
rest_context = load_rest_context(data_root=data_root)

# Load neutral court game IDs
neutral_court_games = load_neutral_court_games(data_root=data_root)

# Load date-keyed last4 historical PBP priors
last4_pbp_priors_by_date = {}
for date_str in team_stats_cache.keys():
    last4_pbp_priors_by_date[date_str] = load_last4_pbp_priors(date_str, data_root=data_root)

# Update predictions
updated = 0
for i, row in enumerate(rows[1:], start=2):  # start=2 for 1-based row
    date = row[data_indices.get('Date')]
    halftime_score = row[data_indices.get('HalftimeScore')]
    pace = row[data_indices.get('PaceProfile')]
    home_team = row[data_indices.get('Home')]
    away_team = row[data_indices.get('Away')]
    game_id = row[data_indices.get('GameID')] if data_indices.get('GameID') is not None else None

    if not halftime_score or not date or not home_team or not away_team:
        continue

    away_ht, home_ht, home_lead, halftime_total = parse_halftime_score(halftime_score)
    if home_lead is None:
        continue

    # Team stats
    date_str = str(date).split(' ')[0]
    stats = team_stats_cache.get(date_str, {})
    home_avg_scored, home_avg_allowed = resolve_team_stats(stats, home_team)
    away_avg_scored, away_avg_allowed = resolve_team_stats(stats, away_team)
    pbp_features = {}
    if game_id not in (None, ''):
        game_id_key = str(game_id).strip()
        if game_id_key not in pbp_feature_cache:
            pbp_feature_cache[game_id_key] = load_game_pbp_features(data_root, game_id_key)
        pbp_features = pbp_feature_cache.get(game_id_key, {})
    features, feature_dict = build_feature_vector(
        date_str,
        pace,
        home_ht,
        away_ht,
        home_avg_scored,
        home_avg_allowed,
        away_avg_scored,
        away_avg_allowed,
        pbp_features,
        game_id=game_id,
        market_lines_cache=market_lines_cache,
        home_team_seo=home_team,
        away_team_seo=away_team,
        rest_context=rest_context,
        neutral_court_games=neutral_court_games,
        last4_pbp_priors=last4_pbp_priors_by_date.get(date_str, {}),
    )

    def predict_target(model_key):
        model = models.get(model_key)
        if model is None:
            return None
        expected_features = getattr(model, 'n_features_in_', len(features))
        model_features = features[:expected_features]
        return model.predict([model_features])[0]

    # Predict
    pred_margin = predict_target('ActualMargin')
    pred_2h = predict_target('Actual2H')
    direct_pred_total = predict_target('ActualTotal')
    if model_strategy.get('total_prediction_strategy') == 'derived_2h' and pred_2h is not None:
        pred_total = feature_dict['halftime_total'] + pred_2h
    else:
        pred_total = direct_pred_total

    # Convert to whole number predictions
    if pred_margin is not None:
        pred_margin = int(round(pred_margin))
    if pred_2h is not None:
        pred_2h = int(round(pred_2h))
    if pred_total is not None:
        pred_total = int(round(pred_total))

    # Line deviation (if you add FD lines to inputs)
    book_2h_val = None
    for nm in ['FD2H', 'FanDuel2H', 'Book2H', 'HalfTotalLine']:
        if nm in data_indices and row[data_indices.get(nm)] is not None:
            try:
                book_2h_val = float(row[data_indices.get(nm)])
                break
            except:
                book_2h_val = None
    book_total_val = None
    for nm in ['FDTotal', 'FanDuelTotal', 'BookTotal', 'FullTotalLine', 'TotalLine']:
        if nm in data_indices and row[data_indices.get(nm)] is not None:
            try:
                book_total_val = float(row[data_indices.get(nm)])
                break
            except:
                book_total_val = None

    line_dev_2h = None
    line_dev_total = None
    if pred_2h is not None and book_2h_val is not None:
        line_dev_2h = pred_2h - book_2h_val
    if pred_total is not None and book_total_val is not None:
        line_dev_total = pred_total - book_total_val

    # Print for debug in live rows
    if line_dev_2h is not None or line_dev_total is not None:
        print(f"Game {row[data_indices.get('GameID')]}: line_dev_2h={line_dev_2h}, line_dev_total={line_dev_total}")

    if pred_margin is None:
        continue

    # Update winner
    winner = home_team if pred_margin > 0 else away_team
    ws[f"G{i}"] = winner

    # Margin range
    abs_margin = abs(pred_margin)
    if abs_margin >= 8:
        margin_range = "6-11"
        confidence = "MEDIUM-HIGH"
    elif abs_margin >= 5:
        margin_range = "3-8"
        confidence = "MEDIUM"
    else:
        margin_range = "1-5"
        confidence = "LOW-MEDIUM"
    ws[f"H{i}"] = margin_range
    ws[f"K{i}"] = confidence

    # 2H range (integers) - single operational band
    if pred_2h is not None:
        _, narrow_half, range_half_width = range_half_widths_for_halftime_total(feature_dict['halftime_total'])
        low = max(0, pred_2h - narrow_half)
        high = pred_2h + narrow_half
        ws[f"I{i}"] = f"{low}-{high}"
        ws[f"S{i}"] = f"{low}-{high}"

    # Total range (integers) - single operational band
    if pred_total is not None:
        _, narrow_half, range_half_width = range_half_widths_for_halftime_total(feature_dict['halftime_total'])
        low = max(0, pred_total - narrow_half)
        high = pred_total + narrow_half
        ws[f"J{i}"] = f"{low}-{high}"
        ws[f"T{i}"] = f"{low}-{high}"

    updated += 1
    if updated % 100 == 0:
        print(f"Updated {updated} rows")

wb.save(path)
print(f"Updated {updated} predictions in workbook")