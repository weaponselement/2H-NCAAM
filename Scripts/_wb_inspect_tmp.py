"""Quick workbook inspection."""
from openpyxl import load_workbook
wb = load_workbook('logs/NCAAM Results.xlsx', data_only=True)
ws = wb['Game_Log']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c) for c in rows[0]]
data = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]
dates = sorted(set(str(r['Date']).split(' ')[0] for r in data if r.get('Date') and r.get('ActualTotal')))
print('Total rows with results:', sum(1 for r in data if r.get('ActualTotal')))
print('Date range:', dates[0], 'to', dates[-1])
print('Unique dates:', len(dates))
print('Sample dates:', dates[:5], '...', dates[-5:])
sample = {k: data[0][k] for k in ['Date','GameID','Away','Home','ActualTotal','ActualMargin']}
print('Sample row:', sample)
