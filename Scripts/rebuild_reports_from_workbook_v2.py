import os
import sys
import subprocess
from openpyxl import load_workbook
from paths import DATA_DIR, SCRIPTS_DIR, NCAAM_RESULTS_XLSX

WORKBOOK = str(NCAAM_RESULTS_XLSX)
SHEET_NAME = "Game_Log"
DATA_ROOT = str(DATA_DIR)
SCRIPT_TO_RUN = str(SCRIPTS_DIR / "step4b_feature_report_from_file_v5_test.py")
SELECTED_DIR = os.path.join(DATA_ROOT, "processed", "selected_games")
BASELINES_DIR = os.path.join(DATA_ROOT, "processed", "baselines")


def main():
    if not os.path.exists(WORKBOOK):
        print(f"Workbook not found: {WORKBOOK}")
        return

    if not os.path.exists(SCRIPT_TO_RUN):
        print(f"Feature script not found: {SCRIPT_TO_RUN}")
        return

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET_NAME]

    jobs = []
    seen = set()

    for row in range(2, ws.max_row + 1):
        gid = ws[f"B{row}"].value
        date_val = ws[f"A{row}"].value

        if gid in (None, "") or date_val in (None, ""):
            continue

        gid = str(gid).strip()
        date_str = str(date_val).strip()

        if not gid or gid in seen:
            continue

        seen.add(gid)

        selected_path = os.path.join(SELECTED_DIR, f"selected_games_{date_str}.json")
        baseline_path = os.path.join(BASELINES_DIR, f"last4_{date_str}.json")

        jobs.append((gid, date_str, selected_path, baseline_path))

    print(f"Unique GameIDs found: {len(jobs)}")

    failures = []

    for i, (gid, date_str, selected_path, baseline_path) in enumerate(jobs, start=1):
        print(f"[{i}/{len(jobs)}] Rebuilding report for {gid} ({date_str})")

        if not os.path.exists(selected_path):
            failures.append((gid, date_str, "missing selected_games file"))
            print(f"  FAIL missing selected_games: {selected_path}")
            continue

        if not os.path.exists(baseline_path):
            failures.append((gid, date_str, "missing baseline manifest"))
            print(f"  FAIL missing baseline manifest: {baseline_path}")
            continue

        cmd = [
            sys.executable,
            SCRIPT_TO_RUN,
            gid,
            "--data-root", DATA_ROOT,
            "--baseline-manifest", baseline_path,
            "--selected-games", selected_path,
        ]

        result = subprocess.run(cmd, capture_output=True, text=True)

        if result.returncode != 0:
            failures.append((gid, date_str, "script returned non-zero"))
            print("  FAIL")
            if result.stdout.strip():
                print(result.stdout.strip())
            if result.stderr.strip():
                print(result.stderr.strip())
        else:
            print("  OK")

    print("")
    print("REPORT REBUILD COMPLETE")
    print(f"Attempted: {len(jobs)}")
    print(f"Failures: {len(failures)}")

    if failures:
        print("Failures:")
        for gid, date_str, note in failures:
            print(f"  - {gid} | {date_str} | {note}")


if __name__ == "__main__":
    main()