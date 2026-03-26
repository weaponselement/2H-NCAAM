from collections import defaultdict
import csv
from openpyxl import load_workbook


def safe_float(v):
    try:
        return float(v)
    except Exception:
        return None


def parse_range(r):
    if r in (None, ''):
        return None
    p = str(r).split('-')
    if len(p) != 2:
        return None
    try:
        lo = float(p[0].strip())
        hi = float(p[1].strip())
        if lo > hi:
            lo, hi = hi, lo
        return lo, hi
    except Exception:
        return None


def in_range(rng, actual):
    if rng is None or actual is None:
        return None
    lo, hi = rng
    return lo <= actual <= hi


def parse_halftime_total(hs):
    if hs in (None, '') or '-' not in str(hs):
        return None
    try:
        a, b = str(hs).split('-', 1)
        return float(a.strip()) + float(b.strip())
    except Exception:
        return None


def ht_bucket(ht):
    if ht is None:
        return 'unknown'
    if ht <= 60:
        return '<=60'
    if ht <= 70:
        return '61-70'
    if ht <= 80:
        return '71-80'
    return '81+'


wb = load_workbook('logs/NCAAM Results.xlsx', data_only=True)
ws = wb['Game_Log']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c) if c is not None else '' for c in rows[0]]
data = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]

market_gids = set()
with open('data/processed/market_lines/canonical_lines.csv', newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        gid = str(row.get('game_id') or '').strip()
        if gid:
            market_gids.add(gid)


def is_conf(r, name):
    return str(r.get('Confidence') or '').strip().upper() == name


def is_pace(r, name):
    return str(r.get('PaceProfile') or '').strip().lower() == name


cohorts = {
    'overall': lambda r: True,
    'confidence=MEDIUM-HIGH': lambda r: is_conf(r, 'MEDIUM-HIGH'),
    'confidence=MEDIUM': lambda r: is_conf(r, 'MEDIUM'),
    'confidence=LOW-MEDIUM': lambda r: is_conf(r, 'LOW-MEDIUM'),
    'pace=run_and_gun': lambda r: is_pace(r, 'run_and_gun'),
    'pace=moderate': lambda r: is_pace(r, 'moderate'),
    'pace=grinder': lambda r: is_pace(r, 'grinder'),
    'market_covered=yes': lambda r: str(r.get('GameID') or '').strip() in market_gids,
    'market_covered=no': lambda r: str(r.get('GameID') or '').strip() not in market_gids,
    'ht_bucket=<=60': lambda r: ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '<=60',
    'ht_bucket=61-70': lambda r: ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '61-70',
    'ht_bucket=71-80': lambda r: ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '71-80',
    'ht_bucket=81+': lambda r: ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '81+',
    'MEDIUM-HIGH & market_covered': lambda r: is_conf(r, 'MEDIUM-HIGH') and str(r.get('GameID') or '').strip() in market_gids,
    'MEDIUM-HIGH & pace=moderate': lambda r: is_conf(r, 'MEDIUM-HIGH') and is_pace(r, 'moderate'),
    'MEDIUM-HIGH & ht_bucket=61-70': lambda r: is_conf(r, 'MEDIUM-HIGH') and ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '61-70',
    'MEDIUM-HIGH & ht_bucket=71-80': lambda r: is_conf(r, 'MEDIUM-HIGH') and ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '71-80',
    'MEDIUM-HIGH & market_covered & moderate': lambda r: is_conf(r, 'MEDIUM-HIGH') and (str(r.get('GameID') or '').strip() in market_gids) and is_pace(r, 'moderate'),
    'MEDIUM-HIGH & market_covered & ht61-70': lambda r: is_conf(r, 'MEDIUM-HIGH') and (str(r.get('GameID') or '').strip() in market_gids) and (ht_bucket(parse_halftime_total(r.get('HalftimeScore'))) == '61-70'),
}

results = []
for name, fn in cohorts.items():
    n = 0
    h2 = 0
    total = 0
    for r in data:
        if not fn(r):
            continue
        a2 = safe_float(r.get('Actual2H'))
        at = safe_float(r.get('ActualTotal'))
        pr2 = parse_range(r.get('Pred2HRange'))
        prt = parse_range(r.get('PredTotalRange'))
        if a2 is None or at is None or pr2 is None or prt is None:
            continue
        n += 1
        if in_range(pr2, a2):
            h2 += 1
        if in_range(prt, at):
            total += 1
    if n > 0:
        results.append((name, n, h2 / n * 100.0, total / n * 100.0))

results.sort(key=lambda x: (x[3], x[2], x[1]), reverse=True)

print('TOP COHORTS (by Total then 2H)')
for name, n, h2p, tp in results[:25]:
    print(f'{name:50s} n={n:4d} 2H={h2p:5.1f}% Total={tp:5.1f}%')

print('\nCANDIDATES WITH BOTH >=80% (min n>=25)')
any_found = False
for name, n, h2p, tp in results:
    if n >= 25 and h2p >= 80.0 and tp >= 80.0:
        any_found = True
        print(f'{name:50s} n={n:4d} 2H={h2p:5.1f}% Total={tp:5.1f}%')
if not any_found:
    print('None in current candidate set.')
