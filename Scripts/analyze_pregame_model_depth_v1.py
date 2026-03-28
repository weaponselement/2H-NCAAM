#!/usr/bin/env python3
"""
Deep-dive analysis of the pregame totals model.
Tests a specific window with:
  - Per-fold consistency
  - Full fixed-threshold curve (no post-hoc picking)
  - Over/Under direction split
  - Monthly performance breakdown
"""

import argparse
import csv
import json
import math
import statistics
import sys
from pathlib import Path

import importlib.util
spec = importlib.util.spec_from_file_location("model_feature_utils",
    Path(__file__).resolve().parent / "model_feature_utils.py")
mfu = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mfu)

from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_ROOT    = PROJECT_ROOT / 'data'
LOGS_DIR     = PROJECT_ROOT / 'logs'
WORKBOOK_PATH       = PROJECT_ROOT / 'logs' / 'NCAAM Results.xlsx'
CANONICAL_LINES_PATH = PROJECT_ROOT / 'data' / 'processed' / 'market_lines' / 'canonical_lines.csv'

PBP_PRIOR_KEYS = [
    'last4_three_rate', 'last4_paint_share', 'last4_ft_rate',
    'last4_turnover_rate', 'last4_orb_rate',
    'last4_possessions_per_team_1h', 'last4_pbp_coverage_count',
]

def safe_float(v):
    try: return float(v)
    except: return None

def payout(odds):
    if odds > 0: return odds / 100.0
    return 100.0 / abs(odds)

# ------------------------------------------------------------------ #
# Data loading (same as evaluator)
# ------------------------------------------------------------------ #
def build_rows():
    wb = load_workbook(str(WORKBOOK_PATH), data_only=True)
    ws = wb['Game_Log']
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else '' for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(v is not None for v in row): continue
        data.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return data

def load_lines():
    out = {}
    with open(str(CANONICAL_LINES_PATH), 'r', encoding='utf-8', newline='') as f:
        for row in csv.DictReader(f):
            gid = str(row.get('game_id') or '').strip()
            if not gid: continue
            out[gid] = safe_float(row.get('total_game'))
    return out

def load_team_stats_for_window(date_str, window):
    baselines_dir = DATA_ROOT / 'processed' / 'baselines'
    p = baselines_dir / f'lastN_{window}_{date_str}.json'
    if not p.exists():
        return mfu.load_team_stats(date_str)
    try:
        with open(p, 'r', encoding='utf-8') as f:
            baseline = json.load(f)
    except Exception:
        return mfu.load_team_stats(date_str)
    stats = {}
    for team, games in (baseline.get('teams') or {}).items():
        if not isinstance(games, list) or not games: continue
        sf = [safe_float(g.get('score_for'))     for g in games]; sf = [x for x in sf if x is not None]
        sa = [safe_float(g.get('score_against')) for g in games]; sa = [x for x in sa if x is not None]
        if sf and sa:
            stats[team] = {'avg_scored': statistics.mean(sf), 'avg_allowed': statistics.mean(sa)}
    return stats

def load_priors_for_window(date_str, window):
    baselines_dir = DATA_ROOT / 'processed' / 'baselines'
    p = baselines_dir / f'lastN_{window}_{date_str}.json'
    if not p.exists():
        return mfu.load_last4_pbp_priors(date_str, data_root=str(DATA_ROOT))
    try:
        with open(p, 'r') as f:
            baseline = json.load(f)
        from step4b_feature_report_from_file_v5_test import load_game_pbp_features
        priors = {}
        for team_seo, games in (baseline.get('teams') or {}).items():
            if not isinstance(games, list) or not games: continue
            prior_rows = []
            for g in games:
                gid = str((g or {}).get('gameID') or '').strip()
                if not gid: continue
                pbp = load_game_pbp_features(str(DATA_ROOT), gid)
                if pbp: prior_rows.append(pbp)
            if not prior_rows:
                priors[team_seo] = {k: v for k, v in mfu.DEFAULT_PBP_FEATURES.items() if 'home_last4' in k}
                continue
            def _mean(key, default):
                vals = [float(r.get(key)) for r in prior_rows if r.get(key) is not None]
                return statistics.mean(vals) if vals else float(default)
            priors[team_seo] = {
                'last4_three_rate':            _mean('home_three_rate',           mfu.DEFAULT_PBP_FEATURES['home_last4_three_rate']),
                'last4_paint_share':           _mean('home_paint_share',          mfu.DEFAULT_PBP_FEATURES['home_last4_paint_share']),
                'last4_ft_rate':               _mean('home_ft_rate',              mfu.DEFAULT_PBP_FEATURES['home_last4_ft_rate']),
                'last4_turnover_rate':         _mean('home_turnover_rate',        mfu.DEFAULT_PBP_FEATURES['home_last4_turnover_rate']),
                'last4_orb_rate':              _mean('home_orb_rate',             mfu.DEFAULT_PBP_FEATURES['home_last4_orb_rate']),
                'last4_possessions_per_team_1h': _mean('possessions_per_team_1h', mfu.DEFAULT_PBP_FEATURES['home_last4_possessions_per_team_1h']),
                'last4_pbp_coverage_count':    float(len(prior_rows)),
            }
        return priors
    except Exception as e:
        return mfu.load_last4_pbp_priors(date_str, data_root=str(DATA_ROOT))

def get_team_priors(priors, team):
    td = (priors or {}).get(str(team or '').strip(), {})
    return {k: float(td.get(k, 0.0) or 0.0) for k in PBP_PRIOR_KEYS}

def expected_total(hs, ha, as_, aa):
    return ((hs + aa) / 2.0) + ((as_ + ha) / 2.0)

def build_feature_row(record, team_stats_by_date, priors_by_date):
    date_str    = str(record.get('Date') or '').split(' ')[0].strip()
    home_team   = str(record.get('Home') or '').strip()
    away_team   = str(record.get('Away') or '').strip()
    actual_total = safe_float(record.get('ActualTotal'))
    game_id     = str(record.get('GameID') or '').strip()
    if not date_str or not home_team or not away_team or actual_total is None or not game_id:
        return None

    stats = team_stats_by_date.get(date_str, {})
    hs, ha = mfu.resolve_team_stats(stats, home_team)
    as_, aa = mfu.resolve_team_stats(stats, away_team)
    exp = expected_total(hs, ha, as_, aa)

    priors = priors_by_date.get(date_str, {})
    hp = get_team_priors(priors, home_team)
    ap = get_team_priors(priors, away_team)

    features = {
        'home_avg_scored':      hs, 'home_avg_allowed':   ha,
        'away_avg_scored':      as_, 'away_avg_allowed':  aa,
        'expected_total':       exp,
        'home_offense_diff':    hs - aa, 'away_offense_diff': as_ - ha,
        'blended_possessions_1h':   (hp['last4_possessions_per_team_1h'] + ap['last4_possessions_per_team_1h']) / 2.0,
        'blended_possessions_full': (hp['last4_possessions_per_team_1h'] + ap['last4_possessions_per_team_1h']),
        'blended_three_rate':   (hp['last4_three_rate']    + ap['last4_three_rate'])    / 2.0,
        'blended_paint_share':  (hp['last4_paint_share']   + ap['last4_paint_share'])   / 2.0,
        'blended_ft_rate':      (hp['last4_ft_rate']       + ap['last4_ft_rate'])       / 2.0,
        'blended_turnover_rate':(hp['last4_turnover_rate'] + ap['last4_turnover_rate']) / 2.0,
        'blended_orb_rate':     (hp['last4_orb_rate']      + ap['last4_orb_rate'])      / 2.0,
        'blended_pbp_coverage': (hp['last4_pbp_coverage_count'] + ap['last4_pbp_coverage_count']) / 2.0,
        'home_last4_three_rate':    hp['last4_three_rate'],   'away_last4_three_rate':    ap['last4_three_rate'],
        'home_last4_ft_rate':       hp['last4_ft_rate'],      'away_last4_ft_rate':       ap['last4_ft_rate'],
        'home_last4_turnover_rate': hp['last4_turnover_rate'],'away_last4_turnover_rate': ap['last4_turnover_rate'],
        'home_last4_orb_rate':      hp['last4_orb_rate'],     'away_last4_orb_rate':      ap['last4_orb_rate'],
        'home_last4_possessions_1h':hp['last4_possessions_per_team_1h'],
        'away_last4_possessions_1h':ap['last4_possessions_per_team_1h'],
        'three_rate_gap':     hp['last4_three_rate']    - ap['last4_three_rate'],
        'ft_rate_gap':        hp['last4_ft_rate']       - ap['last4_ft_rate'],
        'turnover_rate_gap':  hp['last4_turnover_rate'] - ap['last4_turnover_rate'],
        'orb_rate_gap':       hp['last4_orb_rate']      - ap['last4_orb_rate'],
    }
    return {
        'game_id': game_id, 'date': date_str,
        'home': home_team, 'away': away_team,
        'actual_total': actual_total,
        'features': features, 'expected_total': exp,
        'month': date_str[:7],
    }

# ------------------------------------------------------------------ #
# Analysis helpers
# ------------------------------------------------------------------ #
def grade_bets(rows, pred_key, gap_threshold, odds=-110):
    pw = payout(odds)
    bets, wins, ou = [], [], []
    for r in rows:
        cl = r.get('closing_total')
        pred = r.get(pred_key)
        actual = r.get('actual_total')
        if None in (cl, pred, actual): continue
        gap = abs(pred - cl)
        if gap < gap_threshold: continue
        if pred == cl or actual == cl: continue
        direction = 'OVER' if pred > cl else 'UNDER'
        hit = int((direction == 'OVER') == (actual > cl))
        bets.append({'hit': hit, 'direction': direction, 'gap': gap, 'date': r.get('date'), 'month': r.get('month')})
        wins.append(hit)
        ou.append(direction)
    n = len(bets)
    if n == 0:
        return None
    w = sum(wins)
    units = w * pw - (n - w)
    return {
        'bets': n, 'wins': w, 'losses': n - w,
        'hit_rate': round(w / n * 100, 2),
        'units': round(units, 3),
        'roi': round(units / n * 100, 3),
        'bets_detail': bets,
    }

def threshold_curve(rows, pred_key, thresholds, odds=-110):
    results = []
    for t in thresholds:
        g = grade_bets(rows, pred_key, t, odds)
        if g:
            results.append({'threshold': t, 'bets': g['bets'], 'hit': g['hit_rate'], 'roi': g['roi'], 'units': g['units']})
        else:
            results.append({'threshold': t, 'bets': 0, 'hit': 0.0, 'roi': 0.0, 'units': 0.0})
    return results

def over_under_split(graded):
    over_bets  = [b for b in graded['bets_detail'] if b['direction'] == 'OVER']
    under_bets = [b for b in graded['bets_detail'] if b['direction'] == 'UNDER']
    def stats(blist, pw):
        n = len(blist)
        if not n: return None
        w = sum(b['hit'] for b in blist)
        units = w * pw - (n - w)
        return {'bets': n, 'wins': w, 'hit': round(w/n*100, 1), 'roi': round(units/n*100, 2)}
    pw = payout(-110)
    return {'OVER': stats(over_bets, pw), 'UNDER': stats(under_bets, pw)}

def monthly_breakdown(graded):
    by_month = {}
    for b in graded['bets_detail']:
        m = b.get('month', '?')
        by_month.setdefault(m, []).append(b)
    pw = payout(-110)
    rows = []
    for month in sorted(by_month):
        blist = by_month[month]
        n = len(blist); w = sum(b['hit'] for b in blist)
        units = w * pw - (n - w)
        rows.append({'month': month, 'bets': n, 'wins': w, 'hit': round(w/n*100,1), 'roi': round(units/n*100,2)})
    return rows

# ------------------------------------------------------------------ #
# Main
# ------------------------------------------------------------------ #
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--window', type=int, default=5)
    parser.add_argument('--min-train-dates', type=int, default=60)
    parser.add_argument('--fold-size', type=int, default=30)
    args = parser.parse_args()

    window = args.window
    print(f"\n{'='*70}")
    print(f"Deep-Dive Analysis: Window {window} | Walk-Forward")
    print(f"{'='*70}\n")

    workbook_rows = build_rows()
    lines = load_lines()
    unique_dates = sorted({str(r.get('Date') or '').split(' ')[0].strip()
                           for r in workbook_rows if r.get('Date') not in (None, '')})

    print(f"Loading data for {len(unique_dates)} dates...", flush=True)
    team_stats_by_date = {d: load_team_stats_for_window(d, window) for d in unique_dates}
    priors_by_date     = {d: load_priors_for_window(d, window)     for d in unique_dates}

    feature_rows = []
    for record in workbook_rows:
        built = build_feature_row(record, team_stats_by_date, priors_by_date)
        if built is None: continue
        built['closing_total'] = lines.get(built['game_id'])
        feature_rows.append(built)

    feature_rows = [r for r in feature_rows if r.get('closing_total') is not None]
    print(f"Rows with closing lines: {len(feature_rows)}\n", flush=True)

    dates_with_data = sorted({r['date'] for r in feature_rows})
    feature_names   = list(feature_rows[0]['features'].keys())

    # ---- Build walk-forward folds ---------------------------------- #
    min_train = args.min_train_dates
    fold_size = args.fold_size
    folds = []
    idx = min_train
    while idx < len(dates_with_data):
        fd_train = set(dates_with_data[:idx])
        fd_test  = set(dates_with_data[idx:idx + fold_size])
        if fd_test: folds.append((fd_train, fd_test))
        idx += fold_size

    print(f"Walk-forward: {len(folds)} folds, min_train={min_train}, fold_size={fold_size}\n")

    all_test_rows = []
    fold_summaries = []

    for fi, (fd_train, fd_test) in enumerate(folds):
        tr = [r for r in feature_rows if r['date'] in fd_train]
        te = [r for r in feature_rows if r['date'] in fd_test]
        if not tr or not te: continue

        model = RandomForestRegressor(n_estimators=400, min_samples_leaf=3, random_state=42, n_jobs=-1)
        model.fit([[r['features'][n] for n in feature_names] for r in tr],
                  [r['actual_total'] for r in tr])
        preds = model.predict([[r['features'][n] for n in feature_names] for r in te])

        for row, pred in zip(te, preds):
            row['model_pred_total']    = float(pred)
            row['baseline_pred_total'] = float(row['expected_total'])

        all_test_rows.extend(te)

        fd = sorted(fd_test)
        fold_summaries.append({
            'fold': fi + 1, 'dates': f"{fd[0]} -> {fd[-1]}",
            'train': len(tr), 'test': len(te),
        })
        print(f"  Fold {fi+1}: train={len(tr)} | test={len(te)} | {fd[0]} -> {fd[-1]}", flush=True)

    print(f"\nTotal test rows: {len(all_test_rows)}\n")

    thresholds = [0, 2, 4, 5, 6, 7, 8, 9, 10, 12, 15]

    # ================================================================ #
    # 1. FIXED THRESHOLD CURVE (no post-hoc picking)
    # ================================================================ #
    print(f"\n{'='*70}")
    print(f"1. BASELINE MODEL — Full Threshold Curve (no post-hoc selection)")
    print(f"{'='*70}")
    print(f"  {'GAP':>5}  {'BETS':>5}  {'HIT%':>6}  {'UNITS':>7}  {'ROI%':>7}")
    bl_curve = threshold_curve(all_test_rows, 'baseline_pred_total', thresholds)
    for row in bl_curve:
        print(f"  {row['threshold']:>5}  {row['bets']:>5}  {row['hit']:>6.1f}  {row['units']:>7.2f}  {row['roi']:>7.2f}%")

    print(f"\n{'='*70}")
    print(f"1b. RF MODEL — Full Threshold Curve")
    print(f"{'='*70}")
    print(f"  {'GAP':>5}  {'BETS':>5}  {'HIT%':>6}  {'UNITS':>7}  {'ROI%':>7}")
    rf_curve = threshold_curve(all_test_rows, 'model_pred_total', thresholds)
    for row in rf_curve:
        print(f"  {row['threshold']:>5}  {row['bets']:>5}  {row['hit']:>6.1f}  {row['units']:>7.2f}  {row['roi']:>7.2f}%")

    # ================================================================ #
    # 2. PER-FOLD CONSISTENCY at fixed thresholds
    # ================================================================ #
    print(f"\n{'='*70}")
    print(f"2. PER-FOLD CONSISTENCY (Baseline, fixed thresholds)")
    print(f"{'='*70}")
    for fixed_gap in [8, 10]:
        print(f"\n  Gap >= {fixed_gap}:")
        print(f"  {'FOLD':<6}  {'DATES':<30}  {'BETS':>5}  {'HIT%':>6}  {'ROI%':>7}")
        fold_test_buckets = []
        for fi, (fd_train, fd_test) in enumerate(folds):
            fold_rows = [r for r in all_test_rows if r['date'] in fd_test]
            g = grade_bets(fold_rows, 'baseline_pred_total', fixed_gap)
            fd = sorted(fd_test)
            label = f"{fd[0]} -> {fd[-1]}"
            if g:
                print(f"  {fi+1:<6}  {label:<30}  {g['bets']:>5}  {g['hit_rate']:>6.1f}  {g['roi']:>7.2f}%")
                fold_test_buckets.append((g['bets'], g['roi']))
            else:
                print(f"  {fi+1:<6}  {label:<30}  {'0':>5}  {'--':>6}  {'--':>7}")
        if len(fold_test_buckets) >= 2:
            rois = [r for _, r in fold_test_buckets]
            print(f"  -> ROI range: {min(rois):.2f}% to {max(rois):.2f}%  (spread={max(rois)-min(rois):.2f}pp)")

    # ================================================================ #
    # 3. OVER / UNDER SPLIT at best thresholds
    # ================================================================ #
    print(f"\n{'='*70}")
    print(f"3. OVER/UNDER DIRECTION SPLIT (Baseline)")
    print(f"{'='*70}")
    for fixed_gap in [8, 10]:
        g = grade_bets(all_test_rows, 'baseline_pred_total', fixed_gap)
        if not g: continue
        ou = over_under_split(g)
        print(f"\n  Gap >= {fixed_gap} | Total: {g['bets']} bets, {g['hit_rate']}% hit, {g['roi']}% ROI")
        for side, s in ou.items():
            if s: print(f"    {side:<6}: {s['bets']} bets | {s['hit']}% hit | {s['roi']}% ROI")

    # ================================================================ #
    # 4. MONTHLY BREAKDOWN at fixed threshold
    # ================================================================ #
    print(f"\n{'='*70}")
    print(f"4. MONTHLY BREAKDOWN (Baseline, Gap >= 8)")
    print(f"{'='*70}")
    g8 = grade_bets(all_test_rows, 'baseline_pred_total', 8)
    if g8:
        print(f"  {'MONTH':<8}  {'BETS':>5}  {'HIT%':>6}  {'ROI%':>7}")
        for row in monthly_breakdown(g8):
            flag = ' <-- loss month' if row['roi'] < 0 else ''
            print(f"  {row['month']:<8}  {row['bets']:>5}  {row['hit']:>6.1f}  {row['roi']:>7.2f}%{flag}")

    print(f"\n{'='*70}")
    print(f"Analysis complete for Window {window}")
    print(f"{'='*70}\n")

if __name__ == '__main__':
    main()
