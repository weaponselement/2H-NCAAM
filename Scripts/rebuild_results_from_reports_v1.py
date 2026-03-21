import glob
import json
import os
from openpyxl import load_workbook

RESULTS_XLSX = r"C:\NCAA Model\logs\NCAAM Results.xlsx"
REPORTS_DIR = r"C:\NCAA Model\data\processed\reports"


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def first_empty_row(ws):
    row = 2
    while ws[f"B{row}"].value not in (None, ""):
        row += 1
    return row


def normalize_team_name(value):
    if value is None:
        return ""
    return str(value).strip()


def extract_projection_range(projection, key_path):
    obj = projection
    for key in key_path:
        if not isinstance(obj, dict):
            return ""
        obj = obj.get(key, {})
    if isinstance(obj, str):
        return obj
    return ""


def main():
    report_paths = glob.glob(os.path.join(REPORTS_DIR, "feature_report*.json"))
    report_paths.sort(key=lambda p: os.path.getmtime(p))

    if not report_paths:
        print("No report files found.")
        return

    wb = load_workbook(RESULTS_XLSX)
    ws = wb["Game_Log"]

    # Clear everything below header row
    ws.delete_rows(2, ws.max_row)

    seen_game_ids = set()
    rows_written = 0

    for path in report_paths:
        report = load_json(path)

        game_id = str(report.get("gameID", "")).strip()
        if not game_id:
            continue

        # Keep only latest row per game ID if duplicates appear later
        if game_id in seen_game_ids:
            continue
        seen_game_ids.add(game_id)

        halftime = report.get("halftime_score", {})
        projection = report.get("projection", {})
        game_state = report.get("game_state", {})
        teams = report.get("teams", {})

        away = normalize_team_name((teams.get("away") or {}).get("seo", ""))
        home = normalize_team_name((teams.get("home") or {}).get("seo", ""))

        away_ht = halftime.get("away")
        home_ht = halftime.get("home")

        halftime_score = ""
        if away_ht is not None and home_ht is not None:
            halftime_score = f"{away_ht}-{home_ht}"

        pred_margin = projection.get("winner_margin_range", "")
        pred_2h = extract_projection_range(
            projection,
            ["second_half_points_projection", "range"]
        )
        pred_total = extract_projection_range(
            projection,
            ["full_game_total_projection", "range"]
        )

        row = first_empty_row(ws)

        ws[f"A{row}"] = report.get("run_date", "")
        ws[f"B{row}"] = game_id
        ws[f"C{row}"] = away
        ws[f"D{row}"] = home
        ws[f"E{row}"] = halftime_score
        ws[f"F{row}"] = game_state.get("pace_profile", "")
        ws[f"G{row}"] = projection.get("winner_projection", "")
        ws[f"H{row}"] = pred_margin
        ws[f"I{row}"] = pred_2h
        ws[f"J{row}"] = pred_total
        ws[f"K{row}"] = projection.get("confidence", "")

        rows_written += 1

    wb.save(RESULTS_XLSX)

    print(f"Rebuild complete.")
    print(f"Rows written: {rows_written}")
    print(f"Workbook: {RESULTS_XLSX}")


if __name__ == "__main__":
    main()