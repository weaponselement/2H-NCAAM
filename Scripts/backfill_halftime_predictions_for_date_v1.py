import argparse
import glob
import json
import os
import subprocess
import sys

from openpyxl import load_workbook
from paths import NCAAM_RESULTS_XLSX, DATA_DIR, SCRIPTS_DIR

def load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def find_row_by_game_id(ws, game_id: str):
    """Return row number where column B == game_id, else None."""
    row = 2
    while ws[f"B{row}"].value not in (None, ""):
        if ws[f"B{row}"].value == game_id:
            return row
        row += 1
    return None

def has_halftime_file(game_id: str) -> bool:
    folder = os.path.join(str(DATA_DIR), "raw", "pbp_live", str(game_id))
    pattern = os.path.join(folder, "pbp_first_half_*.json")
    return len(glob.glob(pattern)) > 0

def run_py(script_name: str, args_list):
    script_path = os.path.join(str(SCRIPTS_DIR), script_name)
    cmd = [sys.executable, script_path] + args_list
    r = subprocess.run(cmd)
    return r.returncode

def main():
    ap = argparse.ArgumentParser(description="Backfill: pull halftime PBP + generate report + log prediction for all games on a date, skipping any gameIDs already in workbook.")
    ap.add_argument("--date", required=True, help="YYYY-MM-DD")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing predictions")
    args = ap.parse_args()

    date = args.date.strip()
    selected_path = os.path.join(str(DATA_DIR), "processed", "selected_games", f"selected_games_{date}.json")
    baseline_path = os.path.join(str(DATA_DIR), "processed", "baselines", f"last4_{date}.json")

    if not os.path.exists(selected_path):
        print(f"Missing selected_games file: {selected_path}")
        return 2
    if not os.path.exists(baseline_path):
        print(f"Missing baseline manifest: {baseline_path}")
        return 2

    games = load_json(selected_path)
    if not isinstance(games, list):
        print(f"selected_games is not a list: {selected_path}")
        return 2

    wb = load_workbook(str(NCAAM_RESULTS_XLSX))
    ws = wb["Game_Log"]

    total = len(games)
    skipped_existing = 0
    attempted = 0
    ok = 0
    failed = 0

    for i, g in enumerate(games, start=1):
        gid = str((g or {}).get("gameID") or "").strip()
        if not gid:
            continue

        # Skip if already logged in workbook (unless overwrite)
        if find_row_by_game_id(ws, gid) is not None and not args.overwrite:
            skipped_existing += 1
            print(f"[{i}/{total}] SKIP (already logged) gameID={gid}")
            continue

        attempted += 1
        print(f"\n[{i}/{total}] RUN gameID={gid}")

        # 1) Ensure halftime PBP exists; pull if missing
        if not has_halftime_file(gid):
            rc = run_py("step4_pull_halftime_pbp_v2.py", [gid])
            if rc != 0:
                failed += 1
                print(f" FAIL pull_halftime rc={rc} gameID={gid}")
                continue

        # 2) Generate report
        rc = run_py(
            "step4b_feature_report_from_file_v5_test.py",
            [gid, "--baseline-manifest", baseline_path, "--selected-games", selected_path],
        )
        if rc != 0:
            failed += 1
            print(f" FAIL report rc={rc} gameID={gid}")
            continue

        # 3) Log prediction (overwrite if exists to update with new models)
        rc = run_py("log_prediction_to_results_v1.py", [gid, "--overwrite"])
        if rc != 0:
            failed += 1
            print(f" FAIL log rc={rc} gameID={gid}")
            continue

        ok += 1
        print(f" OK logged gameID={gid}")

    print("\nDONE")
    print(f"Total games in selected slate: {total}")
    print(f"Skipped (already logged): {skipped_existing}")
    print(f"Attempted new: {attempted}")
    print(f"OK: {ok}")
    print(f"Failed: {failed}")
    print(f"Workbook: {str(NCAAM_RESULTS_XLSX)}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())