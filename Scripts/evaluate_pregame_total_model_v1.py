import argparse
import csv
import math

from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error

from model_feature_utils import load_last4_pbp_priors, load_team_stats, resolve_team_stats


WORKBOOK_PATH = 'logs/NCAAM Results.xlsx'
SHEET_NAME = 'Game_Log'
CANONICAL_LINES_PATH = 'data/processed/market_lines/canonical_lines.csv'
DATA_ROOT = 'data'


PBP_PRIOR_KEYS = [
    'last4_three_rate',
    'last4_paint_share',
    'last4_ft_rate',
    'last4_turnover_rate',
    'last4_orb_rate',
    'last4_possessions_per_team_1h',
    'last4_pbp_coverage_count',
]


def safe_float(value):
    try:
        return float(value)
    except Exception:
        return None


def build_rows():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c) if c is not None else '' for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        data.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return data


def load_lines():
    out = {}
    with open(CANONICAL_LINES_PATH, 'r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            gid = str(row.get('game_id') or '').strip()
            if not gid:
                continue
            out[gid] = safe_float(row.get('total_game'))
    return out


def get_team_priors(last4_priors, team_name):
    team_data = (last4_priors or {}).get(str(team_name or '').strip(), {})
    return {key: float(team_data.get(key, 0.0) or 0.0) for key in PBP_PRIOR_KEYS}


def expected_total_from_team_stats(home_avg_scored, home_avg_allowed, away_avg_scored, away_avg_allowed):
    expected_home_total = (home_avg_scored + away_avg_allowed) / 2.0
    expected_away_total = (away_avg_scored + home_avg_allowed) / 2.0
    return expected_home_total + expected_away_total


def build_feature_row(record, team_stats_by_date, last4_priors_by_date):
    date_str = str(record.get('Date') or '').split(' ')[0].strip()
    home_team = str(record.get('Home') or '').strip()
    away_team = str(record.get('Away') or '').strip()
    actual_total = safe_float(record.get('ActualTotal'))
    game_id = str(record.get('GameID') or '').strip()

    if not date_str or not home_team or not away_team or actual_total is None or not game_id:
        return None

    stats = team_stats_by_date.get(date_str, {})
    home_avg_scored, home_avg_allowed = resolve_team_stats(stats, home_team)
    away_avg_scored, away_avg_allowed = resolve_team_stats(stats, away_team)
    expected_total = expected_total_from_team_stats(
        home_avg_scored,
        home_avg_allowed,
        away_avg_scored,
        away_avg_allowed,
    )

    priors = last4_priors_by_date.get(date_str, {})
    home_prior = get_team_priors(priors, home_team)
    away_prior = get_team_priors(priors, away_team)

    blended_possessions_1h = (home_prior['last4_possessions_per_team_1h'] + away_prior['last4_possessions_per_team_1h']) / 2.0
    blended_three_rate = (home_prior['last4_three_rate'] + away_prior['last4_three_rate']) / 2.0
    blended_paint_share = (home_prior['last4_paint_share'] + away_prior['last4_paint_share']) / 2.0
    blended_ft_rate = (home_prior['last4_ft_rate'] + away_prior['last4_ft_rate']) / 2.0
    blended_turnover_rate = (home_prior['last4_turnover_rate'] + away_prior['last4_turnover_rate']) / 2.0
    blended_orb_rate = (home_prior['last4_orb_rate'] + away_prior['last4_orb_rate']) / 2.0
    blended_coverage = (home_prior['last4_pbp_coverage_count'] + away_prior['last4_pbp_coverage_count']) / 2.0

    features = {
        'home_avg_scored': home_avg_scored,
        'home_avg_allowed': home_avg_allowed,
        'away_avg_scored': away_avg_scored,
        'away_avg_allowed': away_avg_allowed,
        'expected_total': expected_total,
        'home_offense_diff': home_avg_scored - away_avg_allowed,
        'away_offense_diff': away_avg_scored - home_avg_allowed,
        'blended_possessions_1h': blended_possessions_1h,
        'blended_possessions_full': blended_possessions_1h * 2.0,
        'blended_three_rate': blended_three_rate,
        'blended_paint_share': blended_paint_share,
        'blended_ft_rate': blended_ft_rate,
        'blended_turnover_rate': blended_turnover_rate,
        'blended_orb_rate': blended_orb_rate,
        'blended_pbp_coverage': blended_coverage,
        'home_last4_three_rate': home_prior['last4_three_rate'],
        'away_last4_three_rate': away_prior['last4_three_rate'],
        'home_last4_ft_rate': home_prior['last4_ft_rate'],
        'away_last4_ft_rate': away_prior['last4_ft_rate'],
        'home_last4_turnover_rate': home_prior['last4_turnover_rate'],
        'away_last4_turnover_rate': away_prior['last4_turnover_rate'],
        'home_last4_orb_rate': home_prior['last4_orb_rate'],
        'away_last4_orb_rate': away_prior['last4_orb_rate'],
        'home_last4_possessions_1h': home_prior['last4_possessions_per_team_1h'],
        'away_last4_possessions_1h': away_prior['last4_possessions_per_team_1h'],
        'three_rate_gap': home_prior['last4_three_rate'] - away_prior['last4_three_rate'],
        'ft_rate_gap': home_prior['last4_ft_rate'] - away_prior['last4_ft_rate'],
        'turnover_rate_gap': home_prior['last4_turnover_rate'] - away_prior['last4_turnover_rate'],
        'orb_rate_gap': home_prior['last4_orb_rate'] - away_prior['last4_orb_rate'],
    }

    return {
        'game_id': game_id,
        'date': date_str,
        'home': home_team,
        'away': away_team,
        'actual_total': actual_total,
        'features': features,
        'expected_total': expected_total,
    }


def side_hit(pred_total, closing_total, actual_total):
    if pred_total is None or closing_total is None or actual_total is None:
        return None
    if pred_total == closing_total:
        return None
    if actual_total == closing_total:
        return None  # push
    pred_side = 'OVER' if pred_total > closing_total else 'UNDER'
    actual_side = 'OVER' if actual_total > closing_total else 'UNDER'
    return int(pred_side == actual_side)


def payout_per_unit_risk(odds):
    if odds == 0:
        raise ValueError('Odds cannot be 0')
    if odds > 0:
        return odds / 100.0
    return 100.0 / abs(odds)


def summarize_side_hits(name, rows, pred_key):
    usable = []
    for row in rows:
        closing_total = row.get('closing_total')
        pred_total = row.get(pred_key)
        actual_total = row.get('actual_total')
        hit = side_hit(pred_total, closing_total, actual_total)
        if hit is None:
            continue
        usable.append({
            'hit': hit,
            'gap': abs(pred_total - closing_total),
        })

    if not usable:
        print(f'{name}: no usable rows')
        return

    hit_rate = 100.0 * sum(x['hit'] for x in usable) / len(usable)
    print(f'{name}: n={len(usable)} hit={hit_rate:.1f}%')
    for threshold in [0, 1, 2, 3, 4, 5, 6, 8, 10]:
        subset = [x for x in usable if x['gap'] >= threshold]
        if len(subset) < 10:
            continue
        subset_rate = 100.0 * sum(x['hit'] for x in subset) / len(subset)
        print(f'  gap>={threshold}: n={len(subset)} hit={subset_rate:.1f}%')


def summarize_wager_policy(name, rows, pred_key, odds, min_bets):
    per_win = payout_per_unit_risk(odds)

    graded = []
    for row in rows:
        closing_total = row.get('closing_total')
        pred_total = row.get(pred_key)
        actual_total = row.get('actual_total')
        hit = side_hit(pred_total, closing_total, actual_total)
        if hit is None:
            continue
        graded.append({
            'hit': hit,
            'gap': abs(pred_total - closing_total),
        })

    if not graded:
        print(f'{name}: no graded wagers')
        return

    print(f'{name} (odds {odds:+d}, risk 1u each)')
    best = None
    for threshold in [0, 1, 2, 3, 4, 5, 6, 8, 10]:
        wagers = [x for x in graded if x['gap'] >= threshold]
        n = len(wagers)
        if n < min_bets:
            continue
        wins = sum(x['hit'] for x in wagers)
        losses = n - wins
        units = wins * per_win - losses
        roi = (units / n) * 100.0
        hit_rate = (wins / n) * 100.0
        print(
            f'  gap>={threshold}: bets={n:4d} win%={hit_rate:5.1f}% '
            f'units={units:7.2f} roi={roi:6.2f}%'
        )
        candidate = (units, roi, hit_rate, n, threshold)
        if best is None or candidate > best:
            best = candidate

    if best is not None:
        units, roi, hit_rate, n, threshold = best
        print(
            f'  best_by_units: gap>={threshold} | bets={n} | '
            f'win%={hit_rate:.1f}% | units={units:.2f} | roi={roi:.2f}%'
        )


def evaluate_tiered_policy(name, rows, pred_key, odds, full_gap, half_gap):
    if half_gap > full_gap:
        raise ValueError('half-gap cannot be greater than full-gap')

    per_win = payout_per_unit_risk(odds)
    full_bets = []
    half_bets = []

    for row in rows:
        closing_total = row.get('closing_total')
        pred_total = row.get(pred_key)
        actual_total = row.get('actual_total')
        hit = side_hit(pred_total, closing_total, actual_total)
        if hit is None:
            continue
        gap = abs(pred_total - closing_total)
        if gap >= full_gap:
            full_bets.append(hit)
        elif gap >= half_gap:
            half_bets.append(hit)

    full_wins = sum(full_bets)
    full_losses = len(full_bets) - full_wins
    half_wins = sum(half_bets)
    half_losses = len(half_bets) - half_wins

    full_units = full_wins * per_win - full_losses
    # Half stake risks 0.5u and returns 0.5 * per-win payout.
    half_units = (half_wins * per_win - half_losses) * 0.5

    total_bets = len(full_bets) + len(half_bets)
    total_units = full_units + half_units
    total_risk = len(full_bets) * 1.0 + len(half_bets) * 0.5
    total_wins = full_wins + half_wins
    total_hit_rate = (100.0 * total_wins / total_bets) if total_bets else 0.0
    total_roi = (100.0 * total_units / total_risk) if total_risk > 0 else 0.0

    full_hit = (100.0 * full_wins / len(full_bets)) if full_bets else 0.0
    half_hit = (100.0 * half_wins / len(half_bets)) if half_bets else 0.0

    print(f'{name} tiered policy (FULL>= {full_gap}, HALF>= {half_gap}, PASS otherwise)')
    print(
        f'  FULL: bets={len(full_bets):4d} win%={full_hit:5.1f}% '
        f'units={full_units:7.2f} risk={len(full_bets):6.1f}'
    )
    print(
        f'  HALF: bets={len(half_bets):4d} win%={half_hit:5.1f}% '
        f'units={half_units:7.2f} risk={len(half_bets) * 0.5:6.1f}'
    )
    print(
        f'  ALL : bets={total_bets:4d} win%={total_hit_rate:5.1f}% '
        f'units={total_units:7.2f} risk={total_risk:6.1f} roi={total_roi:6.2f}%'
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--test-fraction', type=float, default=0.25, help='Fraction of unique dates reserved for holdout')
    parser.add_argument('--min-train-dates', type=int, default=20, help='Minimum training dates before evaluating holdout')
    parser.add_argument('--odds', type=int, default=-110, help='American odds for policy ROI simulation (default: -110)')
    parser.add_argument('--min-bets', type=int, default=20, help='Minimum number of wagers to report at a threshold')
    parser.add_argument('--full-gap', type=float, default=10.0, help='Gap threshold for FULL stake in tiered policy')
    parser.add_argument('--half-gap', type=float, default=6.0, help='Gap threshold for HALF stake in tiered policy')
    args = parser.parse_args()

    workbook_rows = build_rows()
    lines = load_lines()

    unique_dates = sorted({str(r.get('Date') or '').split(' ')[0].strip() for r in workbook_rows if r.get('Date') not in (None, '')})
    test_dates_count = max(1, int(math.ceil(len(unique_dates) * args.test_fraction)))
    if len(unique_dates) - test_dates_count < args.min_train_dates:
        test_dates_count = max(1, len(unique_dates) - args.min_train_dates)
    split_index = len(unique_dates) - test_dates_count
    train_dates = set(unique_dates[:split_index])
    test_dates = set(unique_dates[split_index:])

    print(f'Unique dates: {len(unique_dates)} | train={len(train_dates)} | test={len(test_dates)}')
    print(f'Test window: {min(test_dates) if test_dates else "n/a"} -> {max(test_dates) if test_dates else "n/a"}')

    all_needed_dates = sorted({str(r.get('Date') or '').split(' ')[0].strip() for r in workbook_rows if r.get('Date') not in (None, '')})
    team_stats_by_date = {date_str: load_team_stats(date_str) for date_str in all_needed_dates}
    last4_priors_by_date = {date_str: load_last4_pbp_priors(date_str, data_root=DATA_ROOT) for date_str in all_needed_dates}

    feature_rows = []
    for record in workbook_rows:
        built = build_feature_row(record, team_stats_by_date, last4_priors_by_date)
        if built is None:
            continue
        built['closing_total'] = lines.get(built['game_id'])
        feature_rows.append(built)

    feature_rows = [r for r in feature_rows if r.get('closing_total') is not None]
    print(f'Rows with actual totals and closing totals: {len(feature_rows)}')

    feature_names = list(feature_rows[0]['features'].keys()) if feature_rows else []
    train_rows = [r for r in feature_rows if r['date'] in train_dates]
    test_rows = [r for r in feature_rows if r['date'] in test_dates]

    print(f'Train rows: {len(train_rows)} | Test rows: {len(test_rows)}')
    if not train_rows or not test_rows:
        print('Insufficient rows after split.')
        return

    x_train = [[r['features'][name] for name in feature_names] for r in train_rows]
    y_train = [r['actual_total'] for r in train_rows]
    x_test = [[r['features'][name] for name in feature_names] for r in test_rows]
    y_test = [r['actual_total'] for r in test_rows]

    model = RandomForestRegressor(
        n_estimators=400,
        min_samples_leaf=3,
        random_state=42,
        n_jobs=-1,
    )
    model.fit(x_train, y_train)
    preds = model.predict(x_test)

    for row, pred in zip(test_rows, preds):
        row['model_pred_total'] = float(pred)
        row['baseline_pred_total'] = float(row['expected_total'])

    mae_model = mean_absolute_error(y_test, preds)
    mae_baseline = mean_absolute_error(y_test, [r['expected_total'] for r in test_rows])
    print(f'MAE baseline expected_total: {mae_baseline:.3f}')
    print(f'MAE RF pregame total model: {mae_model:.3f}')

    print('')
    summarize_side_hits('BASELINE_EXPECTED_TOTAL vs closing line', test_rows, 'baseline_pred_total')
    summarize_side_hits('RF_PREGAME_TOTAL_MODEL vs closing line', test_rows, 'model_pred_total')

    print('')
    summarize_wager_policy('BASELINE gap policy', test_rows, 'baseline_pred_total', args.odds, args.min_bets)
    summarize_wager_policy('RF model gap policy', test_rows, 'model_pred_total', args.odds, args.min_bets)

    print('')
    evaluate_tiered_policy('BASELINE', test_rows, 'baseline_pred_total', args.odds, args.full_gap, args.half_gap)
    evaluate_tiered_policy('RF model', test_rows, 'model_pred_total', args.odds, args.full_gap, args.half_gap)

    print('')
    importances = sorted(zip(feature_names, model.feature_importances_), key=lambda x: x[1], reverse=True)
    print('Top feature importances')
    for name, value in importances[:15]:
        print(f'  {name:30s} {value:.4f}')


if __name__ == '__main__':
    main()
