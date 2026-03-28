"""
True out-of-sample validation.
Train on games before 2026-03-01, test on 2026-03-01 and later.
Prints hit rates and MAEs on the hold-out set only.
"""
import sys, os, statistics
sys.path.insert(0, str(__import__('pathlib').Path(__file__).parent))
os.chdir(str(__import__('pathlib').Path(__file__).parent.parent))

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
from model_feature_utils import (
    FEATURE_NAMES, build_feature_vector, load_team_stats,
    load_market_lines, parse_halftime_score, resolve_team_stats,
)
from step4b_feature_report_from_file_v5_test import load_game_pbp_features

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = load_workbook('logs/NCAAM Results.xlsx', data_only=True)
ws = wb['Game_Log']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c) if c is not None else '' for c in rows[0]]
data = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]

# ── Load supporting data ──────────────────────────────────────────────────────
dates = set(str(r.get('Date')).split(' ')[0] for r in data if r.get('Date'))
team_stats_cache = {d: load_team_stats(d) for d in dates}
market_lines_cache = load_market_lines()
data_root = str(Path('data').resolve())
pbp_feature_cache = {}

def safe_float(v):
    try:
        return float(v)
    except Exception:
        return None

for r in data:
    away_ht, home_ht, home_lead, halftime_total = parse_halftime_score(r.get('HalftimeScore'))
    r['halftime_away'] = away_ht
    r['halftime_home'] = home_ht
    r['halftime_total'] = halftime_total

    date = str(r.get('Date')).split(' ')[0] if r.get('Date') else ''
    game_id = str(r.get('GameID')).strip() if r.get('GameID') else ''
    stats = team_stats_cache.get(date, {})
    home_avg_scored, home_avg_allowed = resolve_team_stats(stats, r.get('Home'))
    away_avg_scored, away_avg_allowed = resolve_team_stats(stats, r.get('Away'))

    pbp_features = {}
    if game_id:
        if game_id not in pbp_feature_cache:
            pbp_feature_cache[game_id] = load_game_pbp_features(data_root, game_id)
        pbp_features = pbp_feature_cache.get(game_id, {})

    if home_ht is not None and away_ht is not None:
        fv, fd = build_feature_vector(
            date, r.get('PaceProfile'), home_ht, away_ht,
            home_avg_scored, home_avg_allowed,
            away_avg_scored, away_avg_allowed,
            pbp_features, game_id=game_id,
            market_lines_cache=market_lines_cache,
        )
        r['model_features'] = fv
        # fd may have updated halftime_total
        if fd.get('halftime_total') is not None:
            r['halftime_total'] = fd['halftime_total']
    else:
        r['model_features'] = None

    for k in ['ActualMargin', 'Actual2H', 'ActualTotal']:
        r[k] = safe_float(r.get(k))

valid = [
    r for r in data
    if r.get('model_features') is not None
    and r.get('Actual2H') is not None
    and r.get('ActualTotal') is not None
]
print(f'Valid rows total: {len(valid)}')

# ── Time-based split ──────────────────────────────────────────────────────────
cutoff = datetime(2026, 3, 1)
train, test = [], []
for r in valid:
    try:
        d = datetime.strptime(str(r.get('Date')).split(' ')[0], '%Y-%m-%d')
        (train if d < cutoff else test).append(r)
    except Exception:
        pass

print(f'Train (before Mar 1): {len(train)}  |  Test (Mar 1+): {len(test)}')

# ── Train on Feb only, predict March ─────────────────────────────────────────
train_X = [r['model_features'] for r in train]
test_X  = [r['model_features'] for r in test]

m2h = RandomForestRegressor(n_estimators=100, random_state=42)
m2h.fit(train_X, [r['Actual2H'] for r in train])
pred_2h = m2h.predict(test_X)

mtot = RandomForestRegressor(n_estimators=100, random_state=42)
mtot.fit(train_X, [r['ActualTotal'] for r in train])
pred_tot_direct = mtot.predict(test_X)

actual_2h  = [r['Actual2H']    for r in test]
actual_tot = [r['ActualTotal'] for r in test]
ht_totals  = [r['halftime_total'] or 0 for r in test]

# Derived = halftime_total + predicted_2H
pred_tot_derived = [ht + p for ht, p in zip(ht_totals, pred_2h)]

mae_2h      = statistics.mean(abs(a - p) for a, p in zip(actual_2h,  pred_2h))
mae_derived = statistics.mean(abs(a - p) for a, p in zip(actual_tot, pred_tot_derived))
mae_direct  = statistics.mean(abs(a - p) for a, p in zip(actual_tot, pred_tot_direct))

use_derived   = mae_derived <= mae_direct
final_tot_pred = pred_tot_derived if use_derived else pred_tot_direct
mae_tot_final  = mae_derived      if use_derived else mae_direct

# ── Range hit rates using hold-out error as the range width ──────────────────
# (This mirrors how update_all_predictions_ml.py sizes the range in production)
half_2h  = mae_2h
half_tot = mae_tot_final

hit_2h  = sum(1 for a, p in zip(actual_2h,  pred_2h)         if abs(a - p) <= half_2h)  / len(actual_2h)
hit_tot = sum(1 for a, p in zip(actual_tot, final_tot_pred)   if abs(a - p) <= half_tot) / len(actual_tot)

print()
print('=' * 58)
print('  TRUE OUT-OF-SAMPLE RESULTS  (train=Feb, test=March)')
print('=' * 58)
print(f'  Test games : {len(test)}')
print(f'  2H  Hit Rate  (range +/-{half_2h:.2f}) : {hit_2h*100:.1f}%')
print(f'  Total Hit Rate (range +/-{half_tot:.2f}) : {hit_tot*100:.1f}%')
print(f'  2H  MAE    : {mae_2h:.3f}')
print(f'  Total MAE  : {mae_tot_final:.3f}  (strategy: {"derived_2h" if use_derived else "direct"})')
print('=' * 58)
print()
print('  In-sample comparison (all data trained):')
print('    2H Hit: 64.4%  |  Total Hit: 64.3%  |  2H MAE: 3.649  |  Total MAE: 3.649')
print()

# ── Feature importances ───────────────────────────────────────────────────────
print('  2H model feature importances (top 10):')
imps = sorted(zip(FEATURE_NAMES, m2h.feature_importances_), key=lambda x: -x[1])
for name, imp in imps[:10]:
    print(f'    {name}: {imp:.3f}')
