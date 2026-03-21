from pathlib import Path
from openpyxl import load_workbook
import subprocess
import sys
from paths import NCAAM_RESULTS_XLSX

WORKBOOK = str(NCAAM_RESULTS_XLSX)
SHEET_NAME = "Game_Log"
GAME_ID_COL = 2
ACTUAL_WINNER_COL = 12


def is_blank(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def main():
    script_dir = Path(__file__).resolve().parent
    updater = script_dir / "update_results_postgame_v1.py"

    if not Path(WORKBOOK).exists():
        print("Workbook not found:", WORKBOOK)
        sys.exit(1)

    if not updater.exists():
        print("Missing script:", updater)
        sys.exit(1)

    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET_NAME]

    game_ids = []
    seen = set()

    for row in range(2, ws.max_row + 1):

        gid = ws.cell(row=row, column=GAME_ID_COL).value
        actual = ws.cell(row=row, column=ACTUAL_WINNER_COL).value

        if gid in (None, ""):
            continue

        gid = str(gid).strip()

        if gid == "":
            continue

        if not is_blank(actual):
            continue

        if gid in seen:
            continue

        seen.add(gid)
        game_ids.append(gid)

    print("Games needing updates:", len(game_ids))

    if not game_ids:
        print("Nothing to update.")
        return

    failures = []

    for gid in game_ids:

        print("Updating", gid)

        result = subprocess.run([
            sys.executable,
            str(updater),
            gid
        ])

        if result.returncode != 0:
            failures.append(gid)

    print("")
    print("Finished. Attempted:", len(game_ids))

    if failures:
        print("Failures:", failures)
    else:
        print("All games updated successfully.")


if __name__ == "__main__":
    main()