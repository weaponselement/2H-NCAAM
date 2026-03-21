from openpyxl import load_workbook
from paths import NCAAM_RESULTS_XLSX

wb = load_workbook(str(NCAAM_RESULTS_XLSX), data_only=True)
ws = wb["Game_Log"]

missing = []
for r in range(2, ws.max_row + 1):
    gid = ws[f"B{r}"].value
    if gid in (None, ""):
        continue
    actual = ws[f"L{r}"].value
    if actual in (None, ""):
        missing.append((r, str(gid)))

print("missing_count=", len(missing))
print("last_5_missing=", missing[-5:])