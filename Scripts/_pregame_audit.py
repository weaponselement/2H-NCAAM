"""Pregame model data integrity audit — ignores all 2H/halftime columns."""
import sys
from pathlib import Path
from openpyxl import load_workbook
import csv

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# ── 1. Workbook: only the 5 columns pregame model needs ──────────────────────
wb_path = PROJECT_ROOT / "logs" / "NCAAM Results.xlsx"
wb = load_workbook(str(wb_path), data_only=True)
ws = wb["Game_Log"]
rows = list(ws.iter_rows(values_only=True))
headers = [str(c) if c is not None else "" for c in rows[0]]

needed = {"Date", "Home", "Away", "GameID", "ActualTotal"}
missing_headers = needed - set(headers)
if missing_headers:
    print(f"FAIL: workbook missing columns: {missing_headers}")
    sys.exit(1)

date_idx   = headers.index("Date")
home_idx   = headers.index("Home")
away_idx   = headers.index("Away")
gid_idx    = headers.index("GameID")
total_idx  = headers.index("ActualTotal")

total_rows = 0
has_total  = 0
no_total   = 0
bad_total  = 0
dup_check  = {}
dup_ids    = []

for r in rows[1:]:
    if not any(v is not None for v in r):
        continue
    total_rows += 1
    gid = str(r[gid_idx] or "").strip()
    if gid:
        dup_check[gid] = dup_check.get(gid, 0) + 1

    actual = r[total_idx]
    if actual is None or actual == "":
        no_total += 1
    else:
        try:
            v = float(actual)
            if v <= 0:
                bad_total += 1
            else:
                has_total += 1
        except Exception:
            bad_total += 1

dup_ids = [g for g, n in dup_check.items() if n > 1]
print(f"Workbook rows: {total_rows}")
print(f"  ActualTotal present+valid : {has_total}")
print(f"  ActualTotal blank/missing : {no_total}")
print(f"  ActualTotal bad value     : {bad_total}")
print(f"  Duplicate GameIDs         : {len(dup_ids)}")
if dup_ids:
    print(f"  Sample dupes: {dup_ids[:5]}")
wb_ok = (bad_total == 0 and len(dup_ids) == 0)
print(f"  Workbook check: {'PASS' if wb_ok else 'FAIL'}")

# ── 2. Canonical lines CSV ────────────────────────────────────────────────────
lines_path = PROJECT_ROOT / "data" / "processed" / "market_lines" / "canonical_lines.csv"
if not lines_path.exists():
    print(f"\nFAIL: canonical_lines.csv not found at {lines_path}")
    sys.exit(1)

with open(lines_path, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    line_rows = list(reader)
    line_fields = reader.fieldnames or []

needed_line_fields = {"game_id", "date", "away_seo", "home_seo", "total_game"}
missing_line_fields = needed_line_fields - set(line_fields)
total_lines = len(line_rows)
lines_with_total = sum(1 for r in line_rows if r.get("total_game") not in (None, "", "nan"))

print(f"\nCanonical lines: {lines_path.name}")
print(f"  Rows               : {total_lines}")
print(f"  Has total_game     : {lines_with_total}")
print(f"  Missing req fields : {missing_line_fields or 'none'}")
lines_ok = (total_lines > 0 and not missing_line_fields and lines_with_total > 0)
print(f"  Lines check: {'PASS' if lines_ok else 'FAIL'}")

# ── 3. Baselines directory ────────────────────────────────────────────────────
baselines_dir = PROJECT_ROOT / "data" / "processed" / "baselines"
if not baselines_dir.exists():
    print(f"\nFAIL: baselines dir not found: {baselines_dir}")
    sys.exit(1)

lastN_files = list(baselines_dir.glob("lastN_*.json"))
last4_files = list(baselines_dir.glob("last4_*.json"))
print(f"\nBaselines dir: {baselines_dir}")
print(f"  lastN_*.json files : {len(lastN_files)}")
print(f"  last4_*.json files : {len(last4_files)}")
baselines_ok = (len(lastN_files) > 0 or len(last4_files) > 0)
print(f"  Baselines check: {'PASS' if baselines_ok else 'FAIL'}")

# ── 4. Pregame model cache ────────────────────────────────────────────────────
cache_dir = PROJECT_ROOT / "models" / "pregame_total_cache"
pkl_files  = list(cache_dir.glob("pregame_total_rf_w*.pkl")) if cache_dir.exists() else []
meta_files = list(cache_dir.glob("pregame_total_rf_w*.meta.json")) if cache_dir.exists() else []
print(f"\nPregame cache: {cache_dir}")
print(f"  .pkl files  : {[p.name for p in pkl_files]}")
print(f"  .meta files : {[m.name for m in meta_files]}")
cache_ok = len(pkl_files) > 0 and len(meta_files) == len(pkl_files)
print(f"  Cache check: {'PASS' if cache_ok else 'WARN (no cache yet — will build on first run)'}")

# ── 5. Summary ────────────────────────────────────────────────────────────────
print()
all_ok = wb_ok and lines_ok and baselines_ok
print("=" * 50)
print(f"PREGAME AUDIT: {'PASS' if all_ok else 'FAIL'}")
print("=" * 50)
if not all_ok:
    sys.exit(1)
