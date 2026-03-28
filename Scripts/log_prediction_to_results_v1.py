import argparse
import glob
import json
import os
from openpyxl import load_workbook
from paths import DATA_DIR, NCAAM_RESULTS_XLSX

RESULTS_XLSX = str(NCAAM_RESULTS_XLSX)
DATA_ROOT = str(DATA_DIR)


def find_row_by_game_id(ws, game_id: str):
    """Return row number where column B == game_id, else None."""
    row = 2
    while ws[f"B{row}"].value not in (None, ""):
        if ws[f"B{row}"].value == game_id:
            return row
        row += 1
    return None


def latest_report_path(game_id: str):
    pattern = os.path.join(
        DATA_ROOT,
        "processed",
        "reports",
        f"feature_report_v5_test_{game_id}_*.json"
    )
    matches = glob.glob(pattern)

    if not matches:
        pattern = os.path.join(
            DATA_ROOT,
            "processed",
            "reports",
            f"feature_report_v4_{game_id}_*.json"
        )
        matches = glob.glob(pattern)

    if not matches:
        return None

    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return matches[0]


def load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def first_empty_row(ws):
    row = 2
    while ws[f"B{row}"].value not in (None, ""):
        row += 1
    return row


def build_header_index(ws):
    headers = []
    idx = {}
    for cell in ws[1]:
        value = '' if cell.value is None else str(cell.value)
        headers.append(value)
        if value:
            idx[value] = cell.column
    return headers, idx


def ensure_header(ws, headers, idx, header_name):
    if header_name in idx:
        return idx[header_name]
    column = len(headers) + 1
    ws.cell(row=1, column=column, value=header_name)
    headers.append(header_name)
    idx[header_name] = column
    return column


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("game_id")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing entry if it exists")
    args = parser.parse_args()

    report_path = latest_report_path(args.game_id)

    if not report_path:
        print(f"No report found for game {args.game_id}")
        return

    report = load_json(report_path)

    halftime = report.get("halftime_score", {})
    projection = report.get("projection", {})
    game_state = report.get("game_state", {})

    teams = report.get("teams", {})
    away = (teams.get("away") or {}).get("seo", "")
    home = (teams.get("home") or {}).get("seo", "")

    away_ht = halftime.get("away")
    home_ht = halftime.get("home")

    halftime_score = ""
    if away_ht is not None and home_ht is not None:
        halftime_score = f"{away_ht}-{home_ht}"

    pred_margin = projection.get("winner_margin_range", "")
    pred_2h = (projection.get("second_half_points_projection") or {}).get("range", "")
    pred_total = (projection.get("full_game_total_projection") or {}).get("range", "")

    
    wb = load_workbook(RESULTS_XLSX)
    ws = wb["Game_Log"]
    headers, idx = build_header_index(ws)

    base_columns = [
        'Date',
        'GameID',
        'Away',
        'Home',
        'HalftimeScore',
        'PaceProfile',
        'PredWinner',
        'PredMarginRange',
        'Pred2HRange',
        'PredTotalRange',
        'Confidence',
        'TriggerDecision',
        'TriggerName',
        'TriggerHistHit',
        'TriggerHistN',
        'WageredFlag',
        'TriggerVersion',
        'StakeTier',
    ]
    for header_name in base_columns:
        ensure_header(ws, headers, idx, header_name)

    existing_row = find_row_by_game_id(ws, args.game_id)
    if existing_row is not None:
        if not args.overwrite:
            print(f"SKIP: game {args.game_id} already exists in workbook at row {existing_row}")
            print(f"Workbook: {RESULTS_XLSX}")
            return
        else:
            row = existing_row
            print(f"OVERWRITE: updating existing row {row} for game {args.game_id}")
    else:
        row = first_empty_row(ws)

    ws.cell(row=row, column=idx['Date'], value=report.get("run_date", ""))
    ws.cell(row=row, column=idx['GameID'], value=args.game_id)
    ws.cell(row=row, column=idx['Away'], value=away)
    ws.cell(row=row, column=idx['Home'], value=home)
    ws.cell(row=row, column=idx['HalftimeScore'], value=halftime_score)
    ws.cell(row=row, column=idx['PaceProfile'], value=game_state.get("pace_profile", ""))
    ws.cell(row=row, column=idx['PredWinner'], value=projection.get("winner_projection", ""))
    ws.cell(row=row, column=idx['PredMarginRange'], value=pred_margin)
    ws.cell(row=row, column=idx['Pred2HRange'], value=pred_2h)
    ws.cell(row=row, column=idx['PredTotalRange'], value=pred_total)
    ws.cell(row=row, column=idx['Confidence'], value=projection.get("confidence", ""))
    ws.cell(row=row, column=idx['TriggerDecision'], value='')
    ws.cell(row=row, column=idx['TriggerName'], value='')
    ws.cell(row=row, column=idx['TriggerHistHit'], value='')
    ws.cell(row=row, column=idx['TriggerHistN'], value='')
    ws.cell(row=row, column=idx['WageredFlag'], value='N')
    ws.cell(row=row, column=idx['TriggerVersion'], value='')
    ws.cell(row=row, column=idx['StakeTier'], value='PASS')

    wb.save(RESULTS_XLSX)

    print(f"Logged prediction for game {args.game_id} into row {row}")
    print(f"Workbook: {RESULTS_XLSX}")


if __name__ == "__main__":
    main()