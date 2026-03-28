from datetime import datetime
from pathlib import Path
import json
import pickle

from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error

from model_feature_utils import FEATURE_NAMES, build_feature_vector, load_last4_pbp_priors, load_team_stats, load_market_lines, load_neutral_court_games, load_rest_context, parse_halftime_score, resolve_team_stats
from step4b_feature_report_from_file_v5_test import load_game_pbp_features

path = Path('logs/NCAAM Results.xlsx')
if not path.exists():
    raise FileNotFoundError(path)

wb = load_workbook(path, data_only=True)
if 'Game_Log' not in wb.sheetnames:
    raise ValueError('Game_Log not in workbook')

ws = wb['Game_Log']
rows = list(ws.iter_rows(values_only=True))
headers = [str(c) if c is not None else '' for c in rows[0]]
data = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]

# Load team stats for all dates
dates = set(str(r.get('Date')).split(' ')[0] for r in data if r.get('Date'))
team_stats_cache = {}
for d in dates:
    team_stats_cache[d] = load_team_stats(d)

pbp_feature_cache = {}
data_root = str(Path(__file__).resolve().parent.parent / 'data')

print('loaded team stats for', len(dates), 'dates')
print('loaded', len(data), 'rows')

# Load market lines
market_lines_cache = load_market_lines()
print('loaded market lines for', len(market_lines_cache), 'games')

# Load rest context
rest_context = load_rest_context()
print('loaded rest context for', len(rest_context), 'teams')

# Load neutral court game IDs
neutral_court_games = load_neutral_court_games()
print('loaded neutral court games:', len(neutral_court_games))

# Load date-keyed last4 historical PBP priors
last4_pbp_priors_by_date = {}
for d in dates:
    last4_pbp_priors_by_date[d] = load_last4_pbp_priors(d, data_root=data_root)
print('loaded last4 historical pbp priors for', len(last4_pbp_priors_by_date), 'dates')


def safe_float(value):
    try:
        return float(value)
    except Exception:
        return None

# Add features
for r in data:
    away_ht, home_ht, home_lead, halftime_total = parse_halftime_score(r.get('HalftimeScore'))
    r['halftime_away'] = away_ht
    r['halftime_home'] = home_ht
    r['home_lead'] = home_lead
    r['halftime_total'] = halftime_total
    date = str(r.get('Date')).split(' ')[0] if r.get('Date') else ''
    home_team = r.get('Home')
    away_team = r.get('Away')
    game_id = r.get('GameID')
    stats = team_stats_cache.get(date, {})
    home_avg_scored, home_avg_allowed = resolve_team_stats(stats, home_team)
    away_avg_scored, away_avg_allowed = resolve_team_stats(stats, away_team)
    r['home_avg_scored'] = home_avg_scored
    r['home_avg_allowed'] = home_avg_allowed
    r['away_avg_scored'] = away_avg_scored
    r['away_avg_allowed'] = away_avg_allowed
    pbp_features = {}
    if game_id not in (None, ''):
        game_id_key = str(game_id).strip()
        if game_id_key not in pbp_feature_cache:
            pbp_feature_cache[game_id_key] = load_game_pbp_features(data_root, game_id_key)
        pbp_features = pbp_feature_cache.get(game_id_key, {})
    r['pbp_features'] = pbp_features
    if home_ht is not None and away_ht is not None:
        feature_vector, feature_dict = build_feature_vector(
            date,
            r.get('PaceProfile'),
            home_ht,
            away_ht,
            home_avg_scored,
            home_avg_allowed,
            away_avg_scored,
            away_avg_allowed,
            pbp_features,
            game_id=game_id,
            market_lines_cache=market_lines_cache,
            home_team_seo=home_team,
            away_team_seo=away_team,
            rest_context=rest_context,
            neutral_court_games=neutral_court_games,
            last4_pbp_priors=last4_pbp_priors_by_date.get(date, {}),
        )
        r.update(feature_dict)
        r['model_features'] = feature_vector
    else:
        r['model_features'] = None

    r['ActualMargin'] = safe_float(r.get('ActualMargin'))
    r['Actual2H'] = safe_float(r.get('Actual2H'))
    r['ActualTotal'] = safe_float(r.get('ActualTotal'))

# Filter valid rows
valid = [
    r for r in data
    if r.get('model_features') is not None
    and r.get('ActualMargin') is not None
    and r.get('Actual2H') is not None
    and r.get('ActualTotal') is not None
]
print('valid rows for tuning:', len(valid))

# Split by date: train on dates before 2026-03-01, test after
cutoff = datetime(2026, 3, 1)
train = []
test = []
for r in valid:
    try:
        d = datetime.strptime(str(r.get('Date')), '%Y-%m-%d')
        if d < cutoff:
            train.append(r)
        else:
            test.append(r)
    except:
        continue

print('train:', len(train), 'test:', len(test))

targets = ['ActualMargin', 'Actual2H', 'ActualTotal']
error_stats = {}
test_predictions = {}
direct_total_mae = None
derived_total_mae = None
for target in targets:
    print(f'\n--- {target} ---')
    train_X = [r['model_features'] for r in train]
    train_y = [r[target] for r in train]
    test_X = [r['model_features'] for r in test]
    test_y = [r[target] for r in test]

    if target == 'ActualMargin':
        baseline_mae = 7.13
    elif target in ['Actual2H', 'ActualTotal']:
        baseline_mae = 10.33

    model = RandomForestRegressor(n_estimators=100, random_state=42)
    model.fit(train_X, train_y)

    if test_X:
        pred_y = model.predict(test_X)
        mae = mean_absolute_error(test_y, pred_y)
        test_predictions[target] = pred_y
        print(f'test MAE: {mae:.2f}')
        print(f'baseline MAE: {baseline_mae:.2f}')
        print(f'improvement: {baseline_mae - mae:.2f}')
    else:
        # No holdout rows in this slice; keep conservative baseline MAE values.
        mae = baseline_mae
        test_predictions[target] = []
        print('test MAE: n/a (no holdout rows in split)')
        print(f'using baseline MAE fallback: {mae:.2f}')

    if target == 'ActualTotal':
        direct_total_mae = mae
    else:
        error_stats[target] = mae

    importances = model.feature_importances_
    for f, imp in zip(FEATURE_NAMES, importances):
        print(f'  {f}: {imp:.3f}')

if test and 'Actual2H' in test_predictions:
    derived_total_predictions = [r['halftime_total'] + pred_2h for r, pred_2h in zip(test, test_predictions['Actual2H'])]
    derived_total_mae = mean_absolute_error([r['ActualTotal'] for r in test], derived_total_predictions)
    print(f'\nDerived total from 2H MAE: {derived_total_mae:.2f}')
    print(f'Direct total model MAE: {direct_total_mae:.2f}')
else:
    derived_total_mae = direct_total_mae

total_strategy = 'derived_2h' if derived_total_mae is not None and direct_total_mae is not None and derived_total_mae <= direct_total_mae else 'direct_model'
chosen_total_mae = derived_total_mae if total_strategy == 'derived_2h' else direct_total_mae
error_stats['ActualTotal'] = chosen_total_mae
print(f'Chosen total strategy: {total_strategy} (MAE {float(chosen_total_mae):.2f})')

# Persist error stats for prediction range logic
error_path = Path('models') / 'model_error_stats.json'
with open(error_path, 'w', encoding='utf-8') as f:
    json.dump(error_stats, f)
print(f'Saved error stats to {error_path}')

strategy_path = Path('models') / 'model_strategy.json'
with open(strategy_path, 'w', encoding='utf-8') as f:
    json.dump({
        'total_prediction_strategy': total_strategy,
        'direct_total_mae': direct_total_mae,
        'derived_total_mae': derived_total_mae,
        'feature_names': FEATURE_NAMES,
    }, f, indent=2)
print(f'Saved strategy metadata to {strategy_path}')

# Train final models on all data
print('\n--- Final Models on All Data ---')
final_models = {}
for target in targets:
    print(f'Training final {target} model...')
    X_all = [r['model_features'] for r in valid]
    y_all = [r[target] for r in valid]

    final_model = RandomForestRegressor(n_estimators=100, random_state=42)
    final_model.fit(X_all, y_all)
    final_models[target] = final_model

    model_path = Path('models') / f'{target.lower()}_model.pkl'
    model_path.parent.mkdir(exist_ok=True)
    with open(model_path, 'wb') as f:
        pickle.dump(final_model, f)
    print(f'Saved {target} model to {model_path}')

print('All models trained and saved.')
